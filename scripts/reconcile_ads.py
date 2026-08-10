#!/usr/bin/env python3
"""
TikTok Ads Payment Reconciliation — อ่าน screenshot แล้วออก Excel

Usage:
    # แค่อ่านรูปแล้ว deduplicate:
    python reconcile_ads.py /path/to/flat-folder

    # ชนกับ reference file (CSV/Excel):
    python reconcile_ads.py /path/to/folder --reference ref.csv

    # กำหนด output + concurrency:
    python reconcile_ads.py /path/to/folder -o result.xlsx -c 15

Requirements:
    pip install anthropic pandas openpyxl

Expected input:  flat folder ของ screenshots (jpg/png) — ไม่มี subfolder ซ้อน
Expected output: reconciliation.xlsx (หรือตามที่ระบุ)

Reference file format (CSV/Excel) — ต้องมีคอลัมน์ใดคอลัมน์หนึ่ง:
    transaction_id  |  invoice_no  |  amount_thb + transaction_time
"""

from __future__ import annotations

import asyncio
import base64
import json
import sys
import argparse
import re
from pathlib import Path
from typing import Optional

import anthropic
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ─────────────────────────────────────────────
SUPPORTED_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
DEFAULT_CONCURRENCY = 10
OCR_MODEL = "claude-haiku-4-5-20251001"   # เร็ว + ถูก เหมาะกับ structured OCR

EXTRACT_PROMPT = """You are extracting payment transaction data from a screenshot.
The screenshot may be from:
  A) TikTok Ads Manager "Transactions" page — shows a table with multiple rows
  B) A Thai billing portal — shows one transaction with Thai text

Extract ALL visible transaction rows. Return a JSON array only, no explanation.

Each object must have these exact keys (use null if the field is not visible):
{
  "transaction_time": "YYYY-MM-DD HH:MM:SS",
  "transaction_type": "Bill payment",
  "account_name": "company or account name",
  "transaction_id": "full ID — if two parts separated by | or newline, include both joined with |",
  "card_last4": "4-digit suffix of the card",
  "invoice_no": "invoice / document reference (e.g. THTT202602881921 or 2PDDQM5RX2)",
  "status": "Success or ชำระแล้ว or as shown",
  "amount_thb": 19706.61
}

Rules:
- amount_thb: positive number, no currency symbol
- transaction_id: keep full string including | separator
- If no transactions found, return []
- Return ONLY valid JSON array"""


# ─────────────────────────────────────────────
# OCR
# ─────────────────────────────────────────────

async def _ocr_image(
    client: anthropic.AsyncAnthropic,
    image_path: Path,
    sem: asyncio.Semaphore,
    idx: int,
    total: int,
) -> list[dict]:
    async with sem:
        try:
            raw = image_path.read_bytes()
            b64 = base64.b64encode(raw).decode()
            ext = image_path.suffix.lower().lstrip(".")
            media_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "webp": "image/webp"}
            media_type = media_map.get(ext, "image/jpeg")

            resp = await client.messages.create(
                model=OCR_MODEL,
                max_tokens=2048,
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}},
                        {"type": "text", "text": EXTRACT_PROMPT},
                    ],
                }],
            )

            text = resp.content[0].text.strip()
            # strip markdown fences
            text = re.sub(r"^```[a-z]*\n?", "", text)
            text = re.sub(r"\n?```$", "", text)
            rows = json.loads(text.strip())

            for row in rows:
                row["source_file"] = image_path.name

            print(f"  [{idx}/{total}] {image_path.name} → {len(rows)} rows", flush=True)
            return rows

        except Exception as exc:
            print(f"  [{idx}/{total}] ⚠ ERROR {image_path.name}: {exc}", file=sys.stderr, flush=True)
            return [{"source_file": image_path.name, "_error": str(exc)}]


async def ocr_folder(folder: Path, concurrency: int) -> list[dict]:
    images = sorted([f for f in folder.iterdir() if f.suffix.lower() in SUPPORTED_EXTS])
    if not images:
        print(f"No images found in {folder}")
        return []

    print(f"\n📂 {folder.name}")
    print(f"   {len(images)} images | concurrency={concurrency} | model={OCR_MODEL}\n")

    client = anthropic.AsyncAnthropic()
    sem = asyncio.Semaphore(concurrency)
    tasks = [_ocr_image(client, img, sem, i + 1, len(images)) for i, img in enumerate(images)]

    results: list[dict] = []
    for coro in asyncio.as_completed(tasks):
        batch = await coro
        results.extend(batch)

    print(f"\n✅ Extracted {len(results)} raw rows from {len(images)} images")
    return results


# ─────────────────────────────────────────────
# Classification
# ─────────────────────────────────────────────

def _normalize_id(val) -> str:
    if val is None:
        return ""
    return str(val).strip().lower().replace(" ", "")


def classify(
    raw_rows: list[dict],
    reference_df: Optional[pd.DataFrame],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Returns (matched_df, unmatched_df, dup_df, error_df)

    - dup_df      : rows whose transaction_id appears in >1 source file
    - matched_df  : unique rows that ARE in reference (or all unique if no ref)
    - unmatched_df: unique rows NOT in reference (empty if no ref)
    - error_df    : rows where OCR failed
    """
    df = pd.DataFrame(raw_rows)

    # split errors
    if "_error" in df.columns:
        err_df = df[df["_error"].notna()].copy()
        df = df[df["_error"].isna()].drop(columns=["_error"])
    else:
        err_df = pd.DataFrame()

    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), err_df

    df["_txid_norm"] = df["transaction_id"].apply(_normalize_id)
    df["_inv_norm"] = df["invoice_no"].apply(_normalize_id)

    # ── Duplicate detection (same txid from multiple files) ──
    dup_mask = df.duplicated(subset=["_txid_norm"], keep=False) & (df["_txid_norm"] != "")
    dup_df = df[dup_mask].copy()
    unique_df = df[~dup_mask].copy()

    # ── Reference matching ──
    if reference_df is not None:
        ref_ids = set(reference_df.get("transaction_id", pd.Series(dtype=str)).apply(_normalize_id))
        ref_invs = set(reference_df.get("invoice_no", pd.Series(dtype=str)).apply(_normalize_id))
        ref_combined = ref_ids | ref_invs

        match_mask = unique_df["_txid_norm"].isin(ref_combined) | unique_df["_inv_norm"].isin(ref_combined)
        matched_df = unique_df[match_mask].copy()
        unmatched_df = unique_df[~match_mask].copy()
    else:
        matched_df = unique_df.copy()
        unmatched_df = pd.DataFrame()

    # drop helper columns
    for _df in [matched_df, unmatched_df, dup_df]:
        _df.drop(columns=["_txid_norm", "_inv_norm"], inplace=True, errors="ignore")

    return matched_df, unmatched_df, dup_df, err_df


# ─────────────────────────────────────────────
# Excel output
# ─────────────────────────────────────────────

_COL_ORDER = [
    "transaction_time", "amount_thb", "transaction_id", "invoice_no",
    "card_last4", "account_name", "transaction_type", "status", "source_file",
]

def _reorder(df: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in _COL_ORDER if c in df.columns] + [c for c in df.columns if c not in _COL_ORDER]
    return df[cols]


def _apply_style(ws, header_hex: str):
    fill = PatternFill(start_color=header_hex, end_color=header_hex, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 55)


def _write_df(ws, df: pd.DataFrame, header_hex: str):
    if df.empty:
        ws.append(["— ไม่มีรายการ —"])
        return
    df = _reorder(df)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    _apply_style(ws, header_hex)


def write_excel(
    matched: pd.DataFrame,
    unmatched: pd.DataFrame,
    dup: pd.DataFrame,
    errors: pd.DataFrame,
    output_path: Path,
    ref_provided: bool,
    folder_name: str,
):
    wb = Workbook()
    wb.remove(wb.active)

    total_extracted = len(matched) + len(unmatched) + len(dup)

    # ── Summary ──
    ws_s = wb.create_sheet("Summary")
    ws_s.column_dimensions["A"].width = 42
    ws_s.column_dimensions["B"].width = 20

    def row(a, b=""):
        ws_s.append([a, b])

    row("📊 Reconciliation Report")
    row(f"Folder: {folder_name}")
    row("")
    row("รายการทั้งหมดที่ OCR ได้", total_extracted)
    row("  Unique transactions", len(matched) + len(unmatched))
    row("  ⚠ Duplicates (TxID ซ้ำในหลายรูป)", len(dup))
    row("  🔴 Error / อ่านไม่ได้", len(errors))
    row("")

    if ref_provided:
        row("── Matching Result ──")
        row("  ✅ Matched", len(matched))
        row("  ❌ Unmatched (ไม่พบใน reference)", len(unmatched))
        row("  ⚠ Duplicate", len(dup))
    else:
        row("── ผลลัพธ์ (ไม่มี Reference) ──")
        row("  ✅ Unique transactions", len(matched))
        row("  ⚠ Duplicates", len(dup))
        row("")
        row("* หากต้องการ Matched/Unmatched ให้ใส่ --reference")

    ws_s["A1"].font = Font(bold=True, size=13)
    ws_s["A1"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws_s["A1"].font = Font(bold=True, size=13, color="FFFFFF")

    # ── Matched / All Unique ──
    label_matched = "✅ Matched" if ref_provided else "✅ All Unique"
    ws_m = wb.create_sheet(label_matched)
    _write_df(ws_m, matched, "217346")   # green

    # ── Unmatched ──
    if ref_provided:
        ws_u = wb.create_sheet("❌ Unmatched")
        _write_df(ws_u, unmatched, "C0392B")  # red

    # ── Duplicates ──
    ws_d = wb.create_sheet("⚠ Duplicates")
    _write_df(ws_d, dup, "D35400")  # orange

    # ── Errors ──
    if not errors.empty:
        ws_e = wb.create_sheet("🔴 Errors")
        for r in dataframe_to_rows(errors, index=False, header=True):
            ws_e.append(r)

    wb.save(output_path)
    print(f"\n📄 Saved → {output_path.resolve()}")


# ─────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────

async def main():
    parser = argparse.ArgumentParser(
        description="TikTok Ads Payment Reconciliation — reads screenshots → Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("folder", help="Flat folder of screenshot images (jpg/png)")
    parser.add_argument("--reference", "-r", metavar="FILE",
                        help="Reference CSV or Excel with known transactions (optional)")
    parser.add_argument("--output", "-o", default="reconciliation.xlsx",
                        help="Output Excel filename (default: reconciliation.xlsx)")
    parser.add_argument("--concurrency", "-c", type=int, default=DEFAULT_CONCURRENCY,
                        help=f"Max parallel OCR calls (default: {DEFAULT_CONCURRENCY})")
    parser.add_argument("--api-key", "-k", metavar="KEY",
                        help="Anthropic API key (หรือตั้ง env ANTHROPIC_API_KEY)")
    args = parser.parse_args()

    # Set API key
    import os
    if args.api_key:
        os.environ["ANTHROPIC_API_KEY"] = args.api_key
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("❌  ต้องตั้ง ANTHROPIC_API_KEY ก่อน:", file=sys.stderr)
        print("    export ANTHROPIC_API_KEY='sk-ant-...'", file=sys.stderr)
        print("    หรือใช้ flag: --api-key sk-ant-...", file=sys.stderr)
        sys.exit(1)

    folder = Path(args.folder).expanduser().resolve()
    if not folder.is_dir():
        print(f"Error: '{folder}' is not a directory", file=sys.stderr)
        sys.exit(1)

    # Load reference
    reference_df = None
    if args.reference:
        ref = Path(args.reference).expanduser()
        reference_df = pd.read_csv(ref) if ref.suffix.lower() == ".csv" else pd.read_excel(ref)
        print(f"📎 Reference loaded: {len(reference_df)} rows from {ref.name}")

    # OCR
    raw = await ocr_folder(folder, args.concurrency)
    if not raw:
        print("Nothing extracted — exiting.")
        sys.exit(0)

    # Classify
    matched, unmatched, dup, errors = classify(raw, reference_df)

    # Print quick summary
    print("\n─────────────────────────────")
    print(f"  Unique     : {len(matched) + len(unmatched)}")
    print(f"  Duplicates : {len(dup)}")
    if reference_df is not None:
        print(f"  Matched    : {len(matched)}")
        print(f"  Unmatched  : {len(unmatched)}")
    print(f"  Errors     : {len(errors)}")
    print("─────────────────────────────")

    # Write Excel
    output = Path(args.output).expanduser().resolve()
    write_excel(matched, unmatched, dup, errors, output, reference_df is not None, folder.name)


if __name__ == "__main__":
    asyncio.run(main())
