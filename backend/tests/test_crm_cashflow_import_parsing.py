"""Tests for the legacy 8-column import row parser.

Every imported row starts with an unassigned invoice/document status. Values
in the legacy "ใบเสร็จ" column and the current "ใบกำกับภาษี" column are
intentionally ignored; accounting staff classify the document during review.
"""
import csv
import unittest
from io import BytesIO, StringIO
from datetime import date
from decimal import Decimal
from unittest.mock import AsyncMock

from openpyxl import load_workbook

from app.models.crm_cashflow import CrmCashflowImportTemplate
from app.routers.crm_cashflow import (
    DEFAULT_IMPORT_SOURCE_NAME,
    IMPORT_HEADERS_10,
    ImportTemplateColumn,
    _normalize_import_row,
    _normalize_import_row_with_template,
    _prepare_import_rows,
    _validate_import_template_columns,
    download_custom_import_template,
    download_import_template,
)


class _Result:
    """Minimal stand-in for a SQLAlchemy ``Result`` (scalar queries only)."""

    def __init__(self, *, scalar=None):
        self._scalar = scalar

    def scalar_one_or_none(self):
        return self._scalar


class _Company:
    def __init__(self, id=5):
        self.id = id


class LegacyEightColumnImportTests(unittest.TestCase):
    def test_receipt_present_starts_unassigned(self):
        row = ["2025-02-27", "รายจ่าย", "อุปกรณ์คอมพิวเตอร์", "-15000",
               "ซื้อเครื่องปริ้นเตอร์ใหม่", "1", "1", "ฝ่ายจัดซื้อ"]

        normalized = _normalize_import_row(row, legacy_eight_columns=True)

        self.assertIsNone(normalized["invoice"])

    def test_receipt_absent_maps_to_none(self):
        row = ["2025-02-27", "รายรับ", "ยอดขายออนไลน์", "15000",
               "รายได้จากการขายสินค้าออนไลน์", "0", "0", "ฝ่ายขาย"]

        normalized = _normalize_import_row(row, legacy_eight_columns=True)

        self.assertIsNone(normalized["invoice"])

    def test_receipt_column_is_ignored_even_when_malformed(self):
        row = ["2025-02-27", "รายจ่าย", "ค่าเช่า", "-12000",
               "ค่าเช่าสำนักงานประจำเดือน", "ไม่ทราบ", "0", "ฝ่ายบัญชี"]

        normalized = _normalize_import_row(row, legacy_eight_columns=True)

        self.assertIsNone(normalized["invoice"])

    def test_current_ten_column_invoice_value_is_ignored(self):
        row = ["2025-02-27", "รายจ่าย", "ค่าเช่า", "สำนักงาน", "ได้รับแล้ว",
               "0", "0", "-12000", "REF-1", "ฝ่ายบัญชี"]

        normalized = _normalize_import_row(row, legacy_eight_columns=False)

        self.assertIsNone(normalized["invoice"])

    def test_report_export_invoice_value_is_ignored(self):
        row = ["1", "ฝ่ายบัญชี", "2025-02-27", "รายจ่าย", "ค่าเช่า",
               "สำนักงาน", "ได้รับแล้ว", "OFF", "0", "-12000"]

        normalized = _normalize_import_row(
            row, legacy_eight_columns=False, report_export_columns=True,
        )

        self.assertIsNone(normalized["invoice"])

    def test_other_fields_still_parsed_normally(self):
        row = ["2025-02-27", "รายจ่าย", "ค่าน้ำ", "-1200",
               "ค่าน้ำประจำเดือน", "1", "0", "ฝ่ายบัญชี"]

        normalized = _normalize_import_row(row, legacy_eight_columns=True)

        self.assertEqual(normalized["cfstate_date"], date(2025, 2, 27))
        self.assertEqual(normalized["category"], "รายจ่าย")
        self.assertEqual(normalized["source"], "ค่าน้ำ")
        self.assertEqual(normalized["detail"], "ค่าน้ำประจำเดือน")
        self.assertEqual(normalized["income"], Decimal("0"))
        self.assertEqual(normalized["expense"], Decimal("-1200"))
        self.assertEqual(normalized["refrain"], 0)
        self.assertEqual(normalized["department"], "ฝ่ายบัญชี")


class CurrentImportFormatTests(unittest.TestCase):
    def test_positive_expense_is_saved_as_negative_amount(self):
        row = ["2025-02-27", "รายจ่าย", "ค่าเช่า", "สำนักงาน", "", "0",
               "0", "12000", "", "ฝ่ายบัญชี"]

        normalized = _normalize_import_row(row, legacy_eight_columns=False)

        self.assertEqual(normalized["income"], Decimal("0"))
        self.assertEqual(normalized["expense"], Decimal("-12000"))

    def test_blank_source_is_valid(self):
        rows = [
            IMPORT_HEADERS_10,
            ["2025-02-27", "รายจ่าย", "", "สำนักงาน", "", "0",
             "0", "12000", "", "ฝ่ายบัญชี"],
        ]

        prepared = _prepare_import_rows(rows, has_header=True)

        self.assertEqual(prepared[0]["errors"], [])
        self.assertEqual(prepared[0]["data"]["source"], "")
        self.assertEqual(DEFAULT_IMPORT_SOURCE_NAME, "ไม่ระบุ")


class ImportTemplateTests(unittest.IsolatedAsyncioTestCase):
    async def test_xlsx_template_has_separate_positive_income_and_expense_columns(self):
        response = await download_import_template("xlsx")
        content = b"".join([chunk async for chunk in response.body_iterator])
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers, IMPORT_HEADERS_10)
        self.assertTrue(any(
            row[6] == 0 and row[7] > 0
            for row in sheet.iter_rows(min_row=2, values_only=True)
        ))
        self.assertTrue(any(
            row[2] is None
            for row in sheet.iter_rows(min_row=2, values_only=True)
        ))


class DownloadCustomImportTemplateTests(unittest.IsolatedAsyncioTestCase):
    """The per-template download must reflect that template's own column
    order and labels — not the fixed 10-column standard layout."""

    def _template(self):
        return CrmCashflowImportTemplate(
            cfimptpl_id=9,
            comp_id=5,
            cfimptpl_name="ไฟล์จากธนาคาร A",
            cfimptpl_header_row=1,
            cfimptpl_columns=[
                {"field": "department", "label": "แผนก"},
                {"field": "date", "label": "วันที่ทำรายการ"},
                {"field": "category", "label": "ประเภท"},
                {"field": "amount", "label": "จำนวนเงิน"},
            ],
            cfimptpl_status=1,
        )

    async def test_xlsx_headers_and_order_match_the_template(self):
        db = AsyncMock()
        db.execute.return_value = _Result(scalar=self._template())

        response = await download_custom_import_template(
            9, format="xlsx", db=db, company=_Company(5),
        )
        content = b"".join([chunk async for chunk in response.body_iterator])
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(headers, ["แผนก", "วันที่ทำรายการ", "ประเภท", "จำนวนเงิน"])
        # amount column (position 4) keeps a numeric sample, date column
        # (position 2) keeps a real date — order follows the template, not
        # the field's usual position in the standard 10-column layout.
        first_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        self.assertIsInstance(first_row[1], date)
        self.assertIsInstance(first_row[3], (int, float))

    async def test_csv_headers_match_the_template(self):
        db = AsyncMock()
        db.execute.return_value = _Result(scalar=self._template())

        response = await download_custom_import_template(
            9, format="csv", db=db, company=_Company(5),
        )
        content = b"".join([chunk async for chunk in response.body_iterator])
        rows = list(csv.reader(StringIO(content.decode("utf-8-sig"))))

        self.assertEqual(rows[0], ["แผนก", "วันที่ทำรายการ", "ประเภท", "จำนวนเงิน"])
        self.assertEqual(len(rows), 3)  # header + 2 sample rows

    async def test_xlsx_uses_explicit_excel_column_names(self):
        template = self._template()
        template.cfimptpl_columns = [
            {"field": "date", "label": "วันที่", "column": "A"},
            {"field": "income", "label": "รายรับ", "column": "J"},
            {"field": "detail", "label": "รายละเอียด", "column": "AO"},
        ]
        db = AsyncMock()
        db.execute.return_value = _Result(scalar=template)

        response = await download_custom_import_template(
            9, format="xlsx", db=db, company=_Company(5),
        )
        content = b"".join([chunk async for chunk in response.body_iterator])
        sheet = load_workbook(BytesIO(content), data_only=True).active

        self.assertEqual(sheet["A1"].value, "วันที่")
        self.assertEqual(sheet["J1"].value, "รายรับ")
        self.assertEqual(sheet["AO1"].value, "รายละเอียด")
        self.assertIsNone(sheet["B1"].value)

    async def test_xlsx_ignores_system_fields_with_blank_file_columns(self):
        template = self._template()
        template.cfimptpl_columns = [
            {"field": "date", "label": "วันที่", "column": "A"},
            {"field": "category", "label": "หัวข้อ", "column": None},
            {"field": "amount", "label": "จำนวนเงิน", "column": "J"},
            {"field": "detail", "label": "รายละเอียด", "column": None},
        ]
        db = AsyncMock()
        db.execute.return_value = _Result(scalar=template)

        response = await download_custom_import_template(
            9, format="xlsx", db=db, company=_Company(5),
        )
        content = b"".join([chunk async for chunk in response.body_iterator])
        sheet = load_workbook(BytesIO(content), data_only=True).active

        self.assertEqual(sheet["A1"].value, "วันที่")
        self.assertEqual(sheet["J1"].value, "จำนวนเงิน")
        self.assertIsNone(sheet["B1"].value)


class CustomColumnTemplateImportTests(unittest.TestCase):
    """Tests for the user-defined column-mapping template import path."""

    def test_columns_matched_by_position_not_by_field_order(self):
        # Template puts category before date, and department first — parser
        # must follow the template's order, not the field's "natural" order.
        columns = [
            {"field": "department", "label": "แผนก"},
            {"field": "category", "label": "หัวข้อ"},
            {"field": "date", "label": "วันที่"},
            {"field": "income", "label": "รับ"},
            {"field": "expense", "label": "จ่าย"},
        ]
        row = ["ฝ่ายขาย", "รายรับ", "2025-02-27", "15000", "0"]

        normalized = _normalize_import_row_with_template(row, columns)

        self.assertEqual(normalized["department"], "ฝ่ายขาย")
        self.assertEqual(normalized["category"], "รายรับ")
        self.assertEqual(normalized["cfstate_date"], date(2025, 2, 27))
        self.assertEqual(normalized["income"], Decimal("15000"))
        self.assertEqual(normalized["expense"], Decimal("0"))

    def test_skip_field_ignores_column_content(self):
        columns = [
            {"field": "date", "label": "วันที่"},
            {"field": "category", "label": "หัวข้อ"},
            {"field": "skip", "label": "คอลัมน์ที่ไม่ใช้"},
            {"field": "amount", "label": "จำนวนเงิน"},
        ]
        row = ["2025-02-27", "รายจ่าย", "ค่าคำนวณอะไรบางอย่าง", "-500"]

        normalized = _normalize_import_row_with_template(row, columns)

        self.assertEqual(normalized["expense"], Decimal("-500"))
        self.assertEqual(normalized["income"], Decimal("0"))

    def test_combined_amount_column_splits_sign_into_income_and_expense(self):
        columns = [
            {"field": "date", "label": "วันที่"},
            {"field": "category", "label": "หัวข้อ"},
            {"field": "amount", "label": "จำนวนเงิน"},
        ]

        income_row = _normalize_import_row_with_template(
            ["2025-02-27", "รายรับ", "8500"], columns
        )
        expense_row = _normalize_import_row_with_template(
            ["2025-02-27", "รายจ่าย", "-1200"], columns
        )

        self.assertEqual(income_row["income"], Decimal("8500"))
        self.assertEqual(income_row["expense"], Decimal("0"))
        self.assertEqual(expense_row["income"], Decimal("0"))
        self.assertEqual(expense_row["expense"], Decimal("-1200"))

    def test_missing_trailing_columns_are_treated_as_blank(self):
        columns = [
            {"field": "date", "label": "วันที่"},
            {"field": "category", "label": "หัวข้อ"},
            {"field": "amount", "label": "จำนวนเงิน"},
            {"field": "ref", "label": "Ref"},
            {"field": "department", "label": "แผนก"},
        ]
        row = ["2025-02-27", "รายรับ", "100"]  # ref/department columns absent

        normalized = _normalize_import_row_with_template(row, columns)

        self.assertEqual(normalized["ref"], "")
        self.assertEqual(normalized["department"], "")

    def test_explicit_excel_columns_are_used_instead_of_mapping_order(self):
        columns = [
            {"field": "detail", "label": "รายละเอียด", "column": "AO"},
            {"field": "income", "label": "รายรับ", "column": "J"},
            {"field": "date", "label": "วันที่", "column": "A"},
        ]
        row = [None] * 41
        row[0] = "2025-02-27"
        row[9] = "15000"
        row[40] = "ยอดขายออนไลน์"

        normalized = _normalize_import_row_with_template(row, columns)

        self.assertEqual(normalized["cfstate_date"], date(2025, 2, 27))
        self.assertEqual(normalized["income"], Decimal("15000"))
        self.assertEqual(normalized["detail"], "ยอดขายออนไลน์")
        self.assertEqual(normalized["category"], "รายรับ")

    def test_missing_category_is_inferred_from_split_amount_columns(self):
        columns = [
            {"field": "date", "label": "วันที่", "column": "A"},
            {"field": "income", "label": "รายรับ", "column": "J"},
            {"field": "expense", "label": "รายจ่าย", "column": "K"},
        ]
        income_row = ["2025-02-27"] + [None] * 8 + [500, 0]
        expense_row = ["2025-02-28"] + [None] * 8 + [0, 750]

        self.assertEqual(
            _normalize_import_row_with_template(income_row, columns)["category"], "รายรับ"
        )
        self.assertEqual(
            _normalize_import_row_with_template(expense_row, columns)["category"], "รายจ่าย"
        )

    def test_missing_category_is_inferred_from_combined_amount_sign(self):
        columns = [
            {"field": "date", "label": "วันที่", "column": "A"},
            {"field": "amount", "label": "จำนวนเงิน", "column": "B"},
        ]

        self.assertEqual(
            _normalize_import_row_with_template(["2025-02-27", 500], columns)["category"],
            "รายรับ",
        )
        self.assertEqual(
            _normalize_import_row_with_template(["2025-02-28", -750], columns)["category"],
            "รายจ่าย",
        )

    def test_blank_file_column_keeps_optional_system_field_empty(self):
        columns = [
            {"field": "date", "label": "วันที่", "column": "A"},
            {"field": "category", "label": "หัวข้อ", "column": None},
            {"field": "detail", "label": "รายละเอียด", "column": None},
            {"field": "amount", "label": "จำนวนเงิน", "column": "J"},
        ]
        row = ["2025-02-27", "ข้อความที่ต้องไม่ถูกอ่าน"] + [None] * 7 + [500]

        normalized = _normalize_import_row_with_template(row, columns)

        self.assertEqual(normalized["category"], "รายรับ")
        self.assertEqual(normalized["detail"], "")

    def test_unmapped_combined_amount_does_not_override_split_amounts(self):
        columns = [
            {"field": "date", "label": "วันที่", "column": "A"},
            {"field": "income", "label": "รายรับ", "column": "J"},
            {"field": "expense", "label": "รายจ่าย", "column": "K"},
            {"field": "amount", "label": "จำนวนเงิน", "column": None},
        ]
        row = ["2025-02-27"] + [None] * 8 + [500, 0]

        normalized = _normalize_import_row_with_template(row, columns)

        self.assertEqual(normalized["income"], Decimal("500"))
        self.assertEqual(normalized["expense"], Decimal("0"))
        self.assertEqual(normalized["category"], "รายรับ")


class ImportTemplateColumnValidationTests(unittest.TestCase):
    def _columns(self, *fields):
        return [ImportTemplateColumn(field=field, label=field) for field in fields]

    def test_requires_date(self):
        with self.assertRaises(ValueError):
            _validate_import_template_columns(self._columns("amount"))

    def test_category_is_optional(self):
        _validate_import_template_columns(self._columns("date", "amount"))

    def test_requires_at_least_one_amount_shape(self):
        with self.assertRaises(ValueError):
            _validate_import_template_columns(self._columns("date", "category"))

    def test_rejects_both_amount_and_split_income_expense(self):
        with self.assertRaises(ValueError):
            _validate_import_template_columns(
                self._columns("date", "category", "amount", "income")
            )

    def test_rejects_duplicate_field_mapping(self):
        with self.assertRaises(ValueError):
            _validate_import_template_columns(
                self._columns("date", "category", "amount", "date")
            )

    def test_skip_may_repeat_freely(self):
        # "skip" is not a real field, so mapping it to multiple ignored
        # columns must not trip the duplicate-field check.
        _validate_import_template_columns(
            self._columns("date", "category", "amount", "skip", "skip")
        )

    def test_valid_split_income_expense_template_passes(self):
        _validate_import_template_columns(
            self._columns("date", "category", "income", "expense")
        )

    def test_rejects_duplicate_excel_column_mapping(self):
        with self.assertRaises(ValueError):
            _validate_import_template_columns([
                ImportTemplateColumn(field="date", label="วันที่", column="A"),
                ImportTemplateColumn(field="amount", label="จำนวนเงิน", column="A"),
            ])

    def test_excel_column_names_are_normalized(self):
        column = ImportTemplateColumn(field="date", label="วันที่", column=" ao ")
        self.assertEqual(column.column, "AO")

    def test_blank_excel_column_is_normalized_to_none(self):
        column = ImportTemplateColumn(field="detail", label="รายละเอียด", column="  ")
        self.assertIsNone(column.column)

    def test_unmapped_optional_fields_do_not_affect_amount_shape_validation(self):
        _validate_import_template_columns([
            ImportTemplateColumn(field="date", label="วันที่", column="A"),
            ImportTemplateColumn(field="income", label="รายรับ", column=None),
            ImportTemplateColumn(field="expense", label="รายจ่าย", column=None),
            ImportTemplateColumn(field="amount", label="จำนวนเงิน", column="J"),
        ])


if __name__ == "__main__":
    unittest.main()
