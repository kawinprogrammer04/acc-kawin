import io
import asyncio
import tempfile
import inspect
import unittest
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas

from app.services.expense_finance_service import excel_bytes
from app.services.expense_request_service import calculate_totals, render_payment_approval_pdf
from app.services.expense_signature_service import _placement_box, _request_signature_slot, _stamp_pdf
from app.services.approval_service import _request_kind_filter, resolve_approver_for_position, routing_amount
from app.routers.approvals import _employee_organization, _rule_specificity
from app.routers.expense_finance import (
    _accounting_query, _append_legacy_approval_steps, _apply_accounting_pagination, accounting_stats,
    _parse_csv_ints, _parse_csv_values, accounting_view, create_payment,
    replace_payment_proof,
)


class AccountingFilterTests(unittest.TestCase):
    def test_payment_recording_accepts_accounting_view_permission(self):
        dependency = inspect.signature(create_payment).parameters["current_user"].default
        self.assertIs(dependency.dependency, accounting_view)

    def test_payment_proof_upload_accepts_accounting_view_permission(self):
        dependency = inspect.signature(replace_payment_proof).parameters["current_user"].default
        self.assertIs(dependency.dependency, accounting_view)

    def test_csv_filters_trim_empty_values_and_remove_duplicates(self):
        self.assertEqual(
            _parse_csv_values("ready_to_pay, partially_paid,ready_to_pay,,"),
            ["ready_to_pay", "partially_paid"],
        )

    def test_csv_integer_filters_support_multiple_values(self):
        self.assertEqual(_parse_csv_ints("3, 7,3", "department_ids"), [3, 7])

    def test_accounting_query_uses_each_multi_select_filter(self):
        statement = _accounting_query(
            SimpleNamespace(id=9),
            statuses=["ready_to_pay", "partially_paid"],
            department_ids=[3, 7],
            type_ids=[11, 12],
        )
        parameters = list(statement.compile().params.values())
        self.assertIn(["ready_to_pay", "partially_paid"], parameters)
        self.assertIn([3, 7], parameters)
        self.assertIn([11, 12], parameters)

    def test_zero_page_limit_returns_the_unlimited_statement(self):
        statement = SimpleNamespace()
        self.assertIs(_apply_accounting_pagination(statement, 0, 100), statement)

    def test_custom_page_limit_applies_limit_and_offset(self):
        class Statement:
            def __init__(self):
                self.calls = []

            def limit(self, value):
                self.calls.append(("limit", value))
                return self

            def offset(self, value):
                self.calls.append(("offset", value))
                return self

        statement = Statement()
        self.assertIs(_apply_accounting_pagination(statement, 75, 150), statement)
        self.assertEqual(statement.calls, [("limit", 75), ("offset", 150)])

    def test_accounting_stats_use_the_same_filters_as_the_list(self):
        class Result:
            def __init__(self, rows):
                self.rows = rows

            def scalars(self):
                return self

            def all(self):
                return self.rows

        class Database:
            statement = None

            def __init__(self):
                self.results = iter([
                    [SimpleNamespace(
                        id="request-1", status="ready_to_pay",
                        remaining_amount=Decimal("125.00"), net_amount=Decimal("125.00"),
                        settlement_due_date=None,
                    )],
                    [],
                    [],
                ])

            async def execute(self, statement):
                if self.statement is None:
                    self.statement = statement
                return Result(next(self.results))

        database = Database()
        result = asyncio.run(accounting_stats(
            statuses="ready_to_pay", query="ACC-EXP", department_ids="3,7",
            type_ids="11", date_from=date(2026, 8, 1), date_to=date(2026, 8, 31),
            withholding_only=True, db=database, current_user=SimpleNamespace(),
            company=SimpleNamespace(id=9),
        ))

        parameters = list(database.statement.compile().params.values())
        self.assertIn(["ready_to_pay"], parameters)
        self.assertIn([3, 7], parameters)
        self.assertIn([11], parameters)
        self.assertIn("%ACC-EXP%", parameters)
        self.assertEqual(result.ready_to_pay_count, 1)
        self.assertEqual(result.pending_approval_count, 0)
        self.assertEqual(result.transfer_amount_total, Decimal("125.00"))


def request(**overrides):
    values = dict(
        vat_mode="rate", vat_rate=Decimal("7"), vat_amount=Decimal("0"),
        withholding_required=True, withholding_mode="rate", withholding_rate=Decimal("3"),
        withholding_amount=Decimal("0"), discount_amount=Decimal("0"),
        price_mode="exclude_vat", gross_up_enabled=False,
    )
    values.update(overrides)
    return SimpleNamespace(**values)


class ExpenseCalculationTests(unittest.TestCase):
    def test_vat_wht_discount_and_net_are_authoritative(self):
        req = request(discount_amount=Decimal("100"))
        items = [SimpleNamespace(line_total=Decimal("1000"))]
        result = calculate_totals(req, items)
        self.assertEqual(result["subtotal"], Decimal("1000.00"))
        self.assertEqual(result["price_before_vat"], Decimal("900.00"))
        self.assertEqual(result["vat_amount"], Decimal("63.00"))
        self.assertEqual(result["withholding_amount"], Decimal("27.00"))
        self.assertEqual(result["payable_total"], Decimal("936.00"))

    def test_gross_up(self):
        req = request(vat_mode="none", gross_up_enabled=True, withholding_rate=Decimal("3"))
        result = calculate_totals(req, [SimpleNamespace(line_total=Decimal("970"))])
        self.assertEqual(result["withholding_amount"], Decimal("30.00"))
        self.assertEqual(result["grand_total"], Decimal("1000.00"))
        self.assertEqual(result["payable_total"], Decimal("970.00"))

    def test_hr_requested_net_gross_up(self):
        req = request(gross_up_enabled=True, requested_net_amount=Decimal("1000"))
        result = calculate_totals(req, [SimpleNamespace(line_total=Decimal("1100"))])
        self.assertEqual(result["vat_amount"], Decimal("77.00"))
        self.assertEqual(result["withholding_amount"], Decimal("28.55"))
        self.assertEqual(result["grand_total"], Decimal("1177.00"))
        self.assertEqual(result["payable_total"], Decimal("1000.00"))

    def test_installment_payment_amount_overrides_taxable_base(self):
        # A 10,000 baht claim split into installments: VAT/หัก ณ ที่จ่าย must be
        # computed from the 500 baht actually being paid this round, not the
        # 10,000 baht full line-items total — and discount doesn't apply on top.
        req = request(discount_amount=Decimal("100"), installment_payment_amount=Decimal("500"))
        result = calculate_totals(req, [SimpleNamespace(line_total=Decimal("10000"))])
        self.assertEqual(result["subtotal"], Decimal("10000.00"))
        self.assertEqual(result["discount_amount"], Decimal("0.00"))
        self.assertEqual(result["price_before_vat"], Decimal("500.00"))
        self.assertEqual(result["vat_amount"], Decimal("35.00"))
        self.assertEqual(result["withholding_amount"], Decimal("15.00"))
        self.assertEqual(result["payable_total"], Decimal("520.00"))

    def test_installment_payment_amount_absent_is_unchanged(self):
        # No override present (every existing row, every non-chain request) must
        # take byte-for-byte the same path as before this feature existed.
        req = request(discount_amount=Decimal("100"))
        self.assertIsNone(getattr(req, "installment_payment_amount", None))
        result = calculate_totals(req, [SimpleNamespace(line_total=Decimal("1000"))])
        self.assertEqual(result["price_before_vat"], Decimal("900.00"))


class ApprovalRoutingAmountTests(unittest.TestCase):
    def test_installment_document_routes_on_the_full_claim_total(self):
        # A 500 baht installment slice of a 10,000 baht claim must still route
        # to whichever approver tier the FULL 10,000 baht claim requires — not
        # get waved through on a lower tier just because this document's own
        # amount happens to be small.
        req = request(amount=Decimal("520"), installment_target_amount=Decimal("9500"))
        self.assertEqual(routing_amount(req), Decimal("9500"))

    def test_non_installment_document_routes_on_its_own_amount(self):
        req = request(amount=Decimal("1177"))
        self.assertIsNone(getattr(req, "installment_target_amount", None))
        self.assertEqual(routing_amount(req), Decimal("1177"))


class ApprovalRequestKindTests(unittest.TestCase):
    def test_hr_allowance_uses_its_exact_global_fallback(self):
        condition = str(_request_kind_filter("allowance"))
        self.assertIn("approval_rules.request_kind", condition)
        self.assertNotIn("approval_rules.request_kind IS NULL", condition)

    def test_hr_ot_uses_its_exact_global_fallback(self):
        condition = str(_request_kind_filter("ot"))
        self.assertIn("approval_rules.request_kind", condition)
        self.assertNotIn("approval_rules.request_kind IS NULL", condition)

    def test_hr_wildcard_is_excluded_from_direct_payment(self):
        condition = str(_request_kind_filter("direct_payment"))
        self.assertIn("approval_rules.source_system", condition)
        self.assertIn("approval_rules.request_kind IS NULL", condition)

    def test_hr_wildcard_remains_available_to_reimbursement(self):
        condition = str(_request_kind_filter("reimbursement"))
        self.assertNotIn("approval_rules.source_system", condition)
        self.assertIn("approval_rules.request_kind IS NULL", condition)

    def test_department_wide_hr_scope_keeps_lower_specificity(self):
        scope = {
            "company_name": None,
            "department_name": "CRM",
            "requester_position_name": None,
            "expense_type_code": "GENERAL",
            "request_kind": None,
        }
        self.assertEqual(_rule_specificity(scope, None), 2)


class EmployeeApprovalResolutionTests(unittest.TestCase):
    def test_single_employee_in_position_is_implicit_primary_approver(self):
        class Result:
            def __init__(self, value=None, rows=None):
                self.value = value
                self.rows = rows or []

            def scalar_one_or_none(self):
                return self.value

            def scalars(self):
                return self

            def all(self):
                return self.rows

        class Database:
            def __init__(self):
                self.results = iter([Result(), Result(), Result(rows=[42])])

            async def execute(self, _query):
                return next(self.results)

        resolved = asyncio.run(resolve_approver_for_position(
            Database(), 7, datetime(2026, 8, 11, tzinfo=timezone.utc)
        ))
        self.assertEqual(resolved, 42)

    def test_employee_without_department_can_create_expense_request(self):
        class Result:
            def __init__(self, value):
                self.value = value

            def scalar_one_or_none(self):
                return self.value

        position = SimpleNamespace(id=7, company_id=1, department_id=99, name="CEO")
        membership = SimpleNamespace(department_id=None)

        class Database:
            def __init__(self):
                # Position lookup, active position assignment, company membership.
                self.results = iter([Result(position), Result(1), Result(membership)])

            async def execute(self, _query):
                return next(self.results)

            async def get(self, *_args):
                raise AssertionError("NULL membership department must not fall back to the position department")

        resolved_position, resolved_department = asyncio.run(_employee_organization(
            Database(),
            SimpleNamespace(id=42, is_platform_admin=False),
            SimpleNamespace(id=1),
            7,
        ))
        self.assertIs(resolved_position, position)
        self.assertIsNone(resolved_department)


class AccountingApprovalTimelineTests(unittest.TestCase):
    def test_all_legacy_steps_are_appended_in_order(self):
        steps_by_request = {"request-1": []}
        legacy_steps = [
            SimpleNamespace(
                id=90, expense_request_id="request-1", revision=1, step_no=1,
                name="Accounting", approvers=[{"name": "ปิยะธิดา", "status": "approved"}],
                status="approved", completed_at="2026-08-18 15:46:32",
            ),
            SimpleNamespace(
                id=91, expense_request_id="request-1", revision=1, step_no=2,
                name="Manager Accountant", approvers=[{"name": "อิสราภรณ์", "status": "waiting"}],
                status="active", completed_at=None,
            ),
        ]

        _append_legacy_approval_steps(steps_by_request, legacy_steps, {"request-1": 1}, set())

        self.assertEqual([step["step_no"] for step in steps_by_request["request-1"]], [1, 2])
        self.assertEqual(steps_by_request["request-1"][1]["name"], "Manager Accountant")

    def test_native_route_prevents_legacy_duplicate(self):
        native_step = {"id": 10, "step_no": 1, "name": "ACC route"}
        steps_by_request = {"request-1": [native_step]}
        legacy_steps = [SimpleNamespace(
            id=90, expense_request_id="request-1", revision=1, step_no=1,
            name="HR route", approvers=[], status="approved", completed_at=None,
        )]

        _append_legacy_approval_steps(
            steps_by_request, legacy_steps, {"request-1": 1}, {"request-1"},
        )

        self.assertEqual(steps_by_request["request-1"], [native_step])


class ExpenseExportTests(unittest.TestCase):
    def test_formula_injection_is_escaped(self):
        row = SimpleNamespace(
            id="request-1", request_no="=HYPERLINK(\"bad\")", submitted_at=None,
            request_format="reimbursement", expense_type_id=1, company_name_snapshot="Test Co",
            department_id=1, department_name_snapshot="Finance", requester_name_snapshot="+SUM(1,1)",
            recipient_name="Recipient", bank_name="Test Bank", bank_account_name="Tester",
            bank_account_number_encrypted=None, title="@cmd", withholding_decision="none", status="completed",
            gross_amount=1, vat_amount=0, withholding_amount=0, net_amount=1, paid_amount=1, remaining_amount=0,
        )
        workbook = load_workbook(io.BytesIO(excel_bytes(
            [row], expense_type_names={1: "General"}, department_names={1: "Finance"},
        )))
        values = list(workbook.active.values)[1]
        self.assertTrue(values[0].startswith("'="))
        self.assertTrue(values[7].startswith("'+"))
        self.assertTrue(values[6].startswith("'@"))

    def test_columns_match_hr_and_keep_acc_item_column(self):
        row = SimpleNamespace(
            id="request-1", request_no="ACC-EXP-202608-000001", submitted_at=None,
            request_format="reimbursement", expense_type_id=1, company_name_snapshot="Test Co",
            department_id=1, department_name_snapshot="Finance", requester_name_snapshot="Requester",
            recipient_name="Recipient", bank_name="Test Bank", bank_account_name="Tester",
            bank_account_number_encrypted=None, title="รายการของ ACC", withholding_decision=None,
            status="ready_to_pay", gross_amount=100, vat_amount=0, withholding_amount=3,
            net_amount=97, paid_amount=0, remaining_amount=97,
        )
        workbook = load_workbook(io.BytesIO(excel_bytes(
            [row], expense_type_names={1: "General"}, department_names={1: "Finance"},
        )))
        headings = list(workbook.active.values)[0]
        self.assertEqual(headings, (
            "เลขที่คำขอ", "วันที่ส่ง", "ประเภทคำขอ", "หมวดค่าใช้จ่าย", "บริษัท", "แผนก", "รายการ",
            "ผู้ขอ", "ผู้รับเงิน", "ธนาคาร", "ชื่อบัญชี", "เลขบัญชี", "ยอดอนุมัติ",
            "ภาษีหัก ณ ที่จ่าย", "ผลพิจารณาภาษี", "ยอดโอนสุทธิ", "ยอดส่วนต่างเงินทดรอง",
            "จ่ายแล้ว", "คงเหลือ", "สถานะ", "วันที่จ่ายล่าสุด", "เลขอ้างอิง",
        ))
        self.assertEqual(list(workbook.active.values)[1][6], "รายการของ ACC")


class ExpenseRequestPdfTests(unittest.TestCase):
    def test_hr_layout_renders_multiple_pages_and_signature_grid(self):
        req = request(
            id="request-id", request_no="EXP-202608-000099", current_revision=2,
            company_name_snapshot="บริษัท กวิน บราเธอร์ส จำกัด",
            requester_name_snapshot="ผู้ขอ ทดสอบ", recipient_name="ผู้รับ ทดสอบ",
            title="ค่าใช้จ่ายทดสอบ", description="วัตถุประสงค์สำหรับทดสอบ PDF",
            request_format="reimbursement", department_name_snapshot="บัญชี",
            created_at=datetime(2026, 8, 11, tzinfo=timezone.utc), submitted_at=None,
            required_date=date(2026, 8, 15),
        )
        items = [SimpleNamespace(
            description=f"รายการที่ {index}", quantity=Decimal("1"), unit="รายการ",
            unit_price=Decimal("100.0000000000"), line_total=Decimal("100.00"),
        ) for index in range(1, 21)]
        company = SimpleNamespace(
            name_th="บริษัท กวิน บราเธอร์ส จำกัด", name_en="Kawin Brothers Co., Ltd.",
            address="88/6-7 ถนนกาญจนาภิเษก กรุงเทพฯ 10160", phone="082-494-9524",
            tax_id="0105561208119", logo_url=None,
        )
        requester = SimpleNamespace(full_name="ผู้ขอ ทดสอบ", username="requester")
        cells = [{"role": "ผู้อนุมัติลำดับ 1 - CEO", "name": "ผู้อนุมัติ ทดสอบ", "is_requester": False}]
        with tempfile.TemporaryDirectory() as folder:
            target = Path(folder) / "request.pdf"
            render_payment_approval_pdf(
                req, items, company, requester, "Manager Accountant", "ทั่วไป", target,
                "บัญชี", cells,
            )
            reader = PdfReader(str(target))
            self.assertEqual(len(reader.pages), 2)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            self.assertIn("Payment Approval", text)
            self.assertIn("EXP-202608-000099", text)
            self.assertGreater(target.stat().st_size, 10_000)

    def test_installment_document_shows_subtotal_and_installment_amount_rows(self):
        # รวมเงิน must always be the full items subtotal, with a separate row
        # showing the amount actually being requested this installment.
        req = request(
            id="request-id", request_no="EXP-202608-000099-1", current_revision=1,
            company_name_snapshot="บริษัท กวิน บราเธอร์ส จำกัด",
            requester_name_snapshot="ผู้ขอ ทดสอบ", recipient_name="ผู้รับ ทดสอบ",
            title="ค่าใช้จ่ายทดสอบ", description="วัตถุประสงค์สำหรับทดสอบ PDF งวด",
            request_format="reimbursement", department_name_snapshot="บัญชี",
            created_at=datetime(2026, 8, 11, tzinfo=timezone.utc), submitted_at=None,
            required_date=date(2026, 8, 15), installment_enabled=True,
            installment_payment_amount=Decimal("500"),
        )
        items = [SimpleNamespace(
            description="รายการทดสอบ", quantity=Decimal("1"), unit="รายการ",
            unit_price=Decimal("10000.0000000000"), line_total=Decimal("10000.00"),
        )]
        company = SimpleNamespace(
            name_th="บริษัท กวิน บราเธอร์ส จำกัด", name_en="Kawin Brothers Co., Ltd.",
            address="88/6-7 ถนนกาญจนาภิเษก กรุงเทพฯ 10160", phone="082-494-9524",
            tax_id="0105561208119", logo_url=None,
        )
        requester = SimpleNamespace(full_name="ผู้ขอ ทดสอบ", username="requester")
        cells = [{"role": "ผู้อนุมัติลำดับ 1 - CEO", "name": "ผู้อนุมัติ ทดสอบ", "is_requester": False}]
        with tempfile.TemporaryDirectory() as folder:
            target = Path(folder) / "request.pdf"
            render_payment_approval_pdf(
                req, items, company, requester, "Manager Accountant", "ทั่วไป", target,
                "บัญชี", cells,
            )
            reader = PdfReader(str(target))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            # pypdf's text extraction mangles a few Thai vowel glyphs in this font
            # (e.g. "จ่าย" -> "จ่ำย"), so match the unaffected tail of the label.
            self.assertIn("ในงวดนี้", text)
            self.assertIn("10,000.00", text)  # รวมเงิน = full items subtotal
            self.assertIn("500.00", text)  # จำนวนที่ต้องจ่ายในงวดนี้ = installment amount


class SignaturePdfTests(unittest.TestCase):
    def test_primary_signature_slot_matches_hr_grid(self):
        first_step = _request_signature_slot(1, 2)
        self.assertAlmostEqual(first_step["x"], .307)
        self.assertAlmostEqual(first_step["y"], .795878)
        self.assertAlmostEqual(first_step["width"], .155)
        self.assertAlmostEqual(first_step["height"], .026)
        self.assertEqual(first_step["page_number"], 2)
        self.assertEqual(first_step["coordinate_system"], "top_left")
        fourth_step = _request_signature_slot(4, 3)
        self.assertAlmostEqual(fourth_step["x"], .0773)
        self.assertAlmostEqual(fourth_step["y"], .858878)
        self.assertEqual(fourth_step["page_number"], 3)

    def test_browser_top_left_coordinates_are_converted_for_reportlab(self):
        self.assertEqual(
            _placement_box({
                "x": .10, "y": .20, "width": .30, "height": .10,
                "coordinate_system": "top_left",
            }, 100, 200),
            (10.0, 140.0, 30.0, 20.0),
        )

    def test_stamp_page_13_and_rotated_page(self):
        with tempfile.TemporaryDirectory() as folder:
            source = Path(folder) / "source.pdf"
            stream = io.BytesIO()
            pdf = canvas.Canvas(stream)
            for index in range(20):
                pdf.drawString(50, 800, f"page {index + 1}")
                pdf.showPage()
            pdf.save(); stream.seek(0)
            reader = PdfReader(stream); writer = PdfWriter()
            for index, page in enumerate(reader.pages):
                if index == 12: page.rotate(90)
                writer.add_page(page)
            with source.open("wb") as target: writer.write(target)
            signature_stream = io.BytesIO(); signature = canvas.Canvas(signature_stream, pagesize=(150, 50))
            signature.drawString(5, 20, "SIGNED"); signature.save()
            # reportlab produced a PDF, but ImageReader needs a raster. Use a
            # tiny valid PNG fixture instead.
            png = base64_png()
            output = _stamp_pdf(source, png, [{"page_number": 13, "x": .5, "y": .1, "width": .2, "height": .08}])
            stamped = PdfReader(io.BytesIO(output))
            self.assertEqual(len(stamped.pages), 20)
            self.assertEqual(stamped.pages[12].rotation, 0)


def base64_png() -> bytes:
    import base64
    return base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M/wHwAF/gL+XwM7WQAAAABJRU5ErkJggg=="
    )


if __name__ == "__main__":
    unittest.main()
