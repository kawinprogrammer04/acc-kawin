"""Transactional accounting, settlement, export and notification operations."""
from __future__ import annotations

import base64
import hashlib
import html
import io
import re
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment
from sqlalchemy import func, select, update
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import AsyncSession
from weasyprint import HTML

from app.core.config import settings
from app.models.approval import ExpenseRequest, ExpenseType
from app.models.expense_finance import (
    ExpensePayment, ExpenseRequestHistory, ExpenseSettlement, ExpenseSettlementItem,
    ExpenseWithholdingTaxCertificate, SystemNotification,
)
from app.services.expense_request_service import decrypt_account_number

MONEY = Decimal("0.01")


def money(value) -> Decimal:
    return Decimal(value or 0).quantize(MONEY, rounding=ROUND_HALF_UP)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def safe_filename(name: str | None, fallback: str) -> str:
    cleaned = Path(name or fallback).name
    return cleaned if cleaned and cleaned not in {".", ".."} else fallback


def decode_private_file(name: str | None, content: str | None, *, required: bool = False) -> tuple[str | None, bytes | None]:
    if not content:
        if required:
            raise ValueError("กรุณาแนบหลักฐาน")
        return None, None
    try:
        data = base64.b64decode(content, validate=True)
    except Exception as exc:
        raise ValueError("รูปแบบไฟล์ base64 ไม่ถูกต้อง") from exc
    if not data or len(data) > 15 * 1024 * 1024:
        raise ValueError("ไฟล์หลักฐานต้องมีขนาด 1 byte ถึง 15 MB")
    filename = safe_filename(name, "proof.bin")
    if Path(filename).suffix.lower() not in {".pdf", ".jpg", ".jpeg", ".png"}:
        raise ValueError("หลักฐานรองรับเฉพาะ PDF, JPG และ PNG")
    return filename, data


def _store(request_id: str, folder: str, filename: str, data: bytes) -> tuple[str, str]:
    digest = sha256_bytes(data)
    suffix = Path(filename).suffix.lower()[:10]
    path = Path(settings.EXPENSE_REQUEST_UPLOAD_DIR) / request_id / folder / f"{digest}{suffix}"
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        path.write_bytes(data)
    return str(path), digest


def add_history(db: AsyncSession, req: ExpenseRequest, event: str, actor_user_id: int | None,
                from_status: str | None = None, note: str | None = None, snapshot: dict | None = None) -> None:
    db.add(ExpenseRequestHistory(
        company_id=req.company_id, expense_request_id=req.id, revision=req.current_revision,
        event=event, from_status=from_status, to_status=req.status,
        actor_user_id=actor_user_id, note=note, snapshot=snapshot or {},
    ))


def notify(db: AsyncSession, req: ExpenseRequest, user_id: int, kind: str, title: str,
           message: str, dedupe_key: str) -> None:
    db.add(SystemNotification(
        company_id=req.company_id, user_id=user_id, expense_request_id=req.id,
        type=kind, title=title, message=message,
        action_url=f"/expense-requests/{req.id}", dedupe_key=dedupe_key,
    ))


async def _refresh_installment_chain_status(db: AsyncSession, locked: ExpenseRequest) -> None:
    """Recompute the chain-level "still not fully disbursed" status.

    Called after a payment fully closes one document in an installment chain.
    Writes the result onto the root row and fans it out to every sibling so
    list views can read it off any single row without a join. Never touches
    the per-document `status` state machine — that stays exactly as today.
    """
    root_id = locked.installment_chain_root_id
    root = locked if locked.id == root_id else (await db.execute(
        select(ExpenseRequest).where(ExpenseRequest.id == root_id).with_for_update()
    )).scalar_one()
    chain_paid = (await db.execute(select(func.coalesce(func.sum(ExpenseRequest.paid_amount), 0)).where(
        ExpenseRequest.installment_chain_root_id == root_id
    ))).scalar_one()
    new_status = "fully_disbursed" if money(chain_paid) >= money(root.installment_target_amount or 0) else "in_progress"
    if new_status != root.installment_chain_status:
        await db.execute(update(ExpenseRequest).where(
            ExpenseRequest.installment_chain_root_id == root_id
        ).values(installment_chain_status=new_status))
        root.installment_chain_status = new_status
        if locked.id != root.id:
            locked.installment_chain_status = new_status


async def record_payment(db: AsyncSession, req: ExpenseRequest, payload, actor_user_id: int) -> ExpensePayment:
    existing = (await db.execute(select(ExpensePayment).where(
        ExpensePayment.idempotency_key == payload.idempotency_key
    ))).scalar_one_or_none()
    if existing:
        return existing
    locked = (await db.execute(select(ExpenseRequest).where(
        ExpenseRequest.id == req.id, ExpenseRequest.company_id == req.company_id
    ).with_for_update())).scalar_one()
    if locked.status not in {"ready_to_pay", "partially_paid"}:
        raise ValueError(f"บันทึกจ่ายได้เฉพาะสถานะพร้อมจ่ายหรือจ่ายบางส่วน (ปัจจุบัน: {locked.status})")
    amount = money(payload.amount)
    remaining = money(locked.net_amount or locked.amount) - money(locked.paid_amount)
    if amount > remaining:
        raise ValueError("ยอดจ่ายมากกว่ายอดคงเหลือ กรุณาตรวจสอบยอด")
    if amount < remaining and not (locked.installment_enabled and locked.installment_payment_amount is None):
        # Once a document has been split into an installment (installment_payment_amount
        # is set), it must be paid in full in one shot — the "แบ่งจ่ายยังไม่ครบ" state
        # lives across sibling documents in the chain, not as partial payments on one.
        # The legacy "tick the box, pay whatever amount, however many times" behavior
        # (installment_enabled=true but never split into a chain) is unchanged.
        raise ValueError("ระบบไม่รองรับแบ่งจ่ายภายในเอกสารนี้ กรุณาชำระเต็มจำนวนของงวดนี้ หรือสร้างงวดถัดไปสำหรับส่วนที่เหลือ")
    file_name, proof = decode_private_file(payload.proof_file_name, payload.proof_content_base64, required=True)
    proof_path, proof_hash = _store(locked.id, "payments", file_name, proof)
    remaining_after = money(remaining - amount)
    if remaining_after > 0:
        payment_type = "partial"
    else:
        payment_type = "adjustment" if locked.current_revision > 1 and locked.request_format == "advance" else "full"
    payment = ExpensePayment(
        company_id=locked.company_id, expense_request_id=locked.id, revision=locked.current_revision,
        payment_type=payment_type, amount=amount, paid_at=payload.paid_at,
        method=payload.method, reference_no=payload.reference_no, note=payload.note,
        proof_file_name=file_name, proof_file_path=proof_path, proof_sha256=proof_hash,
        recorded_by=actor_user_id, idempotency_key=payload.idempotency_key,
    )
    db.add(payment)
    previous = locked.status
    locked.paid_amount = money(locked.paid_amount) + amount
    locked.remaining_amount = remaining_after
    locked.paid_at = payload.paid_at
    if remaining_after > 0:
        locked.status = "partially_paid"
    elif locked.request_format == "advance" and payment_type == "full":
        locked.status = "settlement_due"
        expense_type = await db.get(ExpenseType, locked.expense_type_id)
        settlement_days = expense_type.settlement_days if expense_type else 7
        locked.settlement_due_date = payload.paid_at.date() + timedelta(days=settlement_days)
    else:
        locked.status = "completed"
        locked.completed_at = datetime.now(timezone.utc)
    if remaining_after <= 0 and locked.installment_chain_root_id is not None:
        await _refresh_installment_chain_status(db, locked)
    add_history(db, locked, "payment_recorded", actor_user_id, previous,
                snapshot={"amount": str(amount), "proof_sha256": proof_hash})
    notify(db, locked, locked.requester_user_id, "paid", "บันทึกจ่ายเงินแล้ว",
           f"{locked.request_no} จ่ายแล้ว {amount:,.2f} บาท", f"paid:{locked.id}:{payload.idempotency_key}")
    await db.commit()
    await db.refresh(payment)
    return payment


async def replace_payment_proof(db: AsyncSession, payment: ExpensePayment, payload,
                                actor_user_id: int) -> ExpensePayment:
    if payment.voided_at:
        raise ValueError("เปลี่ยนหลักฐานของรายการจ่ายที่ยกเลิกแล้วไม่ได้")
    req = (await db.execute(select(ExpenseRequest).where(
        ExpenseRequest.id == payment.expense_request_id,
        ExpenseRequest.company_id == payment.company_id,
    ).with_for_update())).scalar_one()
    file_name, proof = decode_private_file(
        payload.proof_file_name, payload.proof_content_base64, required=True
    )
    proof_path, proof_hash = _store(req.id, "payments", file_name, proof)
    old_hash = payment.proof_sha256
    payment.proof_file_name = file_name
    payment.proof_file_path = proof_path
    payment.proof_sha256 = proof_hash
    payment.updated_at = datetime.now(timezone.utc)
    add_history(db, req, "payment_proof_replaced", actor_user_id, req.status, payload.reason, {
        "payment_id": payment.id, "old_sha256": old_hash, "new_sha256": proof_hash,
    })
    await db.commit()
    await db.refresh(payment)
    return payment


async def void_payment(db: AsyncSession, payment: ExpensePayment, reason: str,
                       actor_user_id: int) -> ExpensePayment:
    if payment.voided_at:
        return payment
    req = (await db.execute(select(ExpenseRequest).where(
        ExpenseRequest.id == payment.expense_request_id,
        ExpenseRequest.company_id == payment.company_id,
    ).with_for_update())).scalar_one()
    active_later = (await db.execute(select(ExpensePayment.id).where(
        ExpensePayment.expense_request_id == req.id,
        ExpensePayment.voided_at.is_(None),
        ExpensePayment.created_at > payment.created_at,
    ).limit(1))).scalar_one_or_none()
    if active_later:
        raise ValueError("ต้องยกเลิกรายการจ่ายล่าสุดก่อน")
    approved_settlement = (await db.execute(select(ExpenseSettlement.id).where(
        ExpenseSettlement.expense_request_id == req.id,
        ExpenseSettlement.status == "approved",
    ).limit(1))).scalar_one_or_none()
    if approved_settlement:
        raise ValueError("ยกเลิกการจ่ายไม่ได้หลังอนุมัติเคลียร์เงินแล้ว")
    previous = req.status
    now = datetime.now(timezone.utc)
    payment.voided_at = now
    payment.voided_by = actor_user_id
    payment.void_reason = reason
    payment.updated_at = now
    req.paid_amount = money(max(Decimal("0"), money(req.paid_amount) - money(payment.amount)))
    req.remaining_amount = money(req.net_amount - req.paid_amount)
    req.paid_at = None
    req.completed_at = None
    req.settlement_due_date = None
    req.status = "ready_to_pay" if req.paid_amount <= 0 else "partially_paid"
    add_history(db, req, "payment_voided", actor_user_id, previous, reason, {
        "payment_id": payment.id, "amount": str(money(payment.amount)),
        "proof_sha256": payment.proof_sha256,
    })
    notify(db, req, req.requester_user_id, "payment_voided", "ยกเลิกรายการจ่ายเงิน",
           reason, f"payment-voided:{payment.id}")
    await db.commit()
    await db.refresh(payment)
    return payment


async def submit_settlement(db: AsyncSession, req: ExpenseRequest, payload, actor_user_id: int) -> ExpenseSettlement:
    locked = (await db.execute(select(ExpenseRequest).where(
        ExpenseRequest.id == req.id, ExpenseRequest.company_id == req.company_id
    ).with_for_update())).scalar_one()
    if locked.requester_user_id != actor_user_id:
        raise PermissionError("เฉพาะเจ้าของคำขอเท่านั้นที่ส่งเคลียร์เงินได้")
    if locked.status != "settlement_due":
        raise ValueError("คำขอนี้ไม่ได้อยู่ระหว่างรอเคลียร์เงิน")
    advance = money(locked.paid_amount)
    actual = money(payload.actual_amount)
    difference = money(actual - advance)
    kind = "equal" if difference == 0 else ("refund" if difference < 0 else "additional")
    proof_path = proof_hash = None
    if kind == "refund":
        file_name, proof = decode_private_file(
            payload.refund_proof_file_name, payload.refund_proof_content_base64, required=True
        )
        proof_path, proof_hash = _store(locked.id, "settlements", file_name, proof)
    settlement = ExpenseSettlement(
        company_id=locked.company_id, expense_request_id=locked.id, revision=locked.current_revision,
        advance_amount=advance, actual_amount=actual, difference_amount=difference,
        settlement_type=kind, status="submitted", note=payload.note,
        refund_proof_path=proof_path, refund_proof_sha256=proof_hash,
        submitted_by=actor_user_id,
    )
    db.add(settlement)
    await db.flush()
    for index, item in enumerate(payload.items, 1):
        db.add(ExpenseSettlementItem(
            settlement_id=settlement.id, sort_order=index, description=item.description,
            quantity=item.quantity, unit=item.unit, unit_price=item.unit_price,
            line_total=money(item.quantity * item.unit_price),
        ))
    previous = locked.status
    if kind == "additional":
        locked.current_revision += 1
        locked.version += 1
        # The adjustment revision is routed only on the additional amount;
        # the original advance and its signatures remain immutable history.
        locked.amount = money(difference)
        locked.gross_amount = money(difference)
        locked.net_amount = money(difference)
        locked.remaining_amount = money(difference)
        locked.status = "settlement_due"
        event = "settlement_additional_revision_created"
    else:
        locked.status = "settlement_review"
        event = "settlement_submitted"
    add_history(db, locked, event, actor_user_id, previous,
                snapshot={"settlement_id": settlement.id, "type": kind, "difference": str(difference)})
    if kind == "additional":
        from app.services import approval_service
        await approval_service.submit_expense_request(db, locked)
    else:
        await db.commit()
    await db.refresh(settlement)
    return settlement


async def review_settlement(db: AsyncSession, settlement: ExpenseSettlement, action: str,
                            comment: str | None, actor_user_id: int) -> ExpenseSettlement:
    settlement = (await db.execute(select(ExpenseSettlement).where(
        ExpenseSettlement.id == settlement.id
    ).with_for_update())).scalar_one()
    req = (await db.execute(select(ExpenseRequest).where(
        ExpenseRequest.id == settlement.expense_request_id
    ).with_for_update())).scalar_one()
    if settlement.status != "submitted" or req.status != "settlement_review":
        raise ValueError("รายการเคลียร์นี้ไม่ได้รอตรวจสอบ")
    previous = req.status
    settlement.reviewed_by = actor_user_id
    settlement.reviewed_at = datetime.now(timezone.utc)
    settlement.review_comment = comment
    if action == "approve":
        settlement.status = "approved"
        req.status = "completed"
        req.settled_at = settlement.reviewed_at
        req.completed_at = settlement.reviewed_at
        event, title = "settlement_approved", "เคลียร์เงินเรียบร้อยแล้ว"
    else:
        settlement.status = "returned"
        req.status = "settlement_due"
        event, title = "settlement_returned", "รายการเคลียร์เงินถูกส่งกลับ"
    add_history(db, req, event, actor_user_id, previous, comment, {"settlement_id": settlement.id})
    notify(db, req, req.requester_user_id, event, title, comment or req.request_no,
           f"{event}:{settlement.id}:{settlement.reviewed_at.date()}")
    await db.commit()
    await db.refresh(settlement)
    return settlement


def excel_bytes(rows: list[ExpenseRequest]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Requests"
    ws.append(["เลขที่", "วันที่", "ผู้ขอ", "บัญชีสำหรับ SCB", "รายการ", "ประเภท", "สถานะ", "ยอดรวม", "VAT", "หัก ณ ที่จ่าย", "ยอดสุทธิ", "จ่ายแล้ว", "คงเหลือ"])
    dangerous = ("=", "+", "-", "@", "\t", "\r")
    def safe(value):
        if isinstance(value, str) and value.startswith(dangerous):
            return "'" + value
        return value
    for r in rows:
        bank_line = " · ".join(part for part in (r.bank_name, r.bank_account_name) if part) or "-"
        account_number = decrypt_account_number(r.bank_account_number_encrypted) or "-"
        ws.append([safe(r.request_no), r.request_date.isoformat(), safe(r.requester_name_snapshot),
                   safe(f"{bank_line}\n{account_number}"), safe(r.title),
                   safe(r.request_format), safe(r.status), float(money(r.gross_amount)), float(money(r.vat_amount)),
                   float(money(r.withholding_amount)), float(money(r.net_amount)), float(money(r.paid_amount)),
                   float(money(r.remaining_amount))])
        ws.cell(row=ws.max_row, column=4).alignment = Alignment(wrapText=True, vertical="top")
    ws.column_dimensions["D"].width = 32
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


async def issue_wht_certificate(db: AsyncSession, req: ExpenseRequest, actor_user_id: int) -> ExpenseWithholdingTaxCertificate:
    if money(req.withholding_amount) <= 0:
        raise ValueError("คำขอนี้ไม่มียอดหัก ณ ที่จ่าย")
    existing = (await db.execute(select(ExpenseWithholdingTaxCertificate).where(
        ExpenseWithholdingTaxCertificate.expense_request_id == req.id
    ))).scalar_one_or_none()
    if existing:
        return existing
    # request_no is normally "EXP-YYYYMM-NNNNNN", but an installment document carries
    # a trailing "-<installment_no>" suffix — strip that before taking the sequence
    # digits, and re-append the installment number so each sibling still gets its
    # own unique certificate number.
    base_request_no = re.sub(r"-\d+$", "", req.request_no or "") if req.installment_chain_root_id else (req.request_no or "")
    installment_suffix = f"-{req.installment_no}" if req.installment_no else ""
    cert_no = f"WHT-{datetime.now().strftime('%Y%m')}-{base_request_no[-6:]}{installment_suffix}"
    taxpayer_id = decrypt_account_number(req.recipient_tax_id_encrypted) or req.taxpayer_id or "-"
    html_doc = f"""<html lang='th'><meta charset='utf-8'><style>@page{{size:A4;margin:20mm}}body{{font-family:'Noto Looped Thai';font-size:13px}}h1{{text-align:center}}table{{width:100%;border-collapse:collapse}}td{{border:1px solid #555;padding:9px}}</style><body><h1>หนังสือรับรองการหักภาษี ณ ที่จ่าย</h1><p>เลขที่ {html.escape(cert_no)}</p><table><tr><td>ผู้ถูกหักภาษี</td><td>{html.escape(req.taxpayer_name or req.recipient_name or '-')}</td></tr><tr><td>เลขประจำตัวผู้เสียภาษี</td><td>{html.escape(taxpayer_id)}</td></tr><tr><td>ฐานภาษี</td><td>{money(req.price_before_vat):,.2f} บาท</td></tr><tr><td>อัตรา</td><td>{money(req.withholding_rate)}%</td></tr><tr><td>ภาษีที่หัก</td><td>{money(req.withholding_amount):,.2f} บาท</td></tr></table></body></html>"""
    data = HTML(string=html_doc).write_pdf()
    path, digest = _store(req.id, "wht", f"{cert_no}.pdf", data)
    cert = ExpenseWithholdingTaxCertificate(
        company_id=req.company_id, expense_request_id=req.id, certificate_no=cert_no,
        tax_rate=req.withholding_rate, base_amount=req.price_before_vat,
        tax_amount=req.withholding_amount, file_path=path, sha256=digest, issued_by=actor_user_id,
    )
    db.add(cert)
    add_history(db, req, "wht_certificate_issued", actor_user_id,
                snapshot={"certificate_no": cert_no, "sha256": digest})
    await db.commit()
    await db.refresh(cert)
    return cert


async def create_due_notifications(db: AsyncSession, today: date | None = None) -> int:
    today = today or datetime.now(ZoneInfo("Asia/Bangkok")).date()
    rows = (await db.execute(select(ExpenseRequest).where(
        ExpenseRequest.status == "settlement_due",
        ExpenseRequest.settlement_due_date.is_not(None),
        ExpenseRequest.settlement_due_date <= today + timedelta(days=3),
    ))).scalars().all()
    created = 0
    for req in rows:
        days = (req.settlement_due_date - today).days
        kind = "settlement_overdue" if days < 0 else ("settlement_due" if days == 0 else "settlement_due_soon")
        result = await db.execute(insert(SystemNotification).values(
            company_id=req.company_id, user_id=req.requester_user_id, expense_request_id=req.id,
            type=kind, title="แจ้งเตือนเคลียร์เงินทดรอง",
            message=f"{req.request_no} ครบกำหนด {req.settlement_due_date.strftime('%d/%m/%Y')}",
            action_url=f"/expense-requests/{req.id}",
            dedupe_key=f"settlement-reminder:{req.id}:{today.isoformat()}:{kind}",
        ).on_conflict_do_nothing(index_elements=["user_id", "dedupe_key"]))
        created += result.rowcount or 0
    await db.commit()
    return created
