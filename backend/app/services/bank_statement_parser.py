"""Parse common Thai bank Statement files into normalized signed transactions."""
from __future__ import annotations

import csv
import hashlib
import io
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable


@dataclass(frozen=True)
class ParsedBankTransaction:
    transaction_date: date
    description: str
    amount: float
    transaction_time: str | None = None
    reference: str | None = None
    channel: str | None = None
    row_hash: str | None = None


@dataclass(frozen=True)
class ParsedBankStatement:
    transactions: list[ParsedBankTransaction]
    processing_method: str


HEADER_ALIASES = {
    "date": {
        "date", "transactiondate", "postingdate", "วันที่", "วันที่ทำรายการ",
        "วันที่รายการ", "วันเดือนปี",
    },
    "description": {
        "description", "details", "detail", "transactiondescription",
        "trdescription", "รายการ", "รายละเอียด", "คำอธิบาย",
    },
    "amount": {"amount", "transactionamount", "ยอด", "ยอดเงิน", "จำนวนเงิน"},
    "debit": {"debit", "withdrawal", "withdraw", "เดบิต", "ถอน", "ยอดถอน", "เงินออก"},
    "credit": {"credit", "deposit", "เครดิต", "ฝาก", "ยอดฝาก", "เงินเข้า"},
    "deposit": {"depositamount", "ยอดเงินเข้า", "ยอดรับ", "รับเงิน"},
    "withdraw": {"withdrawamount", "ยอดเงินออก", "ถอนเงิน"},
    "time": {"time", "transactiontime", "เวลา", "เวลาทำรายการ", "เวลารายการ"},
    "channel": {"channel", "ช่องทาง", "ช่องทางรายการ"},
    "reference": {
        "reference", "ref", "transactionreference", "trcode",
        "transactioncode", "code", "เลขอ้างอิง", "รหัสรายการ",
    },
}


def _normalise_header(value: Any) -> str:
    return re.sub(r"[\s_\-./()]+", "", str(value or "").strip().lower())


def _column_map(row: Iterable[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    normalised = [_normalise_header(value) for value in row]
    for field, aliases in HEADER_ALIASES.items():
        if field == "description":
            for preferred in (
                "description", "รายละเอียด", "คำอธิบาย", "รายการ",
                "transactiondescription", "trdescription", "details", "detail",
            ):
                if preferred in normalised:
                    result[field] = normalised.index(preferred)
                    break
            if field in result:
                continue
        for index, value in enumerate(normalised):
            if value in aliases:
                result[field] = index
                break
    return result


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    for index, row in enumerate(rows[:30]):
        mapping = _column_map(row)
        if "date" in mapping and "description" in mapping and any(
            field in mapping for field in ("amount", "debit", "credit", "deposit", "withdraw")
        ):
            return index, mapping
    return None


def parse_amount(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    negative = (
        (text.startswith("(") and text.endswith(")"))
        or text.endswith("-")
        or bool(re.search(r"\bDR\b", text, re.IGNORECASE))
    )
    cleaned = re.sub(r"[^\d.,+\-]", "", text)
    if not cleaned:
        return None
    if cleaned.count(",") and not cleaned.count("."):
        tail = cleaned.rsplit(",", 1)[-1]
        cleaned = cleaned.replace(",", ".") if len(tail) == 2 else cleaned.replace(",", "")
    else:
        cleaned = cleaned.replace(",", "")
    try:
        amount = float(cleaned.rstrip("-"))
    except ValueError:
        return None
    return -abs(amount) if negative else amount


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and 20_000 < float(value) < 80_000:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    text = re.sub(r"\s+.*$", "", str(value or "").strip())
    for fmt in (
        "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y",
        "%d-%m-%y", "%d.%m.%Y", "%d.%m.%y", "%m/%d/%Y",
    ):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return parsed.replace(year=parsed.year - 543) if parsed.year > 2400 else parsed
        except ValueError:
            continue
    return None


def parse_time(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        seconds = int(round(float(value) * 86400))
        return f"{seconds // 3600 % 24:02d}:{seconds % 3600 // 60:02d}:{seconds % 60:02d}"
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).strftime("%H:%M:%S")
        except ValueError:
            continue
    return None


def make_row_hash(
    transaction_date: date,
    transaction_time: str | None,
    amount: float,
    description: str,
    reference: str | None,
) -> str:
    basis = "|".join([
        transaction_date.isoformat(),
        transaction_time or "",
        f"{amount:.2f}",
        re.sub(r"\s+", " ", description.strip()).lower(),
        (reference or "").strip().lower(),
    ])
    return hashlib.sha256(basis.encode("utf-8")).hexdigest()


def parse_rows(rows: list[list[Any]]) -> list[ParsedBankTransaction]:
    header = _find_header(rows)
    if not header:
        raise ValueError(
            "ไม่พบหัวตาราง กรุณาใช้คอลัมน์ วันที่ รายละเอียด และจำนวนเงิน "
            "หรือคอลัมน์เงินเข้า/เงินออก"
        )
    header_index, columns = header
    result: list[ParsedBankTransaction] = []

    for row in rows[header_index + 1:]:
        if not row or all(value in (None, "") for value in row):
            continue

        def cell(field: str) -> Any:
            index = columns.get(field)
            return row[index] if index is not None and index < len(row) else None

        transaction_date = parse_date(cell("date"))
        description = str(cell("description") or "").strip()
        if "amount" in columns:
            amount = parse_amount(cell("amount"))
        else:
            deposit = parse_amount(cell("deposit"))
            credit = parse_amount(cell("credit"))
            withdraw = parse_amount(cell("withdraw"))
            debit = parse_amount(cell("debit"))
            incoming = deposit if deposit not in (None, 0) else credit
            outgoing = withdraw if withdraw not in (None, 0) else debit
            if incoming not in (None, 0):
                amount = abs(incoming)
            elif outgoing not in (None, 0):
                amount = -abs(outgoing)
            else:
                amount = None
        if not transaction_date or not description or amount in (None, 0):
            continue
        amount = round(float(amount), 2)
        transaction_time = parse_time(cell("time"))
        reference = str(cell("reference") or "").strip()[:200] or None
        channel = str(cell("channel") or "").strip()[:100] or None
        result.append(ParsedBankTransaction(
            transaction_date=transaction_date,
            transaction_time=transaction_time,
            description=description[:1000],
            reference=reference,
            channel=channel,
            amount=amount,
            row_hash=make_row_hash(
                transaction_date, transaction_time, amount, description, reference
            ),
        ))
    if not result:
        raise ValueError("อ่านไฟล์ได้แต่ไม่พบรายการที่มีวันที่ รายละเอียด และจำนวนเงินครบ")
    return result


def _decode_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp874", "tis-620"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("ไม่สามารถอ่าน encoding ของไฟล์ CSV ได้")


def parse_csv(data: bytes) -> list[ParsedBankTransaction]:
    text = _decode_csv(data)
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return parse_rows([list(row) for row in csv.reader(io.StringIO(text), dialect)])


def parse_xlsx(data: bytes) -> list[ParsedBankTransaction]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    errors: list[str] = []
    for sheet in workbook.worksheets:
        try:
            return parse_rows([list(row) for row in sheet.iter_rows(values_only=True)])
        except ValueError as exc:
            errors.append(f"{sheet.title}: {exc}")
    raise ValueError("ไม่พบตารางรายการในไฟล์ Excel — " + "; ".join(errors))


PDF_LINE_PATTERN = re.compile(
    r"^\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{1,2}-\d{1,2})\s+"
    r"(.+?)\s+(\(?[-+]?(?:\d{1,3}(?:,\d{3})*|\d+)(?:\.\d{2})?\)?(?:\s*DR)?)\s*$",
    re.IGNORECASE,
)

OCR_INCOMING_WORDS = (
    "เงินเข้า", "รับโอน", "รับเงิน", "ฝาก", "เครดิต", "credit", "deposit",
    "transfer in", "incoming",
)
OCR_OUTGOING_WORDS = (
    "เงินออก", "โอนออก", "จ่าย", "ชำระ", "ถอน", "เดบิต", "debit", "withdraw",
    "payment", "transfer out",
)
OCR_AMOUNT_PATTERN = re.compile(
    r"\(?[-+]?(?:\d{1,3}(?:,\d{3})*|\d+)(?:\.\d{2})\)?(?:\s*DR)?",
    re.IGNORECASE,
)


def parse_ocr_text(text: str) -> list[ParsedBankTransaction]:
    """Parse OCR output conservatively so a running balance is not imported."""
    result: list[ParsedBankTransaction] = []
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        date_match = re.match(
            r"^(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{1,2}-\d{1,2})\s+(.+)$",
            line,
        )
        if not date_match:
            continue
        transaction_date = parse_date(date_match.group(1))
        remainder = date_match.group(2)
        amounts = list(OCR_AMOUNT_PATTERN.finditer(remainder))
        if not transaction_date or not amounts:
            continue

        lowered = remainder.lower()
        incoming = any(word in lowered for word in OCR_INCOMING_WORDS)
        outgoing = any(word in lowered for word in OCR_OUTGOING_WORDS)
        if incoming == outgoing and len(amounts) > 1:
            continue

        # The last numeric column is commonly the running balance.
        amount_match = amounts[-2] if len(amounts) > 1 else amounts[-1]
        amount = parse_amount(amount_match.group(0))
        if amount in (None, 0):
            continue
        if incoming:
            amount = abs(amount)
        elif outgoing:
            amount = -abs(amount)

        description = (
            remainder[:amount_match.start()] + remainder[amount_match.end():]
        ).strip(" -|")
        if len(amounts) > 1:
            description = description.replace(amounts[-1].group(0), "").strip(" -|")
        description = description or "รายการจาก Statement (OCR)"
        result.append(ParsedBankTransaction(
            transaction_date=transaction_date,
            description=description[:1000],
            amount=round(float(amount), 2),
            row_hash=make_row_hash(
                transaction_date, None, float(amount), description, None
            ),
        ))
    if not result:
        raise ValueError(
            "OCR อ่านตัวอักษรได้ แต่ยังแยกวันที่ รายละเอียด และยอดเงินไม่ได้ "
            "กรุณาใช้ PDF ต้นฉบับ หรือส่งตัวอย่างให้ผู้ดูแลเพิ่มรูปแบบธนาคาร"
        )
    return result


def _run_tesseract(image_path: Path) -> str:
    if not shutil.which("tesseract"):
        raise ValueError("เซิร์ฟเวอร์ยังไม่ได้ติดตั้งระบบ OCR")
    try:
        result = subprocess.run(
            ["tesseract", str(image_path), "stdout", "-l", "tha+eng", "--psm", "6"],
            check=True,
            capture_output=True,
            text=True,
            timeout=90,
        )
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
        raise ValueError("ระบบ OCR ไม่สามารถอ่านรูปภาพนี้ได้") from exc
    return result.stdout


def parse_image_ocr(data: bytes, extension: str) -> list[ParsedBankTransaction]:
    with tempfile.TemporaryDirectory(prefix="statement-ocr-") as temp_dir:
        image_path = Path(temp_dir) / f"statement{extension}"
        image_path.write_bytes(data)
        return parse_ocr_text(_run_tesseract(image_path))


def parse_pdf_ocr(data: bytes) -> list[ParsedBankTransaction]:
    if not shutil.which("pdftoppm"):
        raise ValueError("เซิร์ฟเวอร์ยังไม่ได้ติดตั้งตัวแปลง PDF สำหรับ OCR")
    with tempfile.TemporaryDirectory(prefix="statement-pdf-ocr-") as temp_dir:
        pdf_path = Path(temp_dir) / "statement.pdf"
        output_prefix = Path(temp_dir) / "page"
        pdf_path.write_bytes(data)
        try:
            subprocess.run(
                [
                    "pdftoppm", "-png", "-r", "200", "-f", "1", "-l", "20",
                    str(pdf_path), str(output_prefix),
                ],
                check=True,
                capture_output=True,
                timeout=120,
            )
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
            raise ValueError("ไม่สามารถแปลง PDF สแกนเพื่อทำ OCR ได้") from exc
        images = sorted(Path(temp_dir).glob("page-*.png"))
        if not images:
            raise ValueError("PDF ไม่มีหน้าที่ระบบสามารถอ่านได้")
        return parse_ocr_text("\n".join(_run_tesseract(image) for image in images))


def parse_pdf(data: bytes) -> list[ParsedBankTransaction]:
    import pdfplumber

    fallback: list[ParsedBankTransaction] = []
    with pdfplumber.open(io.BytesIO(data)) as document:
        for page in document.pages:
            for table in page.extract_tables() or []:
                try:
                    parsed = parse_rows([list(row) for row in table if row])
                    if parsed:
                        return parsed
                except ValueError:
                    pass
            for line in (page.extract_text() or "").splitlines():
                match = PDF_LINE_PATTERN.match(line)
                if not match:
                    continue
                transaction_date = parse_date(match.group(1))
                amount = parse_amount(match.group(3))
                description = match.group(2).strip()
                if transaction_date and amount not in (None, 0) and description:
                    fallback.append(ParsedBankTransaction(
                        transaction_date=transaction_date,
                        description=description[:1000],
                        amount=round(float(amount), 2),
                        row_hash=make_row_hash(
                            transaction_date, None, float(amount), description, None
                        ),
                    ))
    if fallback:
        return fallback
    raise ValueError("ไม่สามารถแยกรายการจาก PDF นี้ได้")


def parse_bank_statement_with_metadata(filename: str, data: bytes) -> ParsedBankStatement:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        return ParsedBankStatement(parse_csv(data), "csv")
    if extension in {".xlsx", ".xlsm"}:
        return ParsedBankStatement(parse_xlsx(data), "spreadsheet")
    if extension == ".pdf":
        try:
            return ParsedBankStatement(parse_pdf(data), "pdf_text")
        except ValueError:
            return ParsedBankStatement(parse_pdf_ocr(data), "ocr")
    if extension in {".png", ".jpg", ".jpeg"}:
        return ParsedBankStatement(parse_image_ocr(data, extension), "ocr")
    raise ValueError("รองรับไฟล์ .csv, .xlsx, .xlsm, .pdf, .png, .jpg และ .jpeg")


def parse_bank_statement(filename: str, data: bytes) -> list[ParsedBankTransaction]:
    return parse_bank_statement_with_metadata(filename, data).transactions
