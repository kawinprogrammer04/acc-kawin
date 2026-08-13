"""crm-kawin compatible income/expense statement and tax-invoice tracking.

The API preserves the legacy table/field vocabulary while enforcing acc-kawin
authentication, company tenancy, and user ownership.
"""
from __future__ import annotations

import csv
import io
import json
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Literal, Optional

import os
from pathlib import Path
from urllib.parse import quote
from uuid import uuid4

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from pydantic import BaseModel, Field, field_validator, model_validator
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.dependencies import (
    get_current_company,
    get_current_user,
    require_accountant,
    require_viewer,
)
from app.models.company import Company
from app.core.config import settings
from app.models.crm_cashflow import (
    CrmCashflowCategory,
    CrmCashflowDepartment,
    CrmCashflowImportTemplate,
    CrmCashflowList,
    CrmCashflowStatement,
    CrmCashflowStatementAttachment,
)
from app.models.user import User


router = APIRouter(prefix="/crm-cashflow", tags=["CRM Cashflow"])

DocumentType = Literal["tax_invoice", "cash_bill", "other"]
VerificationStatus = Literal["pending", "verified"]
InvoiceStatus = Literal[
    "none", "pending", "received", "tax_invoice", "cash_bill", "other",
]
DOCUMENT_TYPE_LABELS: dict[str, str] = {
    "tax_invoice": "ใบกำกับภาษี",
    "cash_bill": "บิลเงินสด",
    "other": "อื่นๆ",
}
LEGACY_INVOICE_LABELS = {None: "", 0: "รอใบกำกับ", 1: "ได้รับแล้ว"}

# ── Import template field vocabulary ────────────────────────────────────────
# A custom import template maps an Excel column reference in the uploaded file
# to one of these keys.  Older saved templates without ``column`` remain
# position-based for backwards compatibility.
ImportFieldKey = Literal[
    "date", "category", "source", "detail", "refrain",
    "income", "expense", "amount", "ref", "department", "skip",
]
IMPORT_FIELD_LABELS: dict[str, str] = {
    "date": "วันที่",
    "category": "หัวข้อ",
    "source": "แหล่งที่มา",
    "detail": "รายละเอียด",
    "refrain": "คำนวณต้นทุน",
    "income": "รายรับ",
    "expense": "รายจ่าย",
    "amount": "จำนวนเงิน (+รับ / -จ่าย)",
    "ref": "Ref",
    "department": "แผนก",
    "skip": "ข้าม (ไม่ใช้คอลัมน์นี้)",
}
IMPORT_REQUIRED_FIELDS = {"date"}


# ── Request schemas ──────────────────────────────────────────────────────────
class CategoryCreate(BaseModel):
    cfcat_name: str = Field(min_length=1, max_length=255)

    @field_validator("cfcat_name")
    @classmethod
    def clean_name(cls, value: str) -> str:
        return value.strip()


class CategoryUpdate(BaseModel):
    cfcat_name: Optional[str] = Field(default=None, min_length=1, max_length=255)
    cfcat_status: Optional[int] = None

    @field_validator("cfcat_name")
    @classmethod
    def clean_name(cls, value: Optional[str]) -> Optional[str]:
        return value.strip() if value is not None else value

    @field_validator("cfcat_status")
    @classmethod
    def valid_status(cls, value: Optional[int]) -> Optional[int]:
        if value is not None and value not in (0, 1):
            raise ValueError("สถานะต้องเป็น 0 หรือ 1")
        return value


class SourceCreate(BaseModel):
    cflist_name: str = Field(min_length=1, max_length=255)
    cfcat_id: int

    @field_validator("cflist_name")
    @classmethod
    def clean_name(cls, value: str) -> str:
        return value.strip()


class SourceUpdate(BaseModel):
    cflist_name: Optional[str] = Field(default=None, min_length=1, max_length=255)
    cflist_status: Optional[int] = None
    cflist_hide: Optional[int] = None

    @field_validator("cflist_name")
    @classmethod
    def clean_name(cls, value: Optional[str]) -> Optional[str]:
        return value.strip() if value is not None else value

    @field_validator("cflist_status")
    @classmethod
    def valid_status(cls, value: Optional[int]) -> Optional[int]:
        if value is not None and value not in (0, 1):
            raise ValueError("สถานะต้องเป็น 0 หรือ 1")
        return value


class SourceMove(BaseModel):
    new_cfcat_id: int
    new_cflist_id: Optional[int] = None
    new_list_name: Optional[str] = Field(default=None, max_length=255)

    @model_validator(mode="after")
    def require_target(self):
        if self.new_cflist_id is None and not (self.new_list_name or "").strip():
            raise ValueError("ต้องเลือกแหล่งที่มาปลายทางหรือระบุชื่อใหม่")
        if self.new_cflist_id is not None and self.new_list_name:
            raise ValueError("เลือกแหล่งที่มาปลายทางได้เพียงแบบเดียว")
        return self


class DepartmentCreate(BaseModel):
    cfstate_dep_name: str = Field(min_length=1, max_length=255)

    @field_validator("cfstate_dep_name")
    @classmethod
    def clean_name(cls, value: str) -> str:
        return value.strip()


class DepartmentUpdate(BaseModel):
    cfstate_dep_name: Optional[str] = Field(default=None, min_length=1, max_length=255)
    cfstate_dep_status: Optional[int] = None

    @field_validator("cfstate_dep_name")
    @classmethod
    def clean_name(cls, value: Optional[str]) -> Optional[str]:
        return value.strip() if value is not None else value

    @field_validator("cfstate_dep_status")
    @classmethod
    def valid_status(cls, value: Optional[int]) -> Optional[int]:
        if value is not None and value not in (0, 1):
            raise ValueError("สถานะต้องเป็น 0 หรือ 1")
        return value


class StatementCreate(BaseModel):
    cfstate_date: date
    cfcat_id: int
    cflist_id: int
    cfstate_dep_id: Optional[int] = None
    cfstate_invoice: Optional[int] = None
    cfstate_refrain: int = 1
    cfstate_detail: Optional[str] = None
    cfstate_amount: Decimal
    cfstate_ref: Optional[str] = Field(default=None, max_length=255)

    @field_validator("cfstate_invoice")
    @classmethod
    def valid_invoice(cls, value: Optional[int]) -> Optional[int]:
        if value is not None and value not in (0, 1):
            raise ValueError("สถานะใบกำกับภาษีต้องเป็น null, 0 หรือ 1")
        return value

    @field_validator("cfstate_refrain")
    @classmethod
    def valid_refrain(cls, value: int) -> int:
        if value not in (0, 1):
            raise ValueError("ค่าคำนวณต้นทุนต้องเป็น 0 หรือ 1")
        return value


class StatementBatchCreate(BaseModel):
    items: list[StatementCreate] = Field(min_length=1, max_length=500)
    duplicate_action: str = Field(default="skip", pattern="^(skip|update|create)$")


class StatementFlagsUpdate(BaseModel):
    cfstate_invoice: Optional[int] = None
    cfstate_refrain: Optional[int] = None
    cfstate_verified: Optional[int] = None
    cfstate_document_type: Optional[DocumentType] = None

    @model_validator(mode="after")
    def validate_flags(self):
        editable_fields = {
            "cfstate_invoice", "cfstate_refrain", "cfstate_verified", "cfstate_document_type",
        }
        if not self.model_fields_set.intersection(editable_fields):
            raise ValueError("ไม่มีข้อมูลที่ต้องแก้ไข")
        if self.cfstate_invoice is not None and self.cfstate_invoice not in (0, 1):
            raise ValueError("สถานะใบกำกับภาษีต้องเป็น 0 หรือ 1")
        if self.cfstate_refrain is not None and self.cfstate_refrain not in (0, 1):
            raise ValueError("ค่าคำนวณต้นทุนต้องเป็น 0 หรือ 1")
        if self.cfstate_verified is not None and self.cfstate_verified not in (0, 1):
            raise ValueError("สถานะตรวจสอบแล้วต้องเป็น 0 หรือ 1")
        return self


class ErrorWorkbookPayload(BaseModel):
    error_rows: list[list[Any]]
    error_details: list[str]


class ImportTemplateColumn(BaseModel):
    field: ImportFieldKey
    label: str = Field(min_length=1, max_length=100)
    column: Optional[str] = None

    @field_validator("label")
    @classmethod
    def clean_label(cls, value: str) -> str:
        return value.strip()

    @field_validator("column")
    @classmethod
    def clean_column(cls, value: Optional[str]) -> Optional[str]:
        if value is None:
            return None
        normalized = value.strip().upper()
        if not normalized:
            return None
        if not re.fullmatch(r"[A-Z]+", normalized):
            raise ValueError("ชื่อคอลัมน์ต้องเป็นตัวอักษร เช่น A, J หรือ AO")
        try:
            column_index = column_index_from_string(normalized)
        except ValueError as exc:
            raise ValueError("ชื่อคอลัมน์ต้องอยู่ระหว่าง A ถึง XFD") from exc
        if column_index > 16_384:
            raise ValueError("ชื่อคอลัมน์ต้องอยู่ระหว่าง A ถึง XFD")
        return normalized


def _validate_import_template_columns(columns: list[ImportTemplateColumn]) -> None:
    # Legacy templates did not store ``column`` and are positional. New
    # templates always store the key; a null value explicitly means that the
    # field is not mapped and must remain blank during import.
    fields = [
        column.field for column in columns
        if column.field != "skip" and (
            "column" not in column.model_fields_set or column.column is not None
        )
    ]
    missing = IMPORT_REQUIRED_FIELDS - set(fields)
    if missing:
        names = "', '".join(IMPORT_FIELD_LABELS[key] for key in missing)
        raise ValueError(f"เทมเพลตต้องมีคอลัมน์ '{names}'")
    if len(fields) != len(set(fields)):
        raise ValueError("แต่ละประเภทข้อมูลกำหนดได้เพียงคอลัมน์เดียว (ยกเว้นคอลัมน์ที่ข้าม)")
    mapped_columns = [column.column for column in columns if column.field != "skip" and column.column]
    if len(mapped_columns) != len(set(mapped_columns)):
        raise ValueError("แต่ละชื่อคอลัมน์ในไฟล์กำหนดได้เพียงครั้งเดียว")
    has_split_amount = "income" in fields or "expense" in fields
    if "amount" not in fields and not has_split_amount:
        raise ValueError(
            "เทมเพลตต้องมีคอลัมน์จำนวนเงินอย่างน้อยหนึ่งแบบ (รายรับ/รายจ่าย หรือ จำนวนเงินรวม)"
        )
    if "amount" in fields and has_split_amount:
        raise ValueError(
            "เลือกได้อย่างใดอย่างหนึ่ง: แยกรายรับ/รายจ่าย หรือ จำนวนเงินรวม ไม่ใช่ทั้งสองแบบ"
        )


class ImportTemplateCreate(BaseModel):
    cfimptpl_name: str = Field(min_length=1, max_length=255)
    cfimptpl_header_row: bool = True
    cfimptpl_columns: list[ImportTemplateColumn] = Field(min_length=1)

    @field_validator("cfimptpl_name")
    @classmethod
    def clean_name(cls, value: str) -> str:
        return value.strip()

    @model_validator(mode="after")
    def validate_columns(self):
        _validate_import_template_columns(self.cfimptpl_columns)
        return self


class ImportTemplateUpdate(BaseModel):
    cfimptpl_name: Optional[str] = Field(default=None, min_length=1, max_length=255)
    cfimptpl_header_row: Optional[bool] = None
    cfimptpl_columns: Optional[list[ImportTemplateColumn]] = Field(default=None, min_length=1)
    cfimptpl_status: Optional[int] = None

    @field_validator("cfimptpl_name")
    @classmethod
    def clean_name(cls, value: Optional[str]) -> Optional[str]:
        return value.strip() if value is not None else value

    @model_validator(mode="after")
    def validate_columns(self):
        if self.cfimptpl_columns is not None:
            _validate_import_template_columns(self.cfimptpl_columns)
        return self


# ── Shared query/helpers ─────────────────────────────────────────────────────
async def _owned_row(db: AsyncSession, model, pk_column, row_id: int, comp_id: int, message: str):
    row = (
        await db.execute(
            select(model).where(pk_column == row_id, model.comp_id == comp_id)
        )
    ).scalar_one_or_none()
    if row is None:
        raise HTTPException(404, message)
    return row


async def _ensure_statement_references(
    db: AsyncSession, item: StatementCreate, comp_id: int
) -> None:
    category = await _owned_row(
        db, CrmCashflowCategory, CrmCashflowCategory.cfcat_id,
        item.cfcat_id, comp_id, "ไม่พบหัวข้อ"
    )
    if category.cfcat_status != 1:
        raise HTTPException(409, "หัวข้อนี้ถูกปิดใช้งาน")

    source = await _owned_row(
        db, CrmCashflowList, CrmCashflowList.cflist_id,
        item.cflist_id, comp_id, "ไม่พบแหล่งที่มา"
    )
    if source.cflist_status != 1 or source.cfcat_id != item.cfcat_id:
        raise HTTPException(409, "แหล่งที่มาไม่ตรงกับหัวข้อหรือถูกปิดใช้งาน")

    if item.cfstate_dep_id is not None:
        department = await _owned_row(
            db, CrmCashflowDepartment, CrmCashflowDepartment.cfstate_dep_id,
            item.cfstate_dep_id, comp_id, "ไม่พบแผนก"
        )
        if department.cfstate_dep_status != 1:
            raise HTTPException(409, "แผนกนี้ถูกปิดใช้งาน")


async def _find_duplicate_statement(
    db: AsyncSession,
    comp_id: int,
    cfstate_date: date,
    cfcat_id: int,
    cflist_id: int,
    cfstate_dep_id: Optional[int],
    cfstate_invoice: Optional[int],
    cfstate_refrain: int,
    cfstate_detail: Optional[str],
    cfstate_amount: Decimal,
    cfstate_ref: Optional[str],
) -> Optional[int]:
    """Find an existing statement matching **all** data columns.

    Returns the existing ``cfstate_id`` when a duplicate is found, otherwise
    ``None``.  Empty strings are normalised to ``None`` so that an empty
    detail/ref is treated the same as a NULL in the database.
    """
    detail = cfstate_detail or None
    ref = cfstate_ref or None
    stmt = select(CrmCashflowStatement.cfstate_id).where(
        CrmCashflowStatement.comp_id == comp_id,
        CrmCashflowStatement.cfstate_status == 1,
        CrmCashflowStatement.cfstate_date == cfstate_date,
        CrmCashflowStatement.cfcat_id == cfcat_id,
        CrmCashflowStatement.cflist_id == cflist_id,
        CrmCashflowStatement.cfstate_dep_id == cfstate_dep_id,
        CrmCashflowStatement.cfstate_invoice == cfstate_invoice,
        CrmCashflowStatement.cfstate_refrain == cfstate_refrain,
        CrmCashflowStatement.cfstate_detail == detail,
        CrmCashflowStatement.cfstate_amount == cfstate_amount,
        CrmCashflowStatement.cfstate_ref == ref,
    )
    return (await db.execute(stmt)).scalar_one_or_none()


def _statement_select():
    attachment_count = (
        select(func.count())
        .select_from(CrmCashflowStatementAttachment)
        .where(
            CrmCashflowStatementAttachment.cfstate_id == CrmCashflowStatement.cfstate_id,
            CrmCashflowStatementAttachment.comp_id == CrmCashflowStatement.comp_id,
        )
        .correlate(CrmCashflowStatement)
        .scalar_subquery()
        .label("attachment_count")
    )
    return (
        select(
            CrmCashflowStatement.cfstate_id,
            CrmCashflowStatement.cfstate_date,
            CrmCashflowStatement.cfcat_id,
            CrmCashflowStatement.cflist_id,
            CrmCashflowStatement.user_id,
            CrmCashflowStatement.comp_id,
            CrmCashflowStatement.cfstate_amount,
            CrmCashflowStatement.cfstate_refrain,
            CrmCashflowStatement.cfstate_invoice,
            CrmCashflowStatement.cfstate_document_type,
            CrmCashflowStatement.cfstate_verified,
            CrmCashflowStatement.cfstate_detail,
            CrmCashflowStatement.cfstate_status,
            CrmCashflowStatement.cfstate_dep_id,
            CrmCashflowStatement.cfstate_ref,
            CrmCashflowCategory.cfcat_name,
            CrmCashflowList.cflist_name,
            CrmCashflowDepartment.cfstate_dep_name,
            attachment_count,
            User.full_name.label("user_name"),
            User.username.label("username"),
        )
        .join(CrmCashflowCategory, CrmCashflowCategory.cfcat_id == CrmCashflowStatement.cfcat_id)
        .join(CrmCashflowList, CrmCashflowList.cflist_id == CrmCashflowStatement.cflist_id)
        .join(User, User.id == CrmCashflowStatement.user_id)
        .outerjoin(
            CrmCashflowDepartment,
            CrmCashflowDepartment.cfstate_dep_id == CrmCashflowStatement.cfstate_dep_id,
        )
    )


async def _list_statements(
    db: AsyncSession,
    comp_id: int,
    start_date: Optional[date],
    end_date: Optional[date],
    cfcat_id: Optional[int],
    pending_verification_only: bool = False,
    verification_status: Optional[VerificationStatus] = None,
    invoice_status: Optional[InvoiceStatus] = None,
):
    stmt = _statement_select().where(
        CrmCashflowStatement.comp_id == comp_id,
        CrmCashflowStatement.cfstate_status == 1,
    )
    if start_date is not None:
        stmt = stmt.where(CrmCashflowStatement.cfstate_date >= start_date)
    if end_date is not None:
        stmt = stmt.where(CrmCashflowStatement.cfstate_date <= end_date)
    if cfcat_id is not None:
        stmt = stmt.where(CrmCashflowStatement.cfcat_id == cfcat_id)
    if verification_status is not None:
        stmt = stmt.where(
            CrmCashflowStatement.cfstate_verified
            == (1 if verification_status == "verified" else 0)
        )
    if invoice_status in DOCUMENT_TYPE_LABELS:
        stmt = stmt.where(CrmCashflowStatement.cfstate_document_type == invoice_status)
    elif invoice_status is not None:
        legacy_invoice_value = {"none": None, "pending": 0, "received": 1}[invoice_status]
        stmt = stmt.where(CrmCashflowStatement.cfstate_document_type.is_(None))
        if legacy_invoice_value is None:
            stmt = stmt.where(CrmCashflowStatement.cfstate_invoice.is_(None))
        else:
            stmt = stmt.where(CrmCashflowStatement.cfstate_invoice == legacy_invoice_value)
    if pending_verification_only:
        # Invoice tracking is only relevant to money paid out. Keep every
        # "ใบเสร็จ" status (0/1/null), but exclude positive receipt amounts.
        stmt = stmt.where(
            CrmCashflowStatement.cfstate_verified == 0,
            CrmCashflowStatement.cfstate_amount <= 0,
        )
    stmt = stmt.order_by(
        CrmCashflowStatement.cfstate_date.desc(),
        CrmCashflowStatement.cfstate_id.desc(),
    )
    return (await db.execute(stmt)).mappings().all()


def _serialize_statement(row) -> dict[str, Any]:
    data = dict(row)
    data["cfstate_amount"] = float(data["cfstate_amount"])
    data["cfstate_amount_str"] = f'{data["cfstate_amount"]:,.2f}'
    data["user_name"] = data.get("user_name") or data.get("username")
    return data


def _invoice_status_label(
    cfstate_invoice: Optional[int], cfstate_document_type: Optional[str]
) -> str:
    if cfstate_document_type in DOCUMENT_TYPE_LABELS:
        return DOCUMENT_TYPE_LABELS[cfstate_document_type]
    return LEGACY_INVOICE_LABELS.get(cfstate_invoice, "")


# ── Master data: categories ─────────────────────────────────────────────────
@router.get("/categories", dependencies=[Depends(require_viewer)])
async def list_categories(
    include_inactive: bool = False,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    stmt = select(CrmCashflowCategory).where(CrmCashflowCategory.comp_id == company.id)
    if not include_inactive:
        stmt = stmt.where(CrmCashflowCategory.cfcat_status == 1)
    rows = (await db.execute(stmt.order_by(CrmCashflowCategory.cfcat_id))).scalars().all()
    return [
        {"cfcat_id": row.cfcat_id, "cfcat_name": row.cfcat_name,
         "cfcat_status": row.cfcat_status, "comp_id": row.comp_id}
        for row in rows
    ]


@router.post("/categories", dependencies=[Depends(require_accountant)])
async def create_category(
    payload: CategoryCreate,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    existing = (
        await db.execute(
            select(CrmCashflowCategory).where(
                CrmCashflowCategory.comp_id == company.id,
                func.lower(CrmCashflowCategory.cfcat_name) == payload.cfcat_name.lower(),
            )
        )
    ).scalar_one_or_none()
    if existing:
        if existing.cfcat_status == 0:
            existing.cfcat_status = 1
            await db.commit()
            return {"cfcat_id": existing.cfcat_id, "cfcat_name": existing.cfcat_name,
                    "cfcat_status": existing.cfcat_status, "comp_id": existing.comp_id}
        raise HTTPException(409, "มีชื่อหัวข้อนี้อยู่แล้ว")
    row = CrmCashflowCategory(
        cfcat_name=payload.cfcat_name, cfcat_status=1, comp_id=company.id
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return {"cfcat_id": row.cfcat_id, "cfcat_name": row.cfcat_name,
            "cfcat_status": row.cfcat_status, "comp_id": row.comp_id}


@router.patch("/categories/{category_id}", dependencies=[Depends(require_accountant)])
async def update_category(
    category_id: int,
    payload: CategoryUpdate,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    row = await _owned_row(
        db, CrmCashflowCategory, CrmCashflowCategory.cfcat_id,
        category_id, company.id, "ไม่พบหัวข้อ"
    )
    if payload.cfcat_name is not None and payload.cfcat_name.lower() != row.cfcat_name.lower():
        duplicate = (
            await db.execute(
                select(CrmCashflowCategory.cfcat_id).where(
                    CrmCashflowCategory.comp_id == company.id,
                    func.lower(CrmCashflowCategory.cfcat_name) == payload.cfcat_name.lower(),
                    CrmCashflowCategory.cfcat_id != category_id,
                )
            )
        ).scalar_one_or_none()
        if duplicate:
            raise HTTPException(409, "มีชื่อหัวข้อนี้อยู่แล้ว")
        row.cfcat_name = payload.cfcat_name
    if payload.cfcat_status is not None:
        row.cfcat_status = payload.cfcat_status
    await db.commit()
    return {"cfcat_id": row.cfcat_id, "cfcat_name": row.cfcat_name,
            "cfcat_status": row.cfcat_status, "comp_id": row.comp_id}


@router.delete("/categories/{category_id}", dependencies=[Depends(require_accountant)])
async def delete_category(
    category_id: int,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    row = await _owned_row(
        db, CrmCashflowCategory, CrmCashflowCategory.cfcat_id,
        category_id, company.id, "ไม่พบหัวข้อ"
    )
    # A source under this category still holds a FK to it (even if disabled),
    # so a hard delete would fail until every source is removed first.
    source_count = (
        await db.execute(
            select(func.count()).select_from(CrmCashflowList).where(
                CrmCashflowList.comp_id == company.id,
                CrmCashflowList.cfcat_id == category_id,
            )
        )
    ).scalar_one()
    if source_count:
        raise HTTPException(
            409,
            f"ไม่สามารถลบได้ เนื่องจากยังมีแหล่งที่มาอยู่ภายใต้หัวข้อนี้ กรุณาลบแหล่งที่มาก่อน ({source_count} รายการ)",
        )
    usage_count = (
        await db.execute(
            select(func.count()).select_from(CrmCashflowStatement).where(
                CrmCashflowStatement.comp_id == company.id,
                CrmCashflowStatement.cfcat_id == category_id,
            )
        )
    ).scalar_one()
    if usage_count:
        raise HTTPException(
            409,
            f"ไม่สามารถลบได้ เนื่องจากมีการใช้งานอยู่ กรุณาย้ายข้อมูลก่อน ({usage_count} รายการ)",
        )
    await db.delete(row)
    await db.commit()
    return {"status": 1}


# ── Master data: sources/lists ───────────────────────────────────────────────
@router.get("/sources", dependencies=[Depends(require_viewer)])
async def list_sources(
    cfcat_id: Optional[int] = None,
    include_inactive: bool = False,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    stmt = (
        select(
            CrmCashflowList.cflist_id,
            CrmCashflowList.cflist_name,
            CrmCashflowList.cfcat_id,
            CrmCashflowList.cflist_status,
            CrmCashflowList.comp_id,
            CrmCashflowList.cflist_hide,
            CrmCashflowCategory.cfcat_name,
        )
        .join(CrmCashflowCategory, CrmCashflowCategory.cfcat_id == CrmCashflowList.cfcat_id)
        .where(CrmCashflowList.comp_id == company.id)
    )
    if cfcat_id is not None:
        stmt = stmt.where(CrmCashflowList.cfcat_id == cfcat_id)
    if not include_inactive:
        stmt = stmt.where(
            CrmCashflowList.cflist_status == 1,
            CrmCashflowList.cflist_hide.is_(None),
        )
    rows = (await db.execute(stmt.order_by(CrmCashflowList.cflist_name))).mappings().all()
    return [dict(row) for row in rows]


@router.post("/sources", dependencies=[Depends(require_accountant)])
async def create_source(
    payload: SourceCreate,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    await _owned_row(
        db, CrmCashflowCategory, CrmCashflowCategory.cfcat_id,
        payload.cfcat_id, company.id, "ไม่พบหัวข้อ"
    )
    existing = (
        await db.execute(
            select(CrmCashflowList).where(
                CrmCashflowList.comp_id == company.id,
                CrmCashflowList.cfcat_id == payload.cfcat_id,
                func.lower(CrmCashflowList.cflist_name) == payload.cflist_name.lower(),
            )
        )
    ).scalar_one_or_none()
    if existing:
        if existing.cflist_status == 0 or existing.cflist_hide is not None:
            existing.cflist_status = 1
            existing.cflist_hide = None
            await db.commit()
            return {"cflist_id": existing.cflist_id, "cflist_name": existing.cflist_name,
                    "cfcat_id": existing.cfcat_id, "cflist_status": existing.cflist_status,
                    "comp_id": existing.comp_id, "cflist_hide": existing.cflist_hide}
        raise HTTPException(409, "มีชื่อแหล่งที่มานี้อยู่แล้ว")
    row = CrmCashflowList(
        cflist_name=payload.cflist_name,
        cfcat_id=payload.cfcat_id,
        cflist_status=1,
        comp_id=company.id,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return {"cflist_id": row.cflist_id, "cflist_name": row.cflist_name,
            "cfcat_id": row.cfcat_id, "cflist_status": row.cflist_status,
            "comp_id": row.comp_id, "cflist_hide": row.cflist_hide}


@router.patch("/sources/{source_id}", dependencies=[Depends(require_accountant)])
async def update_source(
    source_id: int,
    payload: SourceUpdate,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    row = await _owned_row(
        db, CrmCashflowList, CrmCashflowList.cflist_id,
        source_id, company.id, "ไม่พบแหล่งที่มา"
    )
    if payload.cflist_name is not None and payload.cflist_name.lower() != row.cflist_name.lower():
        duplicate = (
            await db.execute(
                select(CrmCashflowList.cflist_id).where(
                    CrmCashflowList.comp_id == company.id,
                    CrmCashflowList.cfcat_id == row.cfcat_id,
                    func.lower(CrmCashflowList.cflist_name) == payload.cflist_name.lower(),
                    CrmCashflowList.cflist_id != source_id,
                )
            )
        ).scalar_one_or_none()
        if duplicate:
            raise HTTPException(409, "มีชื่อแหล่งที่มานี้อยู่แล้ว")
        row.cflist_name = payload.cflist_name
    if payload.cflist_status is not None:
        row.cflist_status = payload.cflist_status
    if "cflist_hide" in payload.model_fields_set:
        row.cflist_hide = payload.cflist_hide
    await db.commit()
    return {"cflist_id": row.cflist_id, "cflist_name": row.cflist_name,
            "cfcat_id": row.cfcat_id, "cflist_status": row.cflist_status,
            "comp_id": row.comp_id, "cflist_hide": row.cflist_hide}


@router.delete("/sources/{source_id}", dependencies=[Depends(require_accountant)])
async def delete_source(
    source_id: int,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    row = await _owned_row(
        db, CrmCashflowList, CrmCashflowList.cflist_id,
        source_id, company.id, "ไม่พบแหล่งที่มา"
    )
    # Count every referencing row regardless of status — a hard delete would
    # fail on the FK either way, so the pre-check must match that reality.
    usage_count = (
        await db.execute(
            select(func.count()).select_from(CrmCashflowStatement).where(
                CrmCashflowStatement.comp_id == company.id,
                CrmCashflowStatement.cflist_id == source_id,
            )
        )
    ).scalar_one()
    if usage_count:
        raise HTTPException(
            409,
            f"ไม่สามารถลบได้ เนื่องจากมีการใช้งานอยู่ กรุณาย้ายข้อมูลก่อน ({usage_count} รายการ)",
        )
    await db.delete(row)
    await db.commit()
    return {"status": 1}


@router.post("/sources/{source_id}/move", dependencies=[Depends(require_accountant)])
async def move_source(
    source_id: int,
    payload: SourceMove,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    source = await _owned_row(
        db, CrmCashflowList, CrmCashflowList.cflist_id,
        source_id, company.id, "ไม่พบแหล่งที่มาต้นทาง"
    )
    await _owned_row(
        db, CrmCashflowCategory, CrmCashflowCategory.cfcat_id,
        payload.new_cfcat_id, company.id, "ไม่พบหัวข้อปลายทาง"
    )
    if payload.new_cflist_id is not None:
        target = await _owned_row(
            db, CrmCashflowList, CrmCashflowList.cflist_id,
            payload.new_cflist_id, company.id, "ไม่พบแหล่งที่มาปลายทาง"
        )
        if target.cfcat_id != payload.new_cfcat_id:
            raise HTTPException(409, "แหล่งที่มาปลายทางไม่อยู่ในหัวข้อที่เลือก")
    else:
        name = (payload.new_list_name or "").strip()
        target = CrmCashflowList(
            cflist_name=name,
            cfcat_id=payload.new_cfcat_id,
            cflist_status=1,
            comp_id=company.id,
        )
        db.add(target)
        await db.flush()

    await db.execute(
        update(CrmCashflowStatement)
        .where(
            CrmCashflowStatement.comp_id == company.id,
            CrmCashflowStatement.cflist_id == source_id,
        )
        .values(cfcat_id=payload.new_cfcat_id, cflist_id=target.cflist_id)
    )
    if source.cflist_id != target.cflist_id:
        source.cflist_status = 0
        source.cflist_hide = 1
    await db.commit()
    return {"status": 1, "moved_to": target.cflist_id}


# ── Master data: departments ────────────────────────────────────────────────
@router.get("/departments", dependencies=[Depends(require_viewer)])
async def list_departments(
    include_inactive: bool = False,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    stmt = select(CrmCashflowDepartment).where(
        CrmCashflowDepartment.comp_id == company.id
    )
    if not include_inactive:
        stmt = stmt.where(CrmCashflowDepartment.cfstate_dep_status == 1)
    rows = (
        await db.execute(stmt.order_by(CrmCashflowDepartment.cfstate_dep_id))
    ).scalars().all()
    return [
        {"cfstate_dep_id": row.cfstate_dep_id,
         "cfstate_dep_name": row.cfstate_dep_name,
         "cfstate_dep_status": row.cfstate_dep_status,
         "comp_id": row.comp_id}
        for row in rows
    ]


@router.post("/departments", dependencies=[Depends(require_accountant)])
async def create_department(
    payload: DepartmentCreate,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    existing = (
        await db.execute(
            select(CrmCashflowDepartment).where(
                CrmCashflowDepartment.comp_id == company.id,
                func.lower(CrmCashflowDepartment.cfstate_dep_name)
                == payload.cfstate_dep_name.lower(),
            )
        )
    ).scalar_one_or_none()
    if existing:
        if existing.cfstate_dep_status == 0:
            existing.cfstate_dep_status = 1
            await db.commit()
            return {"cfstate_dep_id": existing.cfstate_dep_id,
                    "cfstate_dep_name": existing.cfstate_dep_name,
                    "cfstate_dep_status": existing.cfstate_dep_status,
                    "comp_id": existing.comp_id}
        raise HTTPException(409, "มีชื่อแผนกนี้อยู่แล้ว")
    row = CrmCashflowDepartment(
        cfstate_dep_name=payload.cfstate_dep_name,
        cfstate_dep_status=1,
        comp_id=company.id,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return {"cfstate_dep_id": row.cfstate_dep_id,
            "cfstate_dep_name": row.cfstate_dep_name,
            "cfstate_dep_status": row.cfstate_dep_status,
            "comp_id": row.comp_id}


@router.patch("/departments/{department_id}", dependencies=[Depends(require_accountant)])
async def update_department(
    department_id: int,
    payload: DepartmentUpdate,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    row = await _owned_row(
        db, CrmCashflowDepartment, CrmCashflowDepartment.cfstate_dep_id,
        department_id, company.id, "ไม่พบแผนก"
    )
    if payload.cfstate_dep_name is not None and payload.cfstate_dep_name.lower() != row.cfstate_dep_name.lower():
        duplicate = (
            await db.execute(
                select(CrmCashflowDepartment.cfstate_dep_id).where(
                    CrmCashflowDepartment.comp_id == company.id,
                    func.lower(CrmCashflowDepartment.cfstate_dep_name)
                    == payload.cfstate_dep_name.lower(),
                    CrmCashflowDepartment.cfstate_dep_id != department_id,
                )
            )
        ).scalar_one_or_none()
        if duplicate:
            raise HTTPException(409, "มีชื่อแผนกนี้อยู่แล้ว")
        row.cfstate_dep_name = payload.cfstate_dep_name
    if payload.cfstate_dep_status is not None:
        row.cfstate_dep_status = payload.cfstate_dep_status
    await db.commit()
    return {"cfstate_dep_id": row.cfstate_dep_id,
            "cfstate_dep_name": row.cfstate_dep_name,
            "cfstate_dep_status": row.cfstate_dep_status,
            "comp_id": row.comp_id}


@router.delete("/departments/{department_id}", dependencies=[Depends(require_accountant)])
async def delete_department(
    department_id: int,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    row = await _owned_row(
        db, CrmCashflowDepartment, CrmCashflowDepartment.cfstate_dep_id,
        department_id, company.id, "ไม่พบแผนก"
    )
    usage_count = (
        await db.execute(
            select(func.count()).select_from(CrmCashflowStatement).where(
                CrmCashflowStatement.comp_id == company.id,
                CrmCashflowStatement.cfstate_dep_id == department_id,
            )
        )
    ).scalar_one()
    if usage_count:
        raise HTTPException(
            409,
            f"ไม่สามารถลบได้ เนื่องจากมีการใช้งานอยู่ กรุณาย้ายข้อมูลก่อน ({usage_count} รายการ)",
        )
    await db.delete(row)
    await db.commit()
    return {"status": 1}


# ── Master data: import templates ───────────────────────────────────────────
def _serialize_import_template(row: CrmCashflowImportTemplate) -> dict[str, Any]:
    return {
        "cfimptpl_id": row.cfimptpl_id,
        "cfimptpl_name": row.cfimptpl_name,
        "cfimptpl_header_row": bool(row.cfimptpl_header_row),
        "cfimptpl_columns": row.cfimptpl_columns,
        "cfimptpl_status": row.cfimptpl_status,
        "comp_id": row.comp_id,
    }


@router.get("/import-templates", dependencies=[Depends(require_viewer)])
async def list_import_templates(
    include_inactive: bool = False,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    stmt = select(CrmCashflowImportTemplate).where(CrmCashflowImportTemplate.comp_id == company.id)
    if not include_inactive:
        stmt = stmt.where(CrmCashflowImportTemplate.cfimptpl_status == 1)
    rows = (
        await db.execute(stmt.order_by(CrmCashflowImportTemplate.cfimptpl_id))
    ).scalars().all()
    return [_serialize_import_template(row) for row in rows]


@router.post("/import-templates", dependencies=[Depends(require_accountant)])
async def create_import_template(
    payload: ImportTemplateCreate,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    existing = (
        await db.execute(
            select(CrmCashflowImportTemplate.cfimptpl_id).where(
                CrmCashflowImportTemplate.comp_id == company.id,
                func.lower(CrmCashflowImportTemplate.cfimptpl_name) == payload.cfimptpl_name.lower(),
                CrmCashflowImportTemplate.cfimptpl_status == 1,
            )
        )
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(409, "มีชื่อเทมเพลตนี้อยู่แล้ว")
    row = CrmCashflowImportTemplate(
        comp_id=company.id,
        cfimptpl_name=payload.cfimptpl_name,
        cfimptpl_header_row=1 if payload.cfimptpl_header_row else 0,
        cfimptpl_columns=[column.model_dump() for column in payload.cfimptpl_columns],
        cfimptpl_status=1,
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return _serialize_import_template(row)


@router.patch("/import-templates/{template_id}", dependencies=[Depends(require_accountant)])
async def update_import_template(
    template_id: int,
    payload: ImportTemplateUpdate,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    row = await _owned_row(
        db, CrmCashflowImportTemplate, CrmCashflowImportTemplate.cfimptpl_id,
        template_id, company.id, "ไม่พบเทมเพลต",
    )
    if payload.cfimptpl_name is not None and payload.cfimptpl_name.lower() != row.cfimptpl_name.lower():
        duplicate = (
            await db.execute(
                select(CrmCashflowImportTemplate.cfimptpl_id).where(
                    CrmCashflowImportTemplate.comp_id == company.id,
                    func.lower(CrmCashflowImportTemplate.cfimptpl_name) == payload.cfimptpl_name.lower(),
                    CrmCashflowImportTemplate.cfimptpl_id != template_id,
                )
            )
        ).scalar_one_or_none()
        if duplicate:
            raise HTTPException(409, "มีชื่อเทมเพลตนี้อยู่แล้ว")
        row.cfimptpl_name = payload.cfimptpl_name
    if payload.cfimptpl_header_row is not None:
        row.cfimptpl_header_row = 1 if payload.cfimptpl_header_row else 0
    if payload.cfimptpl_columns is not None:
        row.cfimptpl_columns = [column.model_dump() for column in payload.cfimptpl_columns]
    if payload.cfimptpl_status is not None:
        row.cfimptpl_status = payload.cfimptpl_status
    await db.commit()
    return _serialize_import_template(row)


@router.delete("/import-templates/{template_id}", dependencies=[Depends(require_accountant)])
async def delete_import_template(
    template_id: int,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    row = await _owned_row(
        db, CrmCashflowImportTemplate, CrmCashflowImportTemplate.cfimptpl_id,
        template_id, company.id, "ไม่พบเทมเพลต",
    )
    await db.delete(row)
    await db.commit()
    return {"status": 1}


# ── Statements and invoice tracking ─────────────────────────────────────────
@router.get("/statements", dependencies=[Depends(require_viewer)])
async def list_statements(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    cfcat_id: Optional[int] = None,
    verification_status: Optional[VerificationStatus] = None,
    invoice_status: Optional[InvoiceStatus] = None,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    if start_date and end_date and start_date > end_date:
        raise HTTPException(422, "วันที่เริ่มต้นต้องไม่เกินวันที่สิ้นสุด")
    rows = await _list_statements(
        db, company.id, start_date, end_date, cfcat_id,
        verification_status=verification_status,
        invoice_status=invoice_status,
    )
    items = [_serialize_statement(row) for row in rows]
    sum_revenue = sum(item["cfstate_amount"] for item in items if item["cfstate_amount"] > 0)
    sum_expenses = sum(item["cfstate_amount"] for item in items if item["cfstate_amount"] <= 0)
    return {
        "items": items,
        "sum_revenue": sum_revenue,
        "sum_expenses": sum_expenses,
        "total": len(items),
    }


@router.get("/invoices", dependencies=[Depends(require_viewer)])
async def list_pending_invoices(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    cfcat_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    rows = await _list_statements(
        db, company.id, start_date, end_date, cfcat_id, pending_verification_only=True
    )
    items = [_serialize_statement(row) for row in rows]
    return {"items": items, "total": len(items)}


@router.get("/statements/export", dependencies=[Depends(require_viewer)])
async def export_statements(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    cfcat_id: Optional[int] = None,
    verification_status: Optional[VerificationStatus] = None,
    invoice_status: Optional[InvoiceStatus] = None,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    rows = await _list_statements(
        db, company.id, start_date, end_date, cfcat_id,
        verification_status=verification_status,
        invoice_status=invoice_status,
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cashflow Statement"
    headers = [
        "#", "แผนก", "วันที่", "หัวข้อ", "แหล่งที่มา", "รายละเอียด",
        "ใบกำกับภาษี", "คำนวณต้นทุน", "รายรับ", "รายจ่าย",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for index, row in enumerate(rows, 1):
        amount = Decimal(row.cfstate_amount)
        sheet.append([
            index,
            row.cfstate_dep_name or "-",
            row.cfstate_date,
            row.cfcat_name,
            row.cflist_name,
            row.cfstate_detail or "",
            _invoice_status_label(row.cfstate_invoice, row.cfstate_document_type),
            "ON" if row.cfstate_refrain == 1 else "OFF",
            amount if amount > 0 else Decimal("0"),
            amount if amount < 0 else Decimal("0"),
        ])
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(
            max(len(str(cell.value or "")) for cell in column) + 2, 50
        )
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"cashflow_statement_{date.today():%Y%m%d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/statements/batch", dependencies=[Depends(require_accountant)])
async def create_statements(
    payload: StatementBatchCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company: Company = Depends(get_current_company),
):
    created = 0
    skipped = 0
    updated = 0
    for item in payload.items:
        await _ensure_statement_references(db, item, company.id)
        existing_id = await _find_duplicate_statement(
            db, company.id,
            item.cfstate_date, item.cfcat_id, item.cflist_id,
            item.cfstate_dep_id, item.cfstate_invoice, item.cfstate_refrain,
            item.cfstate_detail or None, item.cfstate_amount, item.cfstate_ref or None,
        )
        if existing_id:
            if payload.duplicate_action == "update":
                existing = await _owned_row(
                    db, CrmCashflowStatement, CrmCashflowStatement.cfstate_id,
                    existing_id, company.id, "ไม่พบรายการ",
                )
                existing.user_id = current_user.id
                updated += 1
            elif payload.duplicate_action == "create":
                db.add(
                    CrmCashflowStatement(
                        **item.model_dump(),
                        cfstate_status=1,
                        user_id=current_user.id,
                        comp_id=company.id,
                    )
                )
                created += 1
            else:
                skipped += 1
            continue
        db.add(
            CrmCashflowStatement(
                **item.model_dump(),
                cfstate_status=1,
                user_id=current_user.id,
                comp_id=company.id,
            )
        )
        created += 1
    await db.commit()
    return {"status": 1, "created": created, "skipped": skipped, "updated": updated}


@router.post("/statements/check-duplicates", dependencies=[Depends(require_accountant)])
async def check_statement_duplicates(
    payload: StatementBatchCreate,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    """Check which items in the batch are duplicates of existing statements.

    Returns a list of duplicates with the existing statement id and the
    item data so the frontend can show the user what is duplicated.
    """
    duplicates: list[dict[str, Any]] = []
    for index, item in enumerate(payload.items):
        existing_id = await _find_duplicate_statement(
            db, company.id,
            item.cfstate_date, item.cfcat_id, item.cflist_id,
            item.cfstate_dep_id, item.cfstate_invoice, item.cfstate_refrain,
            item.cfstate_detail or None, item.cfstate_amount, item.cfstate_ref or None,
        )
        if existing_id:
            duplicates.append({
                "index": index,
                "existing_id": existing_id,
                "item": item.model_dump(),
            })
    return {"duplicates": duplicates}


@router.patch("/statements/{statement_id}", dependencies=[Depends(require_accountant)])
async def update_statement_flags(
    statement_id: int,
    payload: StatementFlagsUpdate,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    row = await _owned_row(
        db, CrmCashflowStatement, CrmCashflowStatement.cfstate_id,
        statement_id, company.id, "ไม่พบรายการ"
    )
    if payload.cfstate_verified == 1:
        attachment_count = (
            await db.execute(
                select(func.count()).select_from(CrmCashflowStatementAttachment).where(
                    CrmCashflowStatementAttachment.cfstate_id == statement_id,
                    CrmCashflowStatementAttachment.comp_id == company.id,
                )
            )
        ).scalar_one()
        if attachment_count == 0:
            raise HTTPException(
                409,
                "ไม่สามารถตรวจสอบรายการได้ เนื่องจากยังไม่มีไฟล์แนบ กรุณาแนบเอกสารอย่างน้อย 1 ไฟล์ก่อน",
            )
    if payload.cfstate_invoice is not None:
        row.cfstate_invoice = payload.cfstate_invoice
    if payload.cfstate_refrain is not None:
        row.cfstate_refrain = payload.cfstate_refrain
    if payload.cfstate_verified is not None:
        row.cfstate_verified = payload.cfstate_verified
    if "cfstate_document_type" in payload.model_fields_set:
        row.cfstate_document_type = payload.cfstate_document_type
    await db.commit()
    return {"status": 1}


@router.delete("/statements/{statement_id}", dependencies=[Depends(require_accountant)])
async def delete_statement(
    statement_id: int,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    row = await _owned_row(
        db, CrmCashflowStatement, CrmCashflowStatement.cfstate_id,
        statement_id, company.id, "ไม่พบรายการ"
    )
    attachments = (
        await db.execute(
            select(CrmCashflowStatementAttachment).where(
                CrmCashflowStatementAttachment.cfstate_id == statement_id,
                CrmCashflowStatementAttachment.comp_id == company.id,
            )
        )
    ).scalars().all()
    for attachment in attachments:
        try:
            Path(attachment.file_path).unlink(missing_ok=True)
        except OSError:
            pass
        await db.delete(attachment)
    await db.delete(row)
    await db.commit()
    return {"status": 1}


# ── Import/template helpers ─────────────────────────────────────────────────
IMPORT_HEADERS_10 = [
    "วันที่", "หัวข้อ", "แหล่งที่มา", "รายละเอียด", "ใบกำกับภาษี",
    "คำนวณต้นทุน", "รายรับ", "รายจ่าย", "Ref", "แผนก",
]
IMPORT_HEADERS_8 = [
    "วันที่", "หัวข้อ", "รายการ", "จำนวนเงิน", "รายละเอียด", "ใบเสร็จ",
    "เบิกคืน", "แผนก",
]
REPORT_EXPORT_HEADERS = [
    "#", "แผนก", "วันที่", "หัวข้อ", "แหล่งที่มา", "รายละเอียด",
    "ใบกำกับภาษี", "คำนวณต้นทุน", "รายรับ", "รายจ่าย",
]
DEFAULT_IMPORT_SOURCE_NAME = "ไม่ระบุ"
IMPORT_TEMPLATE_SAMPLE = [
    [date(2025, 2, 27), "รายรับ", "ยอดขายออนไลน์", "รายได้จากการขายสินค้าออนไลน์", "", "0", 15000, 0, "", "ฝ่ายขาย"],
    [date(2025, 2, 27), "รายรับ", "ยอดขายหน้าร้าน", "รายได้จากการขายสินค้าหน้าร้าน", "", "0", 8500, 0, "", "ฝ่ายขาย"],
    [date(2025, 2, 27), "รายจ่าย", "ค่าเช่า", "ค่าเช่าสำนักงานประจำเดือน", "", "0", 0, 12000, "", "ฝ่ายบัญชี"],
    [date(2025, 2, 27), "รายจ่าย", "", "ตัวอย่างรายการที่ไม่ระบุแหล่งที่มา", "", "0", 0, 1200, "", "ฝ่ายบัญชี"],
]


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_spreadsheet(filename: str, content: bytes) -> list[list[Any]]:
    extension = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if extension == "csv":
        decoded = None
        for encoding in ("utf-8-sig", "cp874", "utf-8"):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise HTTPException(400, "ไม่สามารถอ่าน encoding ของไฟล์ CSV ได้")
        return [list(row) for row in csv.reader(io.StringIO(decoded))]
    if extension == "xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        return [list(row) for row in workbook.active.iter_rows(values_only=True)]
    if extension == "xls":
        try:
            import xlrd
        except ImportError as exc:  # pragma: no cover - depends on deployment image
            raise HTTPException(500, "เซิร์ฟเวอร์ยังไม่ได้ติดตั้งตัวอ่านไฟล์ XLS") from exc
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        return [sheet.row_values(index) for index in range(sheet.nrows)]
    raise HTTPException(400, "รองรับเฉพาะไฟล์ CSV, XLSX หรือ XLS")


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            parsed = from_excel(value)
            return parsed.date() if isinstance(parsed, datetime) else parsed
        except Exception:
            pass
    text_value = _text(value)
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text_value, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"รูปแบบวันที่ไม่ถูกต้อง: {text_value}")


def _parse_money(value: Any) -> Decimal:
    text_value = _text(value).replace(",", "").replace(" ", "")
    if text_value == "":
        return Decimal("0")
    try:
        return Decimal(text_value)
    except InvalidOperation as exc:
        raise ValueError(f"จำนวนเงินไม่ใช่ตัวเลข: {_text(value)}") from exc


def _parse_refrain(value: Any) -> int:
    normalized = _text(value).lower()
    if normalized in ("1", "on", "คำนวณต้นทุน"):
        return 1
    if normalized in ("0", "off", "", "ไม่คำนวณ", "ไม่คำนวณต้นทุน"):
        return 0
    raise ValueError(f"ค่าคำนวณต้นทุนไม่ถูกต้อง: {_text(value)}")


def _split_amounts(income_value: Any, expense_value: Any) -> tuple[Decimal, Decimal]:
    income = _parse_money(income_value)
    expense = _parse_money(expense_value)
    if expense > 0:
        expense = -expense
    if income != 0 and expense != 0:
        raise ValueError(
            "กรอกจำนวนเงินได้เพียงช่องรายรับหรือรายจ่ายช่องเดียว "
            "โดยอีกช่องต้องเป็น 0"
        )
    return income, expense


def _category_from_amounts(category: Any, income: Decimal, expense: Decimal) -> str:
    """Use an explicit category when supplied; otherwise infer cash direction.

    Split amount files commonly contain no category column at all.  A positive
    income value is income, while any populated expense value (normalised to a
    negative number by ``_split_amounts``) is expense.  A combined amount uses
    its sign in exactly the same way.
    """
    explicit = _text(category)
    if explicit:
        return explicit
    if income > 0:
        return "รายรับ"
    if expense < 0:
        return "รายจ่าย"
    return ""


def _normalize_import_row(
    row: list[Any], legacy_eight_columns: bool, report_export_columns: bool = False
) -> dict[str, Any]:
    padded = list(row) + [None] * max(0, 10 - len(row))
    if legacy_eight_columns:
        amount = _parse_money(padded[3])
        income = amount if amount > 0 else Decimal("0")
        expense = amount if amount < 0 else Decimal("0")
        return {
            "cfstate_date": _parse_date(padded[0]),
            "category": _category_from_amounts(padded[1], income, expense),
            "source": _text(padded[2]),
            "detail": _text(padded[4]),
            # Imported rows always start with an unassigned document status.
            # Accounting staff classify the document during invoice review.
            "invoice": None,
            "refrain": _parse_refrain(padded[6]),
            "income": income,
            "expense": expense,
            "ref": "",
            "department": _text(padded[7]),
        }
    if report_export_columns:
        income, expense = _split_amounts(padded[8], padded[9])
        return {
            "cfstate_date": _parse_date(padded[2]),
            "category": _category_from_amounts(padded[3], income, expense),
            "source": _text(padded[4]),
            "detail": _text(padded[5]),
            "invoice": None,
            "refrain": _parse_refrain(padded[7]),
            "income": income,
            "expense": expense,
            "ref": "",
            "department": _text(padded[1]),
        }
    income, expense = _split_amounts(padded[6], padded[7])
    return {
        "cfstate_date": _parse_date(padded[0]),
        "category": _category_from_amounts(padded[1], income, expense),
        "source": _text(padded[2]),
        "detail": _text(padded[3]),
        "invoice": None,
        "refrain": _parse_refrain(padded[5]),
        "income": income,
        "expense": expense,
        "ref": _text(padded[8]),
        "department": _text(padded[9]),
    }


def _normalize_import_row_with_template(
    row: list[Any], columns: list[dict[str, str]]
) -> dict[str, Any]:
    """Normalize a row using a user-defined Excel-column mapping.

    ``column`` is an A1-style column name (A, J, AO, ...).  Templates saved by
    the previous positional editor have no ``column`` key, so their list index
    remains the fallback position.
    """
    values: dict[str, Any] = {}
    for index, column in enumerate(columns):
        if column["field"] == "skip":
            continue
        if "column" in column and not column["column"]:
            values[column["field"]] = None
            continue
        column_name = column.get("column")
        source_index = column_index_from_string(column_name) - 1 if column_name else index
        values[column["field"]] = row[source_index] if source_index < len(row) else None

    if values.get("amount") is not None:
        amount = _parse_money(values.get("amount"))
        income = amount if amount > 0 else Decimal("0")
        expense = amount if amount < 0 else Decimal("0")
    else:
        income, expense = _split_amounts(values.get("income"), values.get("expense"))

    return {
        "cfstate_date": _parse_date(values.get("date")),
        "category": _category_from_amounts(values.get("category"), income, expense),
        "source": _text(values.get("source")),
        "detail": _text(values.get("detail")),
        # Imported rows always start with an unassigned document status.
        # Accounting staff classify the document during invoice review.
        "invoice": None,
        "refrain": _parse_refrain(values.get("refrain")),
        "income": income,
        "expense": expense,
        "ref": _text(values.get("ref")),
        "department": _text(values.get("department")),
    }


def _prepare_import_rows(
    rows: list[list[Any]],
    has_header: bool,
    template_columns: Optional[list[dict[str, str]]] = None,
):
    if not rows:
        raise HTTPException(400, "ไฟล์ไม่มีข้อมูล")
    if template_columns is not None:
        def normalize(row: list[Any]) -> dict[str, Any]:
            return _normalize_import_row_with_template(row, template_columns)
    else:
        header = [_text(value) for value in rows[0]] if has_header else []
        report_export_columns = header[:10] == REPORT_EXPORT_HEADERS
        legacy_eight_columns = (
            "จำนวนเงิน" in header
            or (not has_header and len(rows[0]) < 10)
        )

        def normalize(row: list[Any]) -> dict[str, Any]:
            return _normalize_import_row(row, legacy_eight_columns, report_export_columns)

    data_rows = rows[1:] if has_header else rows
    prepared = []
    for row_number, row in enumerate(data_rows, start=2 if has_header else 1):
        if not any(_text(value) for value in row):
            continue
        raw = [_text(value) for value in row]
        errors: list[str] = []
        normalized = None
        try:
            normalized = normalize(row)
            if not normalized["category"]:
                errors.append("ไม่ระบุหัวข้อ")
        except ValueError as exc:
            errors.append(str(exc))
        prepared.append(
            {"row_number": row_number, "raw": raw, "data": normalized, "errors": errors}
        )
    return prepared


async def _master_maps(db: AsyncSession, comp_id: int):
    categories = (
        await db.execute(
            select(CrmCashflowCategory).where(
                CrmCashflowCategory.comp_id == comp_id,
                CrmCashflowCategory.cfcat_status == 1,
            )
        )
    ).scalars().all()
    sources = (
        await db.execute(
            select(CrmCashflowList).where(
                CrmCashflowList.comp_id == comp_id,
                CrmCashflowList.cflist_status == 1,
            )
        )
    ).scalars().all()
    departments = (
        await db.execute(
            select(CrmCashflowDepartment).where(
                CrmCashflowDepartment.comp_id == comp_id,
                CrmCashflowDepartment.cfstate_dep_status == 1,
            )
        )
    ).scalars().all()
    return (
        {row.cfcat_name.casefold(): row for row in categories},
        {(row.cflist_name.casefold(), row.cfcat_id): row for row in sources},
        {row.cfstate_dep_name.casefold(): row for row in departments},
    )


async def _find_import_duplicate(
    db: AsyncSession,
    comp_id: int,
    data: dict[str, Any],
    categories: dict[str, CrmCashflowCategory],
    sources: dict[tuple[str, int], CrmCashflowList],
    departments: dict[str, CrmCashflowDepartment],
) -> Optional[int]:
    """Best-effort duplicate lookup for a normalised import row.

    Only checks rows whose referenced category/source/department already
    exist — a row that will create brand-new master data cannot duplicate a
    pre-existing statement, so it is left unchecked (returns ``None``).
    """
    category = categories.get(data["category"].casefold())
    if category is None:
        return None
    source_name = data["source"] or DEFAULT_IMPORT_SOURCE_NAME
    source = sources.get((source_name.casefold(), category.cfcat_id))
    if source is None:
        return None
    cfstate_dep_id = None
    if data["department"]:
        department = departments.get(data["department"].casefold())
        if department is None:
            return None
        cfstate_dep_id = department.cfstate_dep_id
    return await _find_duplicate_statement(
        db, comp_id,
        data["cfstate_date"], category.cfcat_id, source.cflist_id,
        cfstate_dep_id, data["invoice"], data["refrain"],
        data["detail"] or None, data["income"] + data["expense"], data["ref"] or None,
    )


def _parse_row_numbers(value: Optional[str]) -> set[int]:
    """Parse a JSON array or comma-separated string of row numbers."""
    if not value or not value.strip():
        return set()
    text_value = value.strip()
    try:
        parsed = json.loads(text_value)
        if isinstance(parsed, list):
            return {int(entry) for entry in parsed}
    except (ValueError, TypeError):
        pass
    return {int(part) for part in text_value.split(",") if part.strip()}


async def _resolve_import_template_columns(
    db: AsyncSession, company: Company, template_id: Optional[int]
) -> Optional[list[dict[str, str]]]:
    if template_id is None:
        return None
    template = await _owned_row(
        db, CrmCashflowImportTemplate, CrmCashflowImportTemplate.cfimptpl_id,
        template_id, company.id, "ไม่พบเทมเพลต",
    )
    return template.cfimptpl_columns


@router.post("/import/preview", dependencies=[Depends(require_accountant)])
async def preview_import(
    file: UploadFile = File(...),
    header_row: bool = Form(True),
    use_existing_data: bool = Form(False),
    template_id: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    template_columns = await _resolve_import_template_columns(db, company, template_id)
    prepared = _prepare_import_rows(
        _read_spreadsheet(file.filename or "", await file.read()), header_row, template_columns
    )
    categories, sources, departments = await _master_maps(db, company.id)
    duplicates: list[dict[str, Any]] = []
    for item in prepared:
        data = item["data"]
        item["duplicate_id"] = None
        if not data or item["errors"]:
            continue
        if use_existing_data:
            category = categories.get(data["category"].casefold())
            if category is None:
                item["errors"].append(f'ไม่พบหัวข้อ: {data["category"]}')
                continue
            if (
                data["source"]
                and (data["source"].casefold(), category.cfcat_id) not in sources
            ):
                item["errors"].append(f'ไม่พบแหล่งที่มา: {data["source"]}')
            if data["department"] and data["department"].casefold() not in departments:
                item["errors"].append(f'ไม่พบแผนก: {data["department"]}')
            if item["errors"]:
                continue
        duplicate_id = await _find_import_duplicate(
            db, company.id, data, categories, sources, departments
        )
        if duplicate_id:
            item["duplicate_id"] = duplicate_id
            duplicates.append({"row_number": item["row_number"], "existing_id": duplicate_id})
    errors = [item for item in prepared if item["errors"]]
    # Show every parsed row (not just a handful) so the user can review the
    # whole file — including duplicate flags/checkboxes — before importing.
    preview = []
    for item in prepared:
        data = item["data"]
        preview.append({
            "row_number": item["row_number"],
            "raw": item["raw"],
            "data": {
                **(data or {}),
                "cfstate_date": data["cfstate_date"].isoformat() if data else None,
                "income": float(data["income"]) if data else None,
                "expense": float(data["expense"]) if data else None,
            } if data else None,
            "errors": item["errors"],
            "duplicate_id": item["duplicate_id"],
        })
    return {
        "status": 1,
        "headers": IMPORT_HEADERS_10,
        "preview": preview,
        "total_rows": len(prepared),
        "error_count": len(errors),
        "error_rows": [item["raw"] for item in errors],
        "error_details": ["; ".join(item["errors"]) for item in errors],
        "warnings": [f"แถว {item['row_number']}: {'; '.join(item['errors'])}" for item in errors],
        "duplicate_count": len(duplicates),
        "duplicates": duplicates,
    }


@router.post("/import", dependencies=[Depends(require_accountant)])
async def import_statements(
    file: UploadFile = File(...),
    header_row: bool = Form(True),
    use_existing_data: bool = Form(False),
    duplicate_action: str = Form("skip", pattern="^(skip|update|create)$"),
    skip_rows: Optional[str] = Form(None),
    template_id: Optional[int] = Form(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company: Company = Depends(get_current_company),
):
    template_columns = await _resolve_import_template_columns(db, company, template_id)
    prepared = _prepare_import_rows(
        _read_spreadsheet(file.filename or "", await file.read()), header_row, template_columns
    )
    categories, sources, departments = await _master_maps(db, company.id)
    skip_row_numbers = _parse_row_numbers(skip_rows)
    imported = 0
    skipped = 0
    updated = 0
    error_rows: list[list[Any]] = []
    error_details: list[str] = []

    for item in prepared:
        data = item["data"]
        errors = list(item["errors"])
        if item["row_number"] in skip_row_numbers:
            skipped += 1
            continue
        try:
            if data is None or errors:
                raise ValueError("; ".join(errors) or "ข้อมูลไม่ถูกต้อง")
            category = categories.get(data["category"].casefold())
            if category is None:
                if use_existing_data:
                    raise ValueError(f'ไม่พบหัวข้อ: {data["category"]}')
                category = CrmCashflowCategory(
                    cfcat_name=data["category"], cfcat_status=1, comp_id=company.id
                )
                db.add(category)
                await db.flush()
                categories[data["category"].casefold()] = category

            source_name = data["source"] or DEFAULT_IMPORT_SOURCE_NAME
            source_key = (source_name.casefold(), category.cfcat_id)
            source = sources.get(source_key)
            if source is None:
                if use_existing_data and data["source"]:
                    raise ValueError(f'ไม่พบแหล่งที่มา: {data["source"]}')
                source = CrmCashflowList(
                    cflist_name=source_name, cfcat_id=category.cfcat_id,
                    cflist_status=1, comp_id=company.id,
                )
                db.add(source)
                await db.flush()
                sources[source_key] = source

            department = None
            if data["department"]:
                department = departments.get(data["department"].casefold())
                if department is None:
                    if use_existing_data:
                        raise ValueError(f'ไม่พบแผนก: {data["department"]}')
                    department = CrmCashflowDepartment(
                        cfstate_dep_name=data["department"],
                        cfstate_dep_status=1,
                        comp_id=company.id,
                    )
                    db.add(department)
                    await db.flush()
                    departments[data["department"].casefold()] = department

            duplicate_id = await _find_duplicate_statement(
                db, company.id,
                data["cfstate_date"], category.cfcat_id, source.cflist_id,
                department.cfstate_dep_id if department else None,
                data["invoice"], data["refrain"],
                data["detail"] or None, data["income"] + data["expense"], data["ref"] or None,
            )
            if duplicate_id:
                if duplicate_action == "skip":
                    skipped += 1
                    continue
                if duplicate_action == "update":
                    existing = await _owned_row(
                        db, CrmCashflowStatement, CrmCashflowStatement.cfstate_id,
                        duplicate_id, company.id, "ไม่พบรายการ",
                    )
                    existing.user_id = current_user.id
                    updated += 1
                    continue
                # duplicate_action == "create" falls through and inserts a new row

            db.add(CrmCashflowStatement(
                cfstate_date=data["cfstate_date"],
                cfcat_id=category.cfcat_id,
                cflist_id=source.cflist_id,
                user_id=current_user.id,
                comp_id=company.id,
                cfstate_amount=data["income"] + data["expense"],
                cfstate_refrain=data["refrain"],
                cfstate_invoice=data["invoice"],
                cfstate_document_type=None,
                cfstate_verified=0,
                cfstate_detail=data["detail"] or None,
                cfstate_status=1,
                cfstate_dep_id=department.cfstate_dep_id if department else None,
                cfstate_ref=data["ref"] or None,
            ))
            imported += 1
        except ValueError as exc:
            error_rows.append(item["raw"])
            error_details.append(str(exc))

    await db.commit()
    return {
        "status": 1,
        "imported": imported,
        "skipped": skipped,
        "updated": updated,
        "errors": len(error_rows),
        "error_rows": error_rows,
        "error_details": error_details,
    }


@router.get("/import/template", dependencies=[Depends(require_viewer)])
async def download_import_template(format: str = Query("xlsx", pattern="^(xlsx|csv)$")):
    if format == "csv":
        text_buffer = io.StringIO()
        writer = csv.writer(text_buffer)
        writer.writerow(IMPORT_HEADERS_10)
        writer.writerows(IMPORT_TEMPLATE_SAMPLE)
        content = io.BytesIO(text_buffer.getvalue().encode("utf-8-sig"))
        return StreamingResponse(
            content,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="cashflow_import_template.csv"'},
        )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cashflow Import"
    sheet.append(IMPORT_HEADERS_10)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{len(IMPORT_TEMPLATE_SAMPLE) + 1}"
    sheet.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26
    for row in IMPORT_TEMPLATE_SAMPLE:
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row[0].number_format = "yyyy-mm-dd"
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[6].number_format = "#,##0.00"
        row[7].number_format = "#,##0.00"
    for column, width in {
        "A": 14, "B": 16, "C": 24, "D": 38, "E": 18,
        "F": 18, "G": 14, "H": 14, "I": 18, "J": 18,
    }.items():
        sheet.column_dimensions[column].width = width
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="cashflow_import_template.xlsx"'},
    )


def _sample_value_for_import_field(field: str, row_index: int) -> Any:
    """A plausible example value for a template column, so the downloaded
    file isn't just bare headers — mirrors IMPORT_TEMPLATE_SAMPLE's spirit
    but works for any field a custom template chooses to include."""
    if field == "date":
        return date(2025, 2, 27) if row_index == 0 else date(2025, 2, 28)
    if field == "category":
        return "รายรับ" if row_index == 0 else "รายจ่าย"
    if field == "source":
        return "ยอดขายออนไลน์" if row_index == 0 else "ค่าเช่า"
    if field == "detail":
        return "รายได้จากการขายสินค้าออนไลน์" if row_index == 0 else "ค่าเช่าสำนักงานประจำเดือน"
    if field == "refrain":
        return "0"
    if field == "income":
        return 15000 if row_index == 0 else 0
    if field == "expense":
        return 0 if row_index == 0 else 12000
    if field == "amount":
        return 15000 if row_index == 0 else -12000
    if field == "department":
        return "ฝ่ายขาย" if row_index == 0 else "ฝ่ายบัญชี"
    return ""  # ref, skip


@router.get("/import-templates/{template_id}/download", dependencies=[Depends(require_viewer)])
async def download_custom_import_template(
    template_id: int,
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    """Download a blank file laid out at the template's Excel columns."""
    template = await _owned_row(
        db, CrmCashflowImportTemplate, CrmCashflowImportTemplate.cfimptpl_id,
        template_id, company.id, "ไม่พบเทมเพลต",
    )
    columns = template.cfimptpl_columns
    positioned_columns = []
    for index, column in enumerate(columns):
        if "column" in column and not column["column"]:
            continue
        position = column_index_from_string(column["column"]) if column.get("column") else index + 1
        positioned_columns.append((position, column))
    max_column = max((position for position, _ in positioned_columns), default=1)
    headers = [""] * max_column
    sample_rows = [[""] * max_column for _ in range(2)]
    for position, column in positioned_columns:
        headers[position - 1] = column["label"]
        for row_index in range(2):
            sample_rows[row_index][position - 1] = _sample_value_for_import_field(
                column["field"], row_index
            )
    # Content-Disposition must be latin-1 encodable — cfimptpl_name is
    # typically Thai, so strip to ASCII word characters only (plain \w would
    # keep Thai letters, since Python regex treats them as "word" chars too).
    safe_name = re.sub(r"[^A-Za-z0-9_\-]+", "_", template.cfimptpl_name).strip("_") or "import_template"

    if format == "csv":
        text_buffer = io.StringIO()
        writer = csv.writer(text_buffer)
        writer.writerow(headers)
        writer.writerows(sample_rows)
        content = io.BytesIO(text_buffer.getvalue().encode("utf-8-sig"))
        return StreamingResponse(
            content,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}.csv"'},
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import Template"
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sample_rows) + 1}"
    sheet.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26
    for row in sample_rows:
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for position, column in positioned_columns:
            cell = row[position - 1]
            if column["field"] == "date":
                cell.number_format = "yyyy-mm-dd"
            elif column["field"] in ("income", "expense", "amount"):
                cell.number_format = "#,##0.00"
            elif column["field"] == "detail":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, column in positioned_columns:
        width = 38 if column["field"] == "detail" else 18
        sheet.column_dimensions[get_column_letter(index)].width = width
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.xlsx"'},
    )


@router.post("/import/errors", dependencies=[Depends(require_accountant)])
async def download_error_workbook(payload: ErrorWorkbookPayload):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cashflow Import Errors"
    headers = IMPORT_HEADERS_10 + ["สาเหตุข้อผิดพลาด"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for index, raw in enumerate(payload.error_rows):
        values = list(raw[:10]) + [""] * max(0, 10 - len(raw))
        values.append(payload.error_details[index] if index < len(payload.error_details) else "")
        sheet.append(values)
        sheet.cell(row=sheet.max_row, column=11).font = Font(color="FF0000")
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="cashflow_import_errors.xlsx"'},
    )


ALLOWED_ATTACHMENT_TYPES = {
    "image/jpeg", "image/png", "application/pdf",
}
MAX_ATTACHMENT_BYTES = 10 * 1024 * 1024


@router.get("/statements/{statement_id}/attachments", dependencies=[Depends(require_viewer)])
async def list_attachments(
    statement_id: int,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    """List all attachments for a given statement."""
    await _owned_row(
        db, CrmCashflowStatement, CrmCashflowStatement.cfstate_id,
        statement_id, company.id, "ไม่พบรายการ"
    )
    rows = (
        await db.execute(
            select(
                CrmCashflowStatementAttachment.id,
                CrmCashflowStatementAttachment.file_name,
                CrmCashflowStatementAttachment.content_type,
                CrmCashflowStatementAttachment.file_size,
                CrmCashflowStatementAttachment.created_at,
                User.full_name.label("uploaded_by"),
            )
            .join(User, User.id == CrmCashflowStatementAttachment.created_by)
            .where(
                CrmCashflowStatementAttachment.cfstate_id == statement_id,
                CrmCashflowStatementAttachment.comp_id == company.id,
            )
            .order_by(CrmCashflowStatementAttachment.created_at.desc())
        )
    ).mappings().all()
    return [
        {
            "id": str(row.id),
            "file_name": row.file_name,
            "content_type": row.content_type,
            "file_size": row.file_size,
            "created_at": row.created_at.isoformat(),
            "uploaded_by": row.uploaded_by,
        }
        for row in rows
    ]


@router.post("/statements/{statement_id}/attachments", status_code=201, dependencies=[Depends(require_accountant)])
async def upload_attachment(
    statement_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company: Company = Depends(get_current_company),
):
    """Upload an attachment (image or PDF) for a statement."""
    await _owned_row(
        db, CrmCashflowStatement, CrmCashflowStatement.cfstate_id,
        statement_id, company.id, "ไม่พบรายการ"
    )
    existing_count = (
        await db.execute(
            select(func.count()).select_from(CrmCashflowStatementAttachment).where(
                CrmCashflowStatementAttachment.cfstate_id == statement_id,
                CrmCashflowStatementAttachment.comp_id == company.id,
            )
        )
    ).scalar_one()
    if existing_count >= 2:
        raise HTTPException(400, "แนบได้สูงสุด 2 ไฟล์ต่อรายการ")
    if file.content_type not in ALLOWED_ATTACHMENT_TYPES:
        raise HTTPException(400, "รองรับเฉพาะไฟล์ JPG, PNG และ PDF เท่านั้น")
    content = await file.read()
    if len(content) > MAX_ATTACHMENT_BYTES:
        raise HTTPException(400, "ไฟล์มีขนาดใหญ่เกิน 10 MB")
    filename = (file.filename or "file").strip()
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "bin"
    stored_name = f"{uuid4()}.{ext}"
    upload_dir = Path(settings.CRM_CASHFLOW_UPLOAD_DIR)
    upload_dir.mkdir(parents=True, exist_ok=True)
    file_path = upload_dir / stored_name
    file_path.write_bytes(content)
    attachment = CrmCashflowStatementAttachment(
        cfstate_id=statement_id,
        comp_id=company.id,
        file_name=filename,
        stored_name=stored_name,
        file_path=str(file_path),
        content_type=file.content_type,
        file_size=len(content),
        created_by=current_user.id,
    )
    db.add(attachment)
    await db.commit()
    await db.refresh(attachment)
    return {
        "id": str(attachment.id),
        "file_name": attachment.file_name,
        "content_type": attachment.content_type,
        "file_size": attachment.file_size,
        "created_at": attachment.created_at.isoformat(),
    }


@router.get("/statements/{statement_id}/attachments/{attachment_id}", dependencies=[Depends(require_viewer)])
async def view_attachment(
    statement_id: int,
    attachment_id: str,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    """View/download an attachment file."""
    await _owned_row(
        db, CrmCashflowStatement, CrmCashflowStatement.cfstate_id,
        statement_id, company.id, "ไม่พบรายการ"
    )
    attachment = (
        await db.execute(
            select(CrmCashflowStatementAttachment).where(
                CrmCashflowStatementAttachment.id == attachment_id,
                CrmCashflowStatementAttachment.cfstate_id == statement_id,
                CrmCashflowStatementAttachment.comp_id == company.id,
            )
        )
    ).scalar_one_or_none()
    if not attachment or not Path(attachment.file_path).is_file():
        raise HTTPException(404, "ไม่พบไฟล์แนบนี้")
    return FileResponse(
        attachment.file_path,
        media_type=attachment.content_type or "application/octet-stream",
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{quote(attachment.file_name)}"},
    )


@router.delete("/statements/{statement_id}/attachments/{attachment_id}", status_code=204, dependencies=[Depends(require_accountant)])
async def delete_attachment(
    statement_id: int,
    attachment_id: str,
    db: AsyncSession = Depends(get_db),
    company: Company = Depends(get_current_company),
):
    """Delete an attachment."""
    await _owned_row(
        db, CrmCashflowStatement, CrmCashflowStatement.cfstate_id,
        statement_id, company.id, "ไม่พบรายการ"
    )
    attachment = (
        await db.execute(
            select(CrmCashflowStatementAttachment).where(
                CrmCashflowStatementAttachment.id == attachment_id,
                CrmCashflowStatementAttachment.cfstate_id == statement_id,
                CrmCashflowStatementAttachment.comp_id == company.id,
            )
        )
    ).scalar_one_or_none()
    if not attachment:
        raise HTTPException(404, "ไม่พบไฟล์แนบนี้")
    try:
        Path(attachment.file_path).unlink(missing_ok=True)
    except OSError:
        pass
    await db.delete(attachment)
    await db.commit()
