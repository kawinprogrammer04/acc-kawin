from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import json
import logging
import os
import re
import time
import uuid

# Without this, the root logger defaults to WARNING and the OCR debug logs in
# parsers.py (_ocr_one_image's "raw response"/"parsed N rows" INFO lines —
# the only visibility we have into *why* a screenshot failed to parse) are
# silently dropped. `docker logs acc_statement` needs this to show anything.
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")
from contextlib import asynccontextmanager, closing
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
from urllib.parse import quote

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from psycopg import Connection
from psycopg.rows import dict_row

from app.parsers import (
    ParsedTransaction,
    guess_category,
    make_row_hash,
    parse_amount,
    parse_date,
    parse_images_with_vision_model,
    parse_statement_with_metadata,
    parse_time,
)
from app.staging import (
    PreviewNotFoundError,
    cleanup_expired_previews,
    create_preview,
    delete_preview,
    load_preview,
    read_preview_source,
)


APP_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = Path(os.getenv("UPLOAD_DIR", "/app/uploads"))
ROOT_PATH = os.getenv("ROOT_PATH", "/statement").rstrip("/")
MAX_UPLOAD_BYTES = 25 * 1024 * 1024
UPLOAD_JOB_TTL_SECONDS = int(os.getenv("STATEMENT_UPLOAD_JOB_TTL_SECONDS", "3600"))
MAX_PENDING_UPLOAD_JOBS = int(os.getenv("STATEMENT_MAX_PENDING_UPLOAD_JOBS", "3"))
OCR_JOB_CONCURRENCY = int(os.getenv("STATEMENT_OCR_JOB_CONCURRENCY", "1"))
UPLOAD_JOBS: dict[str, dict[str, Any]] = {}
UPLOAD_TASKS: set[asyncio.Task[Any]] = set()
OCR_JOB_SEMAPHORE: asyncio.Semaphore | None = None


class PgConnection:
    def __init__(self, connection: Connection):
        self.connection = connection

    def execute(self, sql: str, params: Any = None):
        return self.connection.execute(sql.replace("?", "%s"), params)

    def executemany(self, sql: str, params_seq: Any):
        with self.connection.cursor() as cursor:
            cursor.executemany(sql.replace("?", "%s"), params_seq)
            return cursor

    def commit(self) -> None:
        self.connection.commit()

    def close(self) -> None:
        self.connection.close()


def postgres_dsn() -> str:
    dsn = os.getenv("STATEMENT_DATABASE_URL") or os.getenv("DATABASE_URL") or ""
    dsn = dsn.replace("postgresql+asyncpg://", "postgresql://", 1)
    dsn = dsn.replace("postgresql+psycopg2://", "postgresql://", 1)
    if dsn:
        return dsn

    password = os.getenv("POSTGRES_PASSWORD")
    if not password:
        raise RuntimeError("Statement service requires STATEMENT_DATABASE_URL, DATABASE_URL, or POSTGRES_PASSWORD")
    user = os.getenv("POSTGRES_USER", "postgres")
    database = os.getenv("POSTGRES_DB", "accounting_db")
    host = os.getenv("DATABASE_HOST", "db")
    port = os.getenv("DATABASE_PORT", "5432")
    return f"postgresql://{user}:{password}@{host}:{port}/{database}"


def connect() -> PgConnection:
    connection = Connection.connect(postgres_dsn(), row_factory=dict_row)
    return PgConnection(connection)


def init_database() -> None:
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    with closing(connect()) as db:
        db.execute(
            "ALTER TABLE statements ADD COLUMN IF NOT EXISTS issuer TEXT"
        )
        db.execute(
            "ALTER TABLE statements ADD COLUMN IF NOT EXISTS statement_type TEXT"
        )
        db.execute(
            "ALTER TABLE statements ADD COLUMN IF NOT EXISTS processing_method TEXT"
        )
        db.execute(
            "ALTER TABLE statements ADD COLUMN IF NOT EXISTS masked_reference TEXT"
        )
        db.execute(
            "ALTER TABLE statements ADD COLUMN IF NOT EXISTS parse_warnings TEXT"
        )
        db.commit()


@asynccontextmanager
async def lifespan(_: FastAPI):
    global OCR_JOB_SEMAPHORE
    init_database()
    OCR_JOB_SEMAPHORE = asyncio.Semaphore(max(1, OCR_JOB_CONCURRENCY))
    try:
        yield
    finally:
        for task in list(UPLOAD_TASKS):
            task.cancel()
        if UPLOAD_TASKS:
            await asyncio.gather(*UPLOAD_TASKS, return_exceptions=True)


app = FastAPI(
    title="Credit Statement Matcher",
    version="1.0.0",
    root_path=ROOT_PATH,
    lifespan=lifespan,
)
app.mount("/static", StaticFiles(directory=APP_DIR / "static"), name="static")

templates = Jinja2Templates(directory=APP_DIR / "templates")
templates.env.filters["money"] = lambda value: f"{float(value or 0):,.2f}"
templates.env.filters["clean_desc"] = lambda d: (d.split(" | ")[-1] if d and " | " in d else (d or ""))


def page(
    request: Request,
    template: str,
    *,
    active: str,
    status_code: int = 200,
    **context: Any,
) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request,
        name=template,
        context={"active": active, "root_path": ROOT_PATH, **context},
        status_code=status_code,
    )


def redirect(path: str) -> RedirectResponse:
    return RedirectResponse(f"{ROOT_PATH}{path}", status_code=303)


def redirect_manual_error(message: str) -> RedirectResponse:
    return redirect(f"/manual-edit?status=unmatched&error={quote(message)}")


def money_cents(value: float | int | str | None) -> int:
    amount = Decimal(str(value or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return int(amount * 100)


def sync_statement_stats(db: PgConnection, statement_id: int | None = None) -> None:
    where = "WHERE id = ?" if statement_id else ""
    values: tuple[Any, ...] = (statement_id,) if statement_id else ()
    statements = db.execute(f"SELECT id FROM statements {where}", values).fetchall()
    for statement in statements:
        sid = int(statement["id"])
        stats = db.execute(
            """
            SELECT
                COUNT(*) AS row_count,
                SUM(CASE WHEN amount > 0 THEN 1 ELSE 0 END) AS deposit_count,
                SUM(CASE WHEN amount < 0 THEN 1 ELSE 0 END) AS withdraw_count,
                COALESCE(SUM(CASE WHEN amount > 0 THEN amount ELSE 0 END), 0) AS total_deposit,
                COALESCE(SUM(CASE WHEN amount < 0 THEN ABS(amount) ELSE 0 END), 0) AS total_withdraw,
                MIN(transaction_date) AS date_from,
                MAX(transaction_date) AS date_to,
                SUM(CASE WHEN match_status = 'matched' THEN 1 ELSE 0 END) AS matched_count,
                SUM(CASE WHEN match_status = 'unmatched' THEN 1 ELSE 0 END) AS unmatched_count,
                SUM(CASE WHEN is_duplicate = 1 THEN 1 ELSE 0 END) AS duplicate_count
            FROM transactions
            WHERE statement_id = ?
            """,
            (sid,),
        ).fetchone()
        db.execute(
            """
            UPDATE statements
            SET row_count = ?, deposit_count = ?, withdraw_count = ?,
                total_deposit = ?, total_withdraw = ?, date_from = ?, date_to = ?,
                matched_count = ?, unmatched_count = ?, duplicate_count = ?
            WHERE id = ?
            """,
            (
                stats["row_count"] or 0,
                stats["deposit_count"] or 0,
                stats["withdraw_count"] or 0,
                stats["total_deposit"] or 0,
                stats["total_withdraw"] or 0,
                stats["date_from"],
                stats["date_to"],
                stats["matched_count"] or 0,
                stats["unmatched_count"] or 0,
                stats["duplicate_count"] or 0,
                sid,
            ),
        )


def audit(db: PgConnection, action: str, entity_type: str, entity_id: str | int, detail: str = "") -> None:
    db.execute(
        """
        INSERT INTO statement_audit_logs (action, entity_type, entity_id, detail)
        VALUES (?, ?, ?, ?)
        """,
        (action, entity_type, str(entity_id), detail[:1000] or None),
    )


def normalise_header(value: Any) -> str:
    return re.sub(r"[\s_\-./()]+", "", str(value or "").strip().lower())


REFERENCE_ALIASES = {
    "reference": {
        "reference",
        "ref",
        "odcode",
        "ordercode",
        "paymentref",
        "เลขอ้างอิง",
        "เลขเอกสาร",
        "เลขออเดอร์",
        "ออเดอร์",
    },
    "date": {"date", "paymentdate", "transactiondate", "วันที่", "วันที่โอน", "วันที่จ่าย"},
    "time": {"time", "paymenttime", "เวลา", "เวลาโอน", "เวลาจ่าย"},
    "amount": {"amount", "payprice", "paymentamount", "ยอด", "ยอดเงิน", "จำนวนเงิน", "ยอดโอน", "ยอดจ่าย"},
    "deposit": {"deposit", "เงินเข้า", "ยอดเงินเข้า", "ยอดรับ", "รับเงิน", "ยอดฝาก"},
    "withdraw": {"withdraw", "withdrawal", "เงินออก", "ยอดเงินออก", "ยอดถอน", "ถอนเงิน"},
    "party_name": {"name", "payname", "customername", "cmname", "ผู้โอน", "ชื่อผู้โอน", "ลูกค้า", "ชื่อลูกค้า"},
    "description": {"description", "details", "detail", "รายการ", "รายละเอียด"},
    "has_attachment": {"hasattachment", "attachment", "slip", "เอกสารแนบ", "สลิป", "มีสลิป"},
    "notes": {"note", "notes", "remark", "หมายเหตุ"},
}


def reference_column_map(row: list[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    normalised = [normalise_header(value) for value in row]
    for field, aliases in REFERENCE_ALIASES.items():
        for index, value in enumerate(normalised):
            if value in aliases:
                result[field] = index
                break
    return result


def parse_bool(value: Any) -> int:
    text = str(value or "").strip().lower()
    if text in {"1", "true", "yes", "y", "มี", "มีเอกสาร", "มีสลิป", "ครบ"}:
        return 1
    return 0


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp874", "tis-620"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("ไม่สามารถอ่าน encoding ของไฟล์ได้")


def parse_reference_rows(rows: list[list[Any]], source_filename: str | None = None) -> list[dict[str, Any]]:
    header_index = -1
    columns: dict[str, int] = {}
    for index, row in enumerate(rows[:25]):
        mapping = reference_column_map(row)
        if "amount" in mapping or "deposit" in mapping or "withdraw" in mapping:
            header_index = index
            columns = mapping
            break
    if header_index < 0:
        raise ValueError("ไม่พบหัวตารางฝั่งเปรียบเทียบ ต้องมีคอลัมน์ยอดเงิน เช่น Amount, Deposit, Withdraw, ยอดเงิน")

    def cell(row: list[Any], field: str) -> Any:
        index = columns.get(field)
        return row[index] if index is not None and index < len(row) else None

    items: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not row or all(value in (None, "") for value in row):
            continue
        reference = str(cell(row, "reference") or "").strip()
        amount = parse_amount(cell(row, "amount"))
        if amount is None:
            deposit = parse_amount(cell(row, "deposit"))
            withdraw = parse_amount(cell(row, "withdraw"))
            if deposit not in (None, 0):
                amount = abs(deposit)
            elif withdraw not in (None, 0):
                amount = -abs(withdraw)
        if amount is None:
            continue
        transaction_date = parse_date(cell(row, "date"))
        transaction_time = parse_time(cell(row, "time"))
        party_name = str(cell(row, "party_name") or cell(row, "description") or "").strip()[:200] or None
        notes = str(cell(row, "notes") or "").strip()[:1000] or None
        if not reference:
            date_part = (transaction_date or "NO-DATE").replace("-", "")
            time_part = (transaction_time or "NO-TIME").replace(":", "")[:6]
            amount_part = str(abs(money_cents(amount)))
            reference = f"AUTO-{date_part}-{time_part}-{amount_part}-{len(items) + 1}"
        row_hash = hashlib.sha256(
            "|".join(
                [
                    reference,
                    transaction_date or "",
                    transaction_time or "",
                    f"{round(float(amount), 2):.2f}",
                    party_name or "",
                ]
            ).encode("utf-8")
        ).hexdigest()
        items.append(
            {
                "source_filename": source_filename,
                "reference": reference[:200],
                "transaction_date": transaction_date,
                "transaction_time": transaction_time,
                "amount": round(float(amount), 2),
                "party_name": party_name,
                "has_attachment": parse_bool(cell(row, "has_attachment")),
                "notes": notes,
                "row_hash": row_hash,
            }
        )
    if not items:
        raise ValueError("อ่านไฟล์ได้แต่ไม่พบรายการฝั่งเปรียบเทียบที่มีค่ายอดเงิน ระบบไม่บังคับ Reference แต่ต้องมี Amount, Deposit หรือ Withdraw")
    return items


def parse_reference_upload(filename: str, data: bytes) -> list[dict[str, Any]]:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        text = decode_text(data)
        sample = text[:8192]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        return parse_reference_rows([list(row) for row in csv.reader(io.StringIO(text), dialect)], filename)
    if extension == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        errors: list[str] = []
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            try:
                return parse_reference_rows(rows, filename)
            except ValueError as exc:
                errors.append(f"{sheet.title}: {exc}")
        raise ValueError("ไม่พบตารางฝั่งเปรียบเทียบในไฟล์ Excel — " + "; ".join(errors))
    if extension == ".pdf":
        import pdfplumber

        errors: list[str] = []
        with pdfplumber.open(io.BytesIO(data)) as document:
            for page_number, page in enumerate(document.pages, start=1):
                for table_index, table in enumerate(page.extract_tables() or [], start=1):
                    rows = [list(row) for row in table if row]
                    try:
                        return parse_reference_rows(rows, filename)
                    except ValueError as exc:
                        errors.append(f"หน้า {page_number} ตาราง {table_index}: {exc}")
        detail = " — " + "; ".join(errors[:3]) if errors else ""
        raise ValueError("ไม่พบตารางฝั่งเปรียบเทียบในไฟล์ PDF ต้องเป็น PDF แบบ text/table ไม่ใช่ภาพสแกน" + detail)
    raise ValueError("รองรับไฟล์ฝั่งเปรียบเทียบเฉพาะ .csv, .xlsx และ .pdf")


COMMON_TEXT_TOKENS = {
    "รับ",
    "โอน",
    "จาก",
    "รับโอน",
    "รับโอนจาก",
    "ธนาคาร",
    "นาย",
    "นาง",
    "นางสาว",
    "บริษัท",
    "จำกัด",
    "co",
    "ltd",
    "company",
    "payment",
    "transfer",
}

BANK_ALIASES = {
    "GSB": {"gsb", "ออมสิน"},
    "KTB": {"ktb", "กรุงไทย"},
    "SCB": {"scb", "ไทยพาณิชย์"},
    "KBANK": {"kbank", "kasikorn", "กสิกร"},
    "BBL": {"bbl", "bangkokbank", "กรุงเทพ"},
    "BAY": {"bay", "krungsri", "กรุงศรี"},
    "TTB": {"ttb", "tmb", "ธนชาต", "ทหารไทย"},
    "UOB": {"uob"},
    "CIMB": {"cimb"},
    "KKP": {"kkp", "เกียรตินาคิน"},
}


def parse_iso_date(value: Any) -> date | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def normalise_match_text(*values: Any) -> str:
    text = " ".join(str(value or "") for value in values)
    return re.sub(r"\s+", " ", text.strip().lower())


# ── OCR-error normalization ────────────────────────────────────────────────────
# Characters that OCR commonly confuses (uppercase → canonical digit/letter)
_OCR_NORM = str.maketrans("OILSBZGAB", "011582648")


def _ocr_normalize(text: str) -> str:
    """Replace common OCR look-alike characters so NJSA6KMPM2 ≈ NJSA6KMPM2."""
    return text.upper().translate(_OCR_NORM)


def _edit_distance(a: str, b: str) -> int:
    """Levenshtein distance between two strings (O(n²) is fine for short codes)."""
    if len(a) < len(b):
        a, b = b, a
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            curr.append(min(
                prev[j] + 1,
                curr[j - 1] + 1,
                prev[j - 1] + (0 if ca == cb else 1),
            ))
        prev = curr
    return prev[len(b)]


def _common_token_match(tx: dict[str, Any], ref: dict[str, Any], min_len: int = 6) -> tuple[int, str]:
    """Bidirectional check: any alphanumeric token (≥min_len chars, contains digit)
    from ANY tx field appearing in ANY ref field, or vice versa.
    Example: tx.description='FACEBK *U4H69MVPM2', ref.party_name='U4H69MVPM2' → match."""
    tx_all = " ".join(str(tx.get(k) or "") for k in
                      ("description", "tr_code", "channel", "reference", "notes")).upper()
    ref_all = " ".join(str(ref.get(k) or "") for k in
                       ("reference", "party_name", "notes")).upper()
    pat = re.compile(r"[A-Z0-9]{" + str(min_len) + r",}")

    # Direction 1: token from tx found somewhere in ref text
    for tok in pat.findall(tx_all):
        if any(c.isdigit() for c in tok) and tok in ref_all:
            return 40, f"พบรหัส '{tok}' ในข้อมูล Ref"

    # Direction 2: token from ref found somewhere in tx text
    for tok in pat.findall(ref_all):
        if any(c.isdigit() for c in tok) and tok in tx_all:
            return 40, f"พบรหัส '{tok}' ในข้อมูล Statement"

    return 0, ""


def _fuzzy_ref_score(tx: dict[str, Any], ref: dict[str, Any]) -> tuple[int, str]:
    """
    Score the reference code from `ref` against ALL text fields in `tx`.
    Handles: exact substring, OCR-normalized substring, per-token edit distance ≤ 2,
    and longest-common-substring overlap ≥ 60% of the reference length.
    """
    ref_code = str(ref.get("reference") or "").strip()
    if not ref_code or len(ref_code) < 4:
        return 0, ""

    tx_haystack = " ".join(str(tx.get(f) or "") for f in
                           ("description", "reference", "notes", "tr_code", "channel"))
    ref_up = ref_code.upper()
    hay_up = tx_haystack.upper()

    # 1. Exact substring
    if ref_up in hay_up:
        return 40, f"รหัส '{ref_code}' ตรงทุกตัว"

    # 2. OCR-normalized substring
    ref_norm = _ocr_normalize(ref_up)
    hay_norm = _ocr_normalize(hay_up)
    if ref_norm in hay_norm:
        return 35, f"รหัส '{ref_code}' ตรงหลัง OCR normalize"

    # 3. Per-token edit distance (only for codes ≤ 20 chars)
    if len(ref_up) <= 20:
        toks    = re.findall(r"[A-Z0-9]{4,}", hay_up)
        toks_n  = re.findall(r"[A-Z0-9]{4,}", hay_norm)
        best, best_label = 99, ""
        for tok, tok_n in [(t, _ocr_normalize(t)) for t in toks] + \
                          [(t, t) for t in toks_n]:
            for r_cmp, t_cmp, normalized in (
                (ref_up,   tok,   False),
                (ref_norm, tok_n, True),
            ):
                if abs(len(t_cmp) - len(r_cmp)) > 3:
                    continue
                d = _edit_distance(r_cmp, t_cmp)
                if d < best:
                    best = d
                    best_label = f"{'OCR ' if normalized else ''}ต่างกัน {d} ตัว ('{tok}')"
        if best <= 1:
            return 32, f"รหัสใกล้เคียง {best_label}"
        if best <= 2:
            return 22, f"รหัสใกล้เคียง {best_label}"

    # 4. Longest-common-substring ≥ 60% of ref length
    min_overlap = max(4, int(len(ref_up) * 0.60))
    for tok in re.findall(r"[A-Z0-9]{4,}", hay_up):
        m = SequenceMatcher(None, ref_up, tok).find_longest_match(
            0, len(ref_up), 0, len(tok)
        )
        if m.size >= min_overlap:
            return 15, f"รหัสซ้อนทับ {m.size}/{len(ref_up)} ตัว ('{tok}')"

    return 0, ""


def match_tokens(*values: Any) -> set[str]:
    text = normalise_match_text(*values)
    tokens = {
        token
        for token in re.findall(r"[a-z0-9]+|[ก-๙]+", text)
        if len(token) >= 2 and token not in COMMON_TEXT_TOKENS
    }
    for canonical, aliases in BANK_ALIASES.items():
        if any(alias in text for alias in aliases):
            tokens.add(canonical.lower())
    return tokens


def date_match_score(statement_date: Any, reference_date: Any) -> tuple[int, str]:
    tx_date = parse_iso_date(statement_date)
    ref_date = parse_iso_date(reference_date)
    if not tx_date or not ref_date:
        return 0, "ไม่มีวันที่ครบ"
    if tx_date == ref_date:
        return 25, "วันเดียวกัน ✓"
    # หลังจากเพิ่ม hard filter แล้ว บรรทัดด้านล่างนี้จะไม่ถูกเรียกอีก
    diff_days = abs((tx_date - ref_date).days)
    return 0, f"วันที่ห่าง {diff_days} วัน"


def text_match_score(tx: dict[str, Any], ref: dict[str, Any]) -> tuple[int, str]:
    tx_text = normalise_match_text(tx["description"], tx["reference"], tx["notes"], tx["channel"])
    ref_text = normalise_match_text(ref["party_name"], ref["reference"], ref["notes"])
    tx_tokens = match_tokens(tx_text)
    ref_tokens = match_tokens(ref_text)
    shared = sorted(tx_tokens & ref_tokens)
    score = 0
    reasons: list[str] = []

    bank_hits = [token.upper() for token in shared if token in {key.lower() for key in BANK_ALIASES}]
    if bank_hits:
        score += 20
        reasons.append("ธนาคารตรง " + ", ".join(bank_hits[:2]))

    name_hits = [token for token in shared if token not in {key.lower() for key in BANK_ALIASES}]
    if name_hits:
        score += min(25, 10 + len(name_hits) * 5)
        reasons.append("ข้อความ/ชื่อตรง " + ", ".join(name_hits[:3]))

    if tx_text and ref_text:
        similarity = SequenceMatcher(None, tx_text, ref_text).ratio()
        if similarity >= 0.75:
            score += 15
            reasons.append("รายละเอียดคล้ายมาก")
        elif similarity >= 0.45:
            score += 8
            reasons.append("รายละเอียดคล้ายบางส่วน")

    if not reasons:
        reasons.append("ยังไม่พบชื่อหรือธนาคารที่ตรง")
    return score, " · ".join(reasons)


def score_reference_candidate(tx: dict[str, Any], ref: dict[str, Any]) -> dict[str, Any]:
    date_score, date_reason = date_match_score(tx["transaction_date"], ref["transaction_date"])
    text_score, text_reason = text_match_score(tx, ref)
    fuzzy_score, fuzzy_reason = _fuzzy_ref_score(tx, ref)
    token_score, token_reason = _common_token_match(tx, ref)
    # take the best code-matching signal
    if token_score >= fuzzy_score:
        best_code_score, best_code_reason = token_score, token_reason
    else:
        best_code_score, best_code_reason = fuzzy_score, fuzzy_reason
    combined_text = max(text_score + best_code_score, text_score)
    score = 45 + date_score + combined_text
    result = dict(ref)
    result["match_score"] = max(0, min(100, score))
    parts = ["ยอดตรงเป๊ะ", date_reason, text_reason]
    if best_code_reason:
        parts.append(best_code_reason)
    result["match_reason"] = " · ".join(p for p in parts if p)
    return result


def candidate_reference_items(db: PgConnection, tx: dict[str, Any], limit: int = 5) -> list[dict[str, Any]]:
    target_cents = abs(money_cents(tx["amount"]))  # ใช้ abs เสมอ: statement เก็บรายจ่ายเป็นลบ
    tx_date = parse_iso_date(tx["transaction_date"])
    rows = db.execute(
        """
        SELECT *
        FROM reference_items
        WHERE match_status = 'unmatched'
        ORDER BY transaction_date IS NULL, transaction_date, id
        LIMIT 800
        """
    ).fetchall()
    scored = []
    for row in rows:
        if abs(abs(money_cents(row["amount"])) - target_cents) > 1:
            continue  # ยอดต้องตรงเสมอ (เปรียบเทียบค่าสัมบูรณ์)
        date_ok = tx_date is not None and parse_iso_date(row["transaction_date"]) == tx_date
        token_ok = _common_token_match(tx, row)[0] > 0
        if not date_ok and not token_ok:
            continue  # ต้องผ่านอย่างน้อยหนึ่งเงื่อนไข: วันที่ตรง หรือ มี token ร่วมกัน
        scored.append(score_reference_candidate(tx, row))
    scored.sort(key=lambda row: (row["match_score"], row["transaction_date"] or "", row["id"]), reverse=True)
    return scored[:limit]


def create_match_group(
    db: PgConnection,
    tx_rows: list[dict[str, Any]],
    *,
    reference: str,
    expected_cents: int,
    has_attachment: int,
    method: str,
    notes: str = "",
    reference_item_id: int | None = None,
) -> int:
    ids = [int(row["id"]) for row in tx_rows]
    total_cents = abs(sum(money_cents(row["amount"]) for row in tx_rows))
    expected_cents = abs(expected_cents)
    if total_cents != expected_cents:
        raise ValueError(
            f"ยอดรวม Statement ไม่ตรงกับยอดอ้างอิง ({total_cents / 100:,.2f} != {expected_cents / 100:,.2f})"
        )
    match_type = "group" if len(ids) > 1 else "single"
    cursor = db.execute(
        """
        INSERT INTO match_groups (
            match_type, reference_item_id, target_reference, expected_amount,
            statement_total, has_attachment, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        RETURNING id
        """,
        (
            match_type,
            reference_item_id,
            reference,
            expected_cents / 100,
            total_cents / 100,
            1 if has_attachment else 0,
            notes.strip()[:1000],
        ),
    )
    group_id = int(cursor.fetchone()["id"])
    db.executemany(
        """
        INSERT INTO match_group_items (match_group_id, transaction_id)
        VALUES (?, ?)
        """,
        [(group_id, tx_id) for tx_id in ids],
    )
    placeholders = ",".join("?" for _ in ids)
    note = notes.strip()[:1000]
    db.execute(
        f"""
        UPDATE transactions
        SET match_status = 'matched',
            match_group_id = ?,
            match_method = ?,
            reference = ?,
            has_attachment = ?,
            notes = CASE
                WHEN ? = '' THEN notes
                WHEN notes IS NULL OR notes = '' THEN ?
                ELSE notes || chr(10) || ?
            END,
            updated_at = CURRENT_TIMESTAMP
        WHERE id IN ({placeholders})
        """,
        [group_id, method, reference, 1 if has_attachment else 0, note, note, note, *ids],
    )
    if reference_item_id:
        db.execute(
            """
            UPDATE reference_items
            SET match_status = 'matched', updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (reference_item_id,),
        )
    audit(
        db,
        "confirm_match",
        "match_group",
        group_id,
        f"{method}: tx={','.join(map(str, ids))}; reference={reference}; amount={expected_cents / 100:,.2f}",
    )
    for statement_id in {int(row["statement_id"]) for row in tx_rows}:
        sync_statement_stats(db, statement_id)
    return group_id


@app.get("/health")
async def health() -> JSONResponse:
    return JSONResponse({"status": "ok", "service": "credit-statement-matcher"})


@app.get("/", include_in_schema=False)
async def index() -> RedirectResponse:
    return redirect("/transactions")


@app.get("/upload", response_class=HTMLResponse)
async def upload_page(request: Request) -> HTMLResponse:
    cleanup_expired_previews(UPLOAD_DIR)
    with closing(connect()) as db:
        statements = db.execute(
            "SELECT * FROM statements ORDER BY uploaded_at DESC, id DESC LIMIT 10"
        ).fetchall()
    return page(request, "upload.html", active="upload", statements=statements)


async def read_upload(file: UploadFile) -> bytes:
    chunks: list[bytes] = []
    size = 0
    while chunk := await file.read(1024 * 1024):
        size += len(chunk)
        if size > MAX_UPLOAD_BYTES:
            raise ValueError("ไฟล์มีขนาดเกิน 25 MB")
        chunks.append(chunk)
    if not chunks:
        raise ValueError("ไฟล์ว่างเปล่า")
    return b"".join(chunks)


def cleanup_upload_jobs(*, now: float | None = None) -> int:
    current_time = time.time() if now is None else now
    expired = [
        token
        for token, job in UPLOAD_JOBS.items()
        if job.get("status") in {"complete", "failed"}
        and current_time - float(job.get("updated_at") or 0)
        > UPLOAD_JOB_TTL_SECONDS
    ]
    for token in expired:
        UPLOAD_JOBS.pop(token, None)
    return len(expired)


def upload_job_progress(job: dict[str, Any], *, now: float | None = None) -> dict[str, Any]:
    current_time = time.time() if now is None else now
    elapsed = max(0, int(current_time - float(job.get("created_at") or current_time)))
    status = str(job.get("status") or "queued")
    if status == "queued":
        step = 0
        message = "ไฟล์อยู่ในคิวและกำลังรอเริ่มอ่าน"
    elif status == "processing":
        processing_elapsed = max(
            0,
            int(current_time - float(job.get("processing_started_at") or current_time)),
        )
        if processing_elapsed < 8:
            step = 1
            message = "กำลังตรวจ text layer และโครงสร้าง Statement"
        elif processing_elapsed < 45:
            step = 2
            message = "กำลังแปลง PDF เป็นภาพและทำ OCR ไทย + อังกฤษ"
        else:
            step = 2
            message = "OCR กำลังอ่านเอกสารหลายหน้า กรุณารอสักครู่"
    elif status == "complete":
        step = 3
        message = "อ่านไฟล์เสร็จแล้ว กำลังเปิดตาราง Preview"
    else:
        step = 3
        message = "ประมวลผลไม่สำเร็จ"
    result: dict[str, Any] = {
        "status": status,
        "step": step,
        "message": message,
        "elapsed_seconds": elapsed,
    }
    if status == "complete" and job.get("preview_token"):
        result["redirect_url"] = (
            f"{ROOT_PATH}/preview/{job['preview_token']}"
        )
    if status == "failed":
        result["error"] = str(job.get("error") or "ไม่สามารถอ่าน Statement ได้")
    return result


async def process_upload_job(
    job_token: str,
    original_name: str,
    data: bytes,
) -> None:
    job = UPLOAD_JOBS.get(job_token)
    if not job:
        return
    semaphore = OCR_JOB_SEMAPHORE
    if semaphore is None:
        job.update(
            status="failed",
            error="ระบบประมวลผลยังไม่พร้อม กรุณาลองใหม่",
            updated_at=time.time(),
        )
        return
    try:
        async with semaphore:
            job.update(
                status="processing",
                processing_started_at=time.time(),
                updated_at=time.time(),
            )
            statement = await asyncio.to_thread(
                parse_statement_with_metadata,
                original_name,
                data,
            )
            preview_token = await asyncio.to_thread(
                create_preview,
                UPLOAD_DIR,
                original_name,
                data,
                statement,
            )
            job.update(
                status="complete",
                preview_token=preview_token,
                updated_at=time.time(),
            )
    except ValueError as exc:
        job.update(status="failed", error=str(exc), updated_at=time.time())
    except asyncio.CancelledError:
        job.update(
            status="failed",
            error="การประมวลผลถูกหยุด กรุณาอัปโหลดใหม่",
            updated_at=time.time(),
        )
        raise
    except Exception:
        job.update(
            status="failed",
            error="เกิดข้อผิดพลาดระหว่างอ่าน Statement กรุณาลองใหม่",
            updated_at=time.time(),
        )


ALLOWED_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
MAX_IMAGE_FILES = 100
MAX_IMAGE_BYTES_PER_FILE = 15 * 1024 * 1024  # 15 MB per image


async def process_image_upload_job(
    job_token: str,
    original_name: str,
    bundle_data: bytes,
    image_items: list[tuple[str, bytes]],
) -> None:
    job = UPLOAD_JOBS.get(job_token)
    if not job:
        return
    semaphore = OCR_JOB_SEMAPHORE
    if semaphore is None:
        job.update(
            status="failed",
            error="ระบบประมวลผลยังไม่พร้อม กรุณาลองใหม่",
            updated_at=time.time(),
        )
        return
    try:
        async with semaphore:
            job.update(
                status="processing",
                processing_started_at=time.time(),
                updated_at=time.time(),
            )
            statement = await parse_images_with_vision_model(image_items)
            preview_token = await asyncio.to_thread(
                create_preview,
                UPLOAD_DIR,
                original_name,
                bundle_data,
                statement,
            )
            job.update(
                status="complete",
                preview_token=preview_token,
                updated_at=time.time(),
            )
    except ValueError as exc:
        job.update(status="failed", error=str(exc), updated_at=time.time())
    except asyncio.CancelledError:
        job.update(
            status="failed",
            error="การประมวลผลถูกหยุด กรุณาอัปโหลดใหม่",
            updated_at=time.time(),
        )
        raise
    except Exception:
        job.update(
            status="failed",
            error="เกิดข้อผิดพลาดระหว่างอ่านรูปภาพ กรุณาลองใหม่",
            updated_at=time.time(),
        )


def queue_upload_job(original_name: str, data: bytes, digest: str) -> str:
    cleanup_upload_jobs()
    active_jobs = [
        job
        for job in UPLOAD_JOBS.values()
        if job.get("status") in {"queued", "processing"}
    ]
    if len(active_jobs) >= MAX_PENDING_UPLOAD_JOBS:
        raise ValueError("ขณะนี้มีไฟล์ OCR อยู่ในคิวหลายรายการ กรุณารอสักครู่แล้วลองใหม่")
    if any(job.get("file_sha256") == digest for job in active_jobs):
        raise ValueError("ไฟล์นี้กำลังประมวลผลอยู่แล้ว กรุณารอผลจากหน้าสถานะเดิม")
    token = uuid.uuid4().hex
    created_at = time.time()
    UPLOAD_JOBS[token] = {
        "token": token,
        "original_name": original_name,
        "file_sha256": digest,
        "status": "queued",
        "created_at": created_at,
        "updated_at": created_at,
    }
    task = asyncio.create_task(process_upload_job(token, original_name, data))
    UPLOAD_TASKS.add(task)
    task.add_done_callback(UPLOAD_TASKS.discard)
    return token


def _preview_rows(
    statement: dict[str, Any],
    submitted: dict[int, dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    # Ads-evidence screenshots (statement_type == "ads_screenshot") land in
    # reference_items, whose transaction_date column is nullable — unlike a
    # real statement's transactions.transaction_date (NOT NULL). Real TikTok
    # screenshots sometimes don't show a date per row at all, so requiring
    # one there would silently discard otherwise-valid rows.
    is_ads_screenshot = statement.get("statement_type") == "ads_screenshot"
    rows: list[dict[str, Any]] = []
    for index, original in enumerate(statement.get("transactions") or []):
        row = dict(original)
        valid = bool(
            (row.get("transaction_date") or is_ads_screenshot)
            and row.get("description")
            and row.get("amount") is not None
        )
        row.update(
            {
                "index": index,
                "include": valid,
                "reviewed": False,
                "requires_review": (
                    float(row.get("confidence") or 0) < 60 or not valid
                ),
                "row_errors": [],
            }
        )
        if submitted and index in submitted:
            row.update(submitted[index])
        rows.append(row)
    return rows


def _render_preview(
    request: Request,
    payload: dict[str, Any],
    *,
    rows: list[dict[str, Any]] | None = None,
    error: str | None = None,
    status_code: int = 200,
) -> HTMLResponse:
    statement = dict(payload["statement"])
    return page(
        request,
        "preview.html",
        active="upload",
        preview_token=payload["token"],
        original_name=payload["original_name"],
        statement=statement,
        rows=rows or _preview_rows(statement),
        error=error,
        status_code=status_code,
    )


def _submitted_preview_rows(
    statement: dict[str, Any], form: Any
) -> tuple[list[ParsedTransaction], list[dict[str, Any]], list[str]]:
    # See _preview_rows — same "date optional for ads_screenshot" exception.
    is_ads_screenshot = statement.get("statement_type") == "ads_screenshot"
    parsed: list[ParsedTransaction] = []
    submitted: dict[int, dict[str, Any]] = {}
    errors: list[str] = []
    originals = list(statement.get("transactions") or [])
    for index, original in enumerate(originals):
        include = str(form.get(f"include_{index}") or "") == "1"
        raw_date = str(form.get(f"transaction_date_{index}") or "").strip()
        raw_description = str(form.get(f"description_{index}") or "").strip()
        raw_amount = str(form.get(f"amount_{index}") or "").strip()
        raw_last4 = re.sub(
            r"\D", "", str(form.get(f"card_last4_{index}") or "")
        )[-4:]
        reviewed = str(form.get(f"reviewed_{index}") or "") == "1"
        transaction_date = parse_date(raw_date)
        amount = parse_amount(raw_amount)
        row_errors: list[str] = []
        requires_review = (
            float(original.get("confidence") or 0) < 60
            or (not original.get("transaction_date") and not is_ads_screenshot)
            or original.get("amount") is None
            or not str(original.get("description") or "").strip()
        )
        if include:
            if not transaction_date and not is_ads_screenshot:
                row_errors.append("วันที่ไม่ถูกต้อง")
            if not raw_description:
                row_errors.append("กรุณาระบุรายละเอียด")
            if amount is None:
                row_errors.append("จำนวนเงินไม่ถูกต้อง")
            if requires_review and not reviewed:
                row_errors.append("กรุณายืนยันว่าได้ตรวจทานแถวนี้แล้ว")
        raw_tr_code_form = str(form.get(f"tr_code_{index}") or original.get("tr_code") or "").strip()
        submitted[index] = {
            "include": include,
            "reviewed": reviewed,
            "requires_review": requires_review,
            "transaction_date": raw_date,
            "description": raw_description,
            "amount": raw_amount,
            "card_last4": raw_last4 or None,
            "tr_code": raw_tr_code_form[:50] or None,
            "row_errors": row_errors,
        }
        if row_errors:
            errors.append(f"แถว {index + 1}: {', '.join(row_errors)}")
            continue
        if not include:
            continue
        signed_amount = round(float(amount), 2)
        transaction_time = parse_time(original.get("transaction_time"))
        channel = str(original.get("channel") or "").strip()[:100] or None
        tr_code = raw_tr_code_form[:50] or None
        row_hash = make_row_hash(
            transaction_date or "",
            transaction_time,
            signed_amount,
            raw_description,
            channel,
            tr_code,
        )
        parsed.append(
            ParsedTransaction(
                transaction_date=transaction_date,
                description=raw_description[:500],
                amount=signed_amount,
                card_last4=raw_last4 or None,
                category=guess_category(raw_description, signed_amount),
                transaction_time=transaction_time,
                deposit_amount=signed_amount if signed_amount > 0 else None,
                withdraw_amount=abs(signed_amount) if signed_amount < 0 else None,
                channel=channel,
                tr_code=tr_code,
                row_hash=row_hash,
                source_page=original.get("source_page"),
                confidence=float(original.get("confidence") or 0),
                warnings=list(original.get("warnings") or []),
            )
        )
    return parsed, _preview_rows(statement, submitted), errors


def _save_as_reference_items(
    db: PgConnection,
    payload: dict[str, Any],
    transactions: list[ParsedTransaction],
) -> int:
    """Save TikTok Ads OCR results as reference_items (not transactions)."""
    source_filename = Path(str(payload.get("original_name") or "tiktok_ads")).name
    existing_hashes = {
        row["row_hash"]
        for row in db.execute(
            "SELECT row_hash FROM reference_items WHERE row_hash IS NOT NULL"
        ).fetchall()
    }
    inserted = 0
    for item in transactions:
        if item.row_hash and item.row_hash in existing_hashes:
            continue
        # reference = payment_reference code (e.g. NJSA6KMPM2); fallback to invoice_no
        reference = (item.tr_code or "").strip()
        if not reference:
            # extract first token from description (invoice_no part)
            reference = (item.description or "").split(" | ")[0].strip() or "TikTok"
        # party_name = invoice_no (first token in description)
        desc_parts = (item.description or "").split(" | ")
        party_name = desc_parts[0].strip() if desc_parts else None
        d = item.transaction_date
        date_str = d.isoformat() if hasattr(d, "isoformat") else (str(d)[:10] if d else None)
        t = item.transaction_time
        time_str = (
            t.strftime("%H:%M:%S") if hasattr(t, "strftime")
            else (str(t)[:8] if t else None)
        )
        channel_name = item.channel or source_filename
        db.execute(
            """
            INSERT INTO reference_items (
                source_filename, reference, transaction_date, transaction_time,
                amount, party_name, notes, row_hash
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                channel_name,
                reference,
                date_str,
                time_str,
                item.amount,
                party_name,
                item.description,
                item.row_hash,
            ),
        )
        inserted += 1
        if item.row_hash:
            existing_hashes.add(item.row_hash)
    audit(
        db,
        "upload_images_as_references",
        "reference_items",
        source_filename,
        f"inserted {inserted} rows from TikTok Ads screenshots",
    )
    db.commit()
    return inserted


def _save_confirmed_preview(
    db: PgConnection,
    payload: dict[str, Any],
    data: bytes,
    transactions: list[ParsedTransaction],
) -> int:
    digest = str(payload["file_sha256"])
    duplicate = db.execute(
        "SELECT id FROM statements WHERE file_sha256 = ?", (digest,)
    ).fetchone()
    if duplicate:
        raise ValueError("ไฟล์นี้เคยถูกอัปโหลดแล้ว")

    original_name = Path(str(payload["original_name"])).name
    statement = dict(payload["statement"])
    stored_name = f"{uuid.uuid4().hex}{Path(original_name).suffix.lower()}"
    stored_path = UPLOAD_DIR / stored_name
    stored_path.write_bytes(data)
    try:
        cursor = db.execute(
            """
            INSERT INTO statements (
                original_filename, stored_filename, file_sha256, row_count,
                issuer, statement_type, processing_method, masked_reference,
                parse_warnings
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            RETURNING id
            """,
            (
                original_name,
                stored_name,
                digest,
                len(transactions),
                statement.get("issuer"),
                statement.get("statement_type"),
                statement.get("extraction_method"),
                statement.get("masked_reference"),
                "\n".join(statement.get("warnings") or [])[:2000] or None,
            ),
        )
        statement_id = int(cursor.fetchone()["id"])
        existing_hashes = {
            row["row_hash"]
            for row in db.execute(
                "SELECT row_hash FROM transactions WHERE row_hash IS NOT NULL"
            ).fetchall()
        }
        seen_hashes: set[str] = set()
        insert_rows: list[tuple[Any, ...]] = []
        for item in transactions:
            duplicate_row = bool(
                item.row_hash
                and (
                    item.row_hash in existing_hashes
                    or item.row_hash in seen_hashes
                )
            )
            insert_rows.append(
                (
                    statement_id,
                    item.transaction_date,
                    item.description,
                    item.amount,
                    item.transaction_time,
                    item.deposit_amount,
                    item.withdraw_amount,
                    item.channel,
                    item.tr_code,
                    item.row_hash,
                    1 if duplicate_row else 0,
                    item.card_last4,
                    item.category,
                )
            )
            if item.row_hash:
                seen_hashes.add(item.row_hash)
        if insert_rows:
            db.executemany(
                """
                INSERT INTO transactions (
                    statement_id, transaction_date, description, amount,
                    transaction_time, deposit_amount, withdraw_amount, channel,
                    tr_code, row_hash, is_duplicate, card_last4, category
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                insert_rows,
            )
        audit(
            db,
            "confirm_statement_preview",
            "statement",
            statement_id,
            (
                f"parser={statement.get('issuer') or 'unknown'}; "
                f"method={statement.get('extraction_method') or 'unknown'}; "
                f"rows={len(transactions)}"
            ),
        )
        sync_statement_stats(db, statement_id)
        db.commit()
        return statement_id
    except Exception:
        stored_path.unlink(missing_ok=True)
        raise


@app.post("/upload", response_class=HTMLResponse)
async def upload_statement(request: Request, file: UploadFile = File(...)) -> Response:
    original_name = Path(file.filename or "statement").name
    try:
        data = await read_upload(file)
        digest = hashlib.sha256(data).hexdigest()
        with closing(connect()) as db:
            duplicate = db.execute(
                "SELECT id FROM statements WHERE file_sha256 = ?", (digest,)
            ).fetchone()
        if duplicate:
            raise ValueError("ไฟล์นี้เคยถูกอัปโหลดแล้ว")
        job_token = queue_upload_job(original_name, data, digest)
    except ValueError as exc:
        return page(
            request,
            "upload.html",
            active="upload",
            statements=[],
            error=str(exc),
            status_code=422,
        )
    return redirect(f"/upload/processing/{job_token}")


@app.get("/upload/processing/{job_token}", response_class=HTMLResponse)
async def upload_processing_page(
    request: Request,
    job_token: str,
) -> HTMLResponse:
    cleanup_upload_jobs()
    job = UPLOAD_JOBS.get(job_token)
    if not re.fullmatch(r"[0-9a-f]{32}", job_token) or not job:
        return page(
            request,
            "upload.html",
            active="upload",
            statements=[],
            error="ไม่พบงานประมวลผล กรุณาอัปโหลดไฟล์ใหม่",
            status_code=404,
        )
    return page(
        request,
        "processing.html",
        active="upload",
        job_token=job_token,
        original_name=job["original_name"],
    )


@app.get("/upload/jobs/{job_token}")
async def upload_job_status(job_token: str) -> JSONResponse:
    cleanup_upload_jobs()
    job = UPLOAD_JOBS.get(job_token)
    if not re.fullmatch(r"[0-9a-f]{32}", job_token) or not job:
        return JSONResponse(
            {"status": "missing", "error": "ไม่พบงานประมวลผล"},
            status_code=404,
            headers={"Cache-Control": "no-store"},
        )
    return JSONResponse(
        upload_job_progress(job),
        headers={"Cache-Control": "no-store"},
    )


@app.get("/preview/{preview_token}", response_class=HTMLResponse)
async def statement_preview_page(
    request: Request,
    preview_token: str,
) -> HTMLResponse:
    try:
        payload = load_preview(UPLOAD_DIR, preview_token)
    except PreviewNotFoundError as exc:
        return page(
            request,
            "upload.html",
            active="upload",
            statements=[],
            error=str(exc),
            status_code=404,
        )
    return _render_preview(request, payload)


@app.post("/upload/confirm", response_class=HTMLResponse)
async def confirm_statement_upload(request: Request) -> HTMLResponse:
    form = await request.form()
    token = str(form.get("preview_token") or "")
    try:
        payload = load_preview(UPLOAD_DIR, token)
        transactions, rows, errors = _submitted_preview_rows(
            dict(payload["statement"]), form
        )
        if errors:
            visible_errors = "; ".join(errors[:6])
            if len(errors) > 6:
                visible_errors += f"; และอีก {len(errors) - 6} แถว"
            return _render_preview(
                request,
                payload,
                rows=rows,
                error=(
                    f"พบ {len(errors)} แถวที่ยังยืนยันไม่ได้ — "
                    f"{visible_errors}"
                ),
                status_code=422,
            )
        statement_type = str(
            (payload.get("statement") or {}).get("statement_type") or ""
        )
        if statement_type == "ads_screenshot":
            # TikTok Ads screenshots → save as reference_items (not transactions)
            with closing(connect()) as db:
                inserted = _save_as_reference_items(db, payload, transactions)
            delete_preview(UPLOAD_DIR, token)
            return redirect(f"/references?imported=1&inserted={inserted}")
        data = read_preview_source(UPLOAD_DIR, payload)
        with closing(connect()) as db:
            statement_id = _save_confirmed_preview(
                db, payload, data, transactions
            )
        delete_preview(UPLOAD_DIR, token)
    except (PreviewNotFoundError, ValueError) as exc:
        return page(
            request,
            "upload.html",
            active="upload",
            statements=[],
            error=str(exc),
            status_code=422,
        )
    return redirect(f"/transactions?statement_id={statement_id}&uploaded=1")


@app.post("/upload/cancel")
async def cancel_statement_upload(
    preview_token: str = Form(...),
) -> RedirectResponse:
    try:
        delete_preview(UPLOAD_DIR, preview_token)
    except PreviewNotFoundError:
        pass
    return redirect("/upload")


@app.post("/upload/images", response_class=HTMLResponse)
async def upload_image_batch(
    request: Request,
    files: list[UploadFile] = File(...),
) -> Response:
    import json as _json

    try:
        valid = [f for f in files if f.filename]
        if not valid:
            raise ValueError("กรุณาเลือกไฟล์รูปภาพอย่างน้อย 1 ไฟล์")
        if len(valid) > MAX_IMAGE_FILES:
            raise ValueError(f"อัปโหลดได้ไม่เกิน {MAX_IMAGE_FILES} ไฟล์ต่อครั้ง")

        image_items: list[tuple[str, bytes]] = []
        filenames: list[str] = []

        for upload in valid:
            name = Path(upload.filename or "").name or "image.jpg"
            ext = Path(name).suffix.lower()
            if ext not in ALLOWED_IMAGE_EXTS:
                raise ValueError(f"ไฟล์ {name}: รองรับเฉพาะ JPG, PNG, WEBP")
            chunks: list[bytes] = []
            size = 0
            while chunk := await upload.read(1024 * 1024):
                size += len(chunk)
                if size > MAX_IMAGE_BYTES_PER_FILE:
                    raise ValueError(
                        f"ไฟล์ {name} มีขนาดเกิน {MAX_IMAGE_BYTES_PER_FILE // (1024 * 1024)} MB"
                    )
                chunks.append(chunk)
            data = b"".join(chunks)
            if not data:
                continue
            image_items.append((name, data))
            filenames.append(name)

        if not image_items:
            raise ValueError("ไม่พบไฟล์รูปภาพที่อ่านได้")

        cleanup_upload_jobs()
        active = [j for j in UPLOAD_JOBS.values() if j.get("status") in {"queued", "processing"}]
        if len(active) >= MAX_PENDING_UPLOAD_JOBS:
            raise ValueError("มีงานอยู่ในคิวหลายรายการ กรุณารอสักครู่แล้วลองใหม่")

        bundle = _json.dumps(
            {
                "type": "image_batch",
                "batch_id": uuid.uuid4().hex,
                "files": filenames,
                "count": len(filenames),
            },
            ensure_ascii=False,
        ).encode()

        token = uuid.uuid4().hex
        created_at = time.time()
        original_name = f"images_{len(filenames)}_files.imgbatch"
        UPLOAD_JOBS[token] = {
            "token": token,
            "original_name": original_name,
            "file_sha256": hashlib.sha256(bundle).hexdigest(),
            "status": "queued",
            "created_at": created_at,
            "updated_at": created_at,
        }
        task = asyncio.create_task(
            process_image_upload_job(token, original_name, bundle, image_items)
        )
        UPLOAD_TASKS.add(task)
        task.add_done_callback(UPLOAD_TASKS.discard)

    except ValueError as exc:
        with closing(connect()) as db:
            stmts = db.execute(
                "SELECT * FROM statements ORDER BY uploaded_at DESC, id DESC LIMIT 10"
            ).fetchall()
        return page(
            request,
            "upload.html",
            active="upload",
            statements=stmts,
            error=str(exc),
            status_code=422,
        )

    return redirect(f"/upload/processing/{token}")


@app.post("/statements/{statement_id}/delete")
async def delete_statement(statement_id: int) -> RedirectResponse:
    with closing(connect()) as db:
        statement = db.execute(
            "SELECT stored_filename FROM statements WHERE id = ?", (statement_id,)
        ).fetchone()
        if not statement:
            raise HTTPException(status_code=404, detail="Statement not found")
        db.execute("DELETE FROM statements WHERE id = ?", (statement_id,))
        db.commit()
    (UPLOAD_DIR / statement["stored_filename"]).unlink(missing_ok=True)
    return redirect("/upload")


@app.get("/review", response_class=HTMLResponse)
async def review_page(
    request: Request,
    statement_id: int | None = None,
    issue: str | None = None,
    uploaded: int | None = None,
) -> HTMLResponse:
    with closing(connect()) as db:
        statements = db.execute(
            "SELECT * FROM statements ORDER BY uploaded_at DESC, id DESC"
        ).fetchall()
        if statement_id is None and statements:
            statement_id = int(statements[0]["id"])

        clauses: list[str] = []
        values: list[Any] = []
        if statement_id:
            clauses.append("t.statement_id = ?")
            values.append(statement_id)
        if issue == "unmatched":
            clauses.append("t.match_status = 'unmatched'")
        elif issue == "duplicates":
            clauses.append("t.is_duplicate = 1")
        elif issue == "missing-attachments":
            clauses.append("t.match_status = 'matched' AND t.has_attachment = 0")
        elif issue == "matched":
            clauses.append("t.match_status = 'matched'")

        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        rows = db.execute(
            f"""
            SELECT t.*, s.original_filename,
                   mg.target_reference, mg.reference_item_id, mg.notes AS match_notes,
                   t.match_method,
                   ri.source_filename AS ref_source_filename,
                   ri.transaction_date AS ref_date,
                   ri.party_name AS ref_party_name,
                   ri.reference AS ref_ocr_reference
            FROM transactions t
            JOIN statements s ON s.id = t.statement_id
            LEFT JOIN match_groups mg ON mg.id = t.match_group_id
            LEFT JOIN reference_items ri ON ri.id = mg.reference_item_id
            {where}
            ORDER BY
                CASE t.match_status WHEN 'unmatched' THEN 0 WHEN 'matched' THEN 1 ELSE 2 END,
                t.is_duplicate DESC,
                t.transaction_date DESC,
                t.transaction_time DESC,
                t.id DESC
            LIMIT 250
            """,
            values,
        ).fetchall()
        totals = db.execute(
            f"""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN t.match_status = 'matched' THEN 1 ELSE 0 END) AS matched,
                SUM(CASE WHEN t.match_status = 'unmatched' THEN 1 ELSE 0 END) AS unmatched,
                SUM(CASE WHEN t.match_status = 'ignored' THEN 1 ELSE 0 END) AS ignored,
                SUM(CASE WHEN t.is_duplicate = 1 THEN 1 ELSE 0 END) AS duplicates,
                SUM(CASE WHEN t.match_status = 'matched' AND t.has_attachment = 0 THEN 1 ELSE 0 END) AS missing_attachments,
                COALESCE(SUM(CASE WHEN t.amount > 0 THEN t.amount ELSE 0 END), 0) AS deposits
            FROM transactions t
            {"WHERE t.statement_id = ?" if statement_id else ""}
            """,
            (statement_id,) if statement_id else (),
        ).fetchone()
        ref_stats = db.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN match_status = 'matched' THEN 1 ELSE 0 END) AS matched,
                SUM(CASE WHEN match_status = 'unmatched' THEN 1 ELSE 0 END) AS unmatched,
                SUM(CASE WHEN has_attachment = 0 THEN 1 ELSE 0 END) AS missing_attachments
            FROM reference_items
            """
        ).fetchone()
        candidates_by_tx: dict[int, list[dict[str, Any]]] = {}
        for row in rows:
            if row["match_status"] == "unmatched" and not row["is_duplicate"]:
                candidates_by_tx[int(row["id"])] = candidate_reference_items(db, row, limit=3)

    return page(
        request,
        "review.html",
        active="review",
        statements=statements,
        selected_statement_id=statement_id,
        issue=issue or "",
        uploaded=bool(uploaded),
        rows=rows,
        totals=totals,
        ref_stats=ref_stats,
        candidates_by_tx=candidates_by_tx,
    )


def transaction_filters(
    statement_id: int | None,
    status: str | None,
    card: str | None,
    query: str | None,
) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    values: list[Any] = []
    if statement_id:
        clauses.append("t.statement_id = ?")
        values.append(statement_id)
    if status in {"unmatched", "matched", "ignored"}:
        clauses.append("t.match_status = ?")
        values.append(status)
    if card:
        clauses.append("t.card_last4 = ?")
        values.append(re.sub(r"\D", "", card)[-4:])
    if query:
        clauses.append("(t.description LIKE ? OR t.reference LIKE ? OR t.notes LIKE ?)")
        term = f"%{query.strip()}%"
        values.extend((term, term, term))
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    return where, values


SUMMARY_PLATFORM_LABELS = {
    "facebook": "Facebook/Meta",
    "tiktok": "TikTok",
    "google": "Google",
    "payment": "ชำระบัตร/เงินคืน",
    "other": "อื่น ๆ/ไม่ระบุ",
}

SUMMARY_STATUS_LABELS = {
    "matched": "ตรวจเรียบร้อย",
    "unmatched": "ต้องตรวจ",
    "duplicates": "รายการซ้ำ",
    "missing-attachments": "ไม่มีหลักฐาน",
    "ignored": "ไม่นำมาคำนวณ",
}


def summary_platform_sql(alias: str = "t") -> str:
    """Return the shared, user-facing platform grouping for statement rows."""
    haystack = (
        f"LOWER(COALESCE({alias}.channel, '') || ' ' || "
        f"COALESCE({alias}.description, ''))"
    )
    category = f"LOWER(COALESCE({alias}.category, ''))"
    return f"""
        CASE
            WHEN {category} LIKE '%%ชำระบัตร%%'
              OR {category} LIKE '%%คืนเงิน%%' THEN 'payment'
            WHEN {haystack} LIKE '%%facebook%%'
              OR {haystack} LIKE '%%meta%%' THEN 'facebook'
            WHEN {haystack} LIKE '%%tiktok%%' THEN 'tiktok'
            WHEN {haystack} LIKE '%%google%%' THEN 'google'
            ELSE 'other'
        END
    """


def summary_transaction_filter_parts(
    *,
    date_from: str | None = None,
    date_to: str | None = None,
    card_last4: str | None = None,
    platform: str | None = None,
    status: str | None = None,
    statement_id: int | None = None,
    alias: str = "t",
) -> tuple[list[str], list[Any]]:
    """Build one set of filters shared by summary, review, and exports."""
    clauses: list[str] = []
    values: list[Any] = []

    start = parse_iso_date(date_from)
    end = parse_iso_date(date_to)
    if date_from and not start:
        raise HTTPException(status_code=422, detail="วันที่เริ่มต้นไม่ถูกต้อง")
    if date_to and not end:
        raise HTTPException(status_code=422, detail="วันที่สิ้นสุดไม่ถูกต้อง")
    if start and end and start > end:
        raise HTTPException(status_code=422, detail="วันที่เริ่มต้นต้องไม่อยู่หลังวันที่สิ้นสุด")
    if start:
        clauses.append(f"{alias}.transaction_date >= ?")
        values.append(start.isoformat())
    if end:
        clauses.append(f"{alias}.transaction_date <= ?")
        values.append(end.isoformat())

    if card_last4:
        clean_last4 = re.sub(r"\D", "", card_last4)[-4:]
        if len(clean_last4) != 4:
            raise HTTPException(status_code=422, detail="เลขท้ายบัตรต้องมี 4 หลัก")
        clauses.append(f"{alias}.card_last4 = ?")
        values.append(clean_last4)

    if platform:
        if platform not in SUMMARY_PLATFORM_LABELS:
            raise HTTPException(status_code=422, detail="แพลตฟอร์มที่เลือกไม่ถูกต้อง")
        clauses.append(f"({summary_platform_sql(alias)}) = ?")
        values.append(platform)

    if status:
        if status == "duplicates":
            clauses.append(f"{alias}.is_duplicate = 1")
        elif status == "missing-attachments":
            clauses.append(
                f"{alias}.match_status = 'matched' AND {alias}.has_attachment = 0"
            )
        elif status in {"matched", "unmatched", "ignored"}:
            clauses.append(f"{alias}.match_status = ?")
            values.append(status)
        else:
            raise HTTPException(status_code=422, detail="สถานะที่เลือกไม่ถูกต้อง")

    if statement_id is not None:
        if statement_id <= 0:
            raise HTTPException(status_code=422, detail="ไฟล์ Statement ที่เลือกไม่ถูกต้อง")
        clauses.append(f"{alias}.statement_id = ?")
        values.append(statement_id)
    return clauses, values


def summary_transaction_filters(**kwargs: Any) -> tuple[str, list[Any]]:
    clauses, values = summary_transaction_filter_parts(**kwargs)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    return where, values


@app.get("/transactions", response_class=HTMLResponse)
async def transactions_page(
    request: Request,
    statement_id: int | None = None,
    status: str | None = None,
    card: str | None = None,
    q: str | None = None,
    uploaded: int | None = None,
) -> HTMLResponse:
    where, values = transaction_filters(statement_id, status, card, q)
    with closing(connect()) as db:
        transactions = db.execute(
            f"""
            SELECT
                t.*,
                s.original_filename,
                s.issuer,
                s.statement_type,
                s.masked_reference,
                c.name AS card_name,
                mg.target_reference,
                mg.match_type
            FROM transactions t
            JOIN statements s ON s.id = t.statement_id
            LEFT JOIN cards c ON c.last4 = t.card_last4
            LEFT JOIN match_groups mg ON mg.id = t.match_group_id
            {where}
            ORDER BY t.transaction_date DESC, t.id DESC
            LIMIT 1000
            """,
            values,
        ).fetchall()
        statements = db.execute(
            "SELECT id, original_filename FROM statements ORDER BY uploaded_at DESC"
        ).fetchall()
        cards = db.execute("SELECT * FROM cards ORDER BY name").fetchall()
    return page(
        request,
        "transactions.html",
        active="transactions",
        transactions=transactions,
        statements=statements,
        cards=cards,
        filters={"statement_id": statement_id, "status": status, "card": card, "q": q},
        uploaded=bool(uploaded),
    )


def _rows_to_json(rows: list[Any]) -> str:
    def _default(obj: Any) -> Any:
        if isinstance(obj, Decimal):
            return float(obj)
        if hasattr(obj, "isoformat"):
            return obj.isoformat()
        return str(obj)
    return json.dumps([dict(r) for r in rows], default=_default, ensure_ascii=False)


@app.get("/manual-edit", response_class=HTMLResponse)
async def manual_edit_page(
    request: Request,
    error: str | None = None,
) -> HTMLResponse:
    with closing(connect()) as db:
        statements = db.execute(
            "SELECT id, original_filename FROM statements ORDER BY uploaded_at DESC"
        ).fetchall()
        transactions = db.execute(
            """
            SELECT t.id, t.statement_id, t.transaction_date, t.transaction_time,
                   t.description, t.amount, t.tr_code, t.channel, t.card_last4,
                   t.match_status, t.is_duplicate
            FROM transactions t
            WHERE t.match_status = 'unmatched' AND t.is_duplicate = 0
            ORDER BY t.statement_id, t.transaction_date DESC, t.id DESC
            LIMIT 1000
            """
        ).fetchall()
        reference_items = db.execute(
            """
            SELECT id, reference, amount, transaction_date, transaction_time,
                   party_name, source_filename, has_attachment
            FROM reference_items
            WHERE match_status = 'unmatched'
            ORDER BY transaction_date, id
            """
        ).fetchall()
        warning_stats = db.execute(
            """
            SELECT
                SUM(CASE WHEN match_status = 'unmatched' THEN 1 ELSE 0 END) AS unmatched,
                SUM(CASE WHEN is_duplicate = 1 THEN 1 ELSE 0 END) AS duplicates
            FROM transactions
            """
        ).fetchone()
    return page(
        request,
        "manual_edit.html",
        active="manual-edit",
        statements=statements,
        transactions_json=_rows_to_json(transactions),
        reference_items_json=_rows_to_json(reference_items),
        error=error,
        warning_stats=warning_stats,
    )


@app.post("/transactions/{transaction_id}/update")
async def update_transaction(
    transaction_id: int,
    transaction_date: str = Form(...),
    description: str = Form(...),
    amount: float = Form(...),
    transaction_time: str = Form(""),
    channel: str = Form(""),
    tr_code: str = Form(""),
    card_last4: str = Form(""),
    category: str = Form("ยังไม่จัดหมวดหมู่"),
    match_status: str = Form("unmatched"),
    reference: str = Form(""),
    notes: str = Form(""),
    next_path: str = Form("/manual-edit"),
) -> RedirectResponse:
    if match_status not in {"unmatched", "matched", "ignored"}:
        raise HTTPException(status_code=422, detail="Invalid match status")
    safe_next = next_path if next_path in {"/review", "/manual-edit", "/transactions"} else "/manual-edit"
    last4 = re.sub(r"\D", "", card_last4)[-4:] or None
    with closing(connect()) as db:
        before = db.execute(
            "SELECT statement_id FROM transactions WHERE id = ?", (transaction_id,)
        ).fetchone()
        result = db.execute(
            """
            UPDATE transactions
            SET transaction_date = ?, description = ?, amount = ?,
                transaction_time = ?, channel = ?, tr_code = ?,
                card_last4 = ?, category = ?, match_status = ?,
                reference = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                transaction_date,
                description.strip()[:500],
                amount,
                transaction_time.strip()[:8] or None,
                channel.strip()[:100] or None,
                tr_code.strip()[:50] or None,
                last4,
                category.strip()[:100] or "ยังไม่จัดหมวดหมู่",
                match_status,
                reference.strip()[:200] or None,
                notes.strip()[:1000] or None,
                transaction_id,
            ),
        )
        if before:
            audit(db, "update_transaction", "transaction", transaction_id, "manual field update")
            sync_statement_stats(db, int(before["statement_id"]))
        db.commit()
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Transaction not found")
    return redirect(safe_next)


@app.post("/matches/manual")
async def create_manual_match(
    transaction_ids: str = Form(...),
    target_reference: str = Form(...),
    expected_amount: float = Form(...),
    has_attachment: int = Form(0),
    notes: str = Form(""),
) -> RedirectResponse:
    ids = sorted({int(value) for value in re.findall(r"\d+", transaction_ids)})
    if not ids:
        return redirect_manual_error("กรุณาเลือกรายการ Statement")
    reference = target_reference.strip()[:200]
    if not reference:
        return redirect_manual_error("กรุณาระบุเลขอ้างอิง/รายการที่นำมาชน")

    placeholders = ",".join("?" for _ in ids)
    with closing(connect()) as db:
        rows = db.execute(
            f"""
            SELECT id, statement_id, amount, match_status
            FROM transactions
            WHERE id IN ({placeholders})
            """,
            ids,
        ).fetchall()
        if len(rows) != len(ids):
            return redirect_manual_error("ไม่พบรายการ Statement บางรายการ")
        if any(row["match_status"] == "matched" for row in rows):
            return redirect_manual_error("มีรายการที่ถูกจับคู่แล้ว กรุณาเลือกรายการใหม่")

        try:
            create_match_group(
                db,
                rows,
                reference=reference,
                expected_cents=money_cents(expected_amount),
                has_attachment=1 if has_attachment else 0,
                method="manual_group" if len(ids) > 1 else "manual_single",
                notes=notes,
            )
        except ValueError as exc:
            return redirect_manual_error(
                str(exc)
            )
        db.commit()
    return redirect("/manual-edit?status=unmatched")


@app.post("/references/{ref_id}/update")
async def update_reference_item(
    ref_id: int,
    reference: str = Form(...),
    amount: float = Form(...),
    transaction_date: str = Form(""),
    transaction_time: str = Form(""),
    party_name: str = Form(""),
    notes: str = Form(""),
    next_path: str = Form("/references"),
) -> RedirectResponse:
    safe_next = next_path if next_path.startswith("/") else "/references"
    with closing(connect()) as db:
        result = db.execute(
            """
            UPDATE reference_items
            SET reference = ?, amount = ?, transaction_date = ?,
                transaction_time = ?, party_name = ?, notes = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (
                reference.strip()[:200],
                amount,
                transaction_date.strip() or None,
                transaction_time.strip()[:8] or None,
                party_name.strip()[:300] or None,
                notes.strip()[:1000] or None,
                ref_id,
            ),
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Reference item not found")
        audit(db, "update_reference_item", "reference_items", ref_id, "manual field update")
        db.commit()
    return redirect(safe_next)


@app.get("/references", response_class=HTMLResponse)
async def references_page(
    request: Request,
    error: str | None = None,
    imported: int | None = None,
    deleted: int | None = None,
) -> HTMLResponse:
    with closing(connect()) as db:
        items = db.execute(
            """
            SELECT *
            FROM reference_items
            ORDER BY match_status = 'matched', transaction_date DESC, id DESC
            LIMIT 500
            """
        ).fetchall()
        stats = db.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN match_status = 'matched' THEN 1 ELSE 0 END) AS matched,
                SUM(CASE WHEN match_status = 'unmatched' THEN 1 ELSE 0 END) AS unmatched,
                SUM(CASE WHEN has_attachment = 0 THEN 1 ELSE 0 END) AS missing_attachments
            FROM reference_items
            """
        ).fetchone()
        sources = db.execute(
            """
            SELECT
                source_filename,
                COUNT(*) AS total,
                SUM(CASE WHEN match_status = 'matched' THEN 1 ELSE 0 END) AS matched,
                SUM(CASE WHEN match_status = 'unmatched' THEN 1 ELSE 0 END) AS unmatched,
                MAX(created_at) AS imported_at
            FROM reference_items
            WHERE source_filename IS NOT NULL AND source_filename <> ''
            GROUP BY source_filename
            ORDER BY imported_at DESC, source_filename ASC
            """
        ).fetchall()
    return page(
        request,
        "references.html",
        active="references",
        items=items,
        stats=stats,
        sources=sources,
        error=error,
        imported=bool(imported),
        deleted=bool(deleted),
    )


@app.post("/references/upload")
async def upload_references(request: Request, file: UploadFile = File(...)):
    original_name = Path(file.filename or "references").name
    try:
        data = await read_upload(file)
        items = parse_reference_upload(original_name, data)
    except ValueError as exc:
        return redirect(f"/references?error={quote(str(exc))}")

    with closing(connect()) as db:
        existing_hashes = {
            row["row_hash"]
            for row in db.execute(
                "SELECT row_hash FROM reference_items WHERE row_hash IS NOT NULL"
            ).fetchall()
        }
        inserted = 0
        for item in items:
            if item["row_hash"] in existing_hashes:
                continue
            db.execute(
                """
                INSERT INTO reference_items (
                    source_filename, reference, transaction_date, transaction_time,
                    amount, party_name, has_attachment, notes, row_hash
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    item["source_filename"],
                    item["reference"],
                    item["transaction_date"],
                    item["transaction_time"],
                    item["amount"],
                    item["party_name"],
                    item["has_attachment"],
                    item["notes"],
                    item["row_hash"],
                ),
            )
            inserted += 1
            existing_hashes.add(item["row_hash"])
        audit(db, "upload_references", "reference_items", original_name, f"inserted {inserted} rows")
        db.commit()
    return redirect("/references?imported=1")


@app.post("/references/source/delete")
async def delete_reference_source(source_filename: str = Form(...)) -> RedirectResponse:
    filename = Path(source_filename).name.strip()
    if not filename:
        return redirect(f"/references?error={quote('ไม่พบชื่อไฟล์ที่ต้องการลบ')}")

    with closing(connect()) as db:
        stats = db.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN match_status = 'matched' THEN 1 ELSE 0 END) AS matched
            FROM reference_items
            WHERE source_filename = ?
            """,
            (filename,),
        ).fetchone()
        if not stats or not stats["total"]:
            return redirect(f"/references?error={quote('ไม่พบข้อมูลจากไฟล์นี้')}")
        if stats["matched"]:
            return redirect(
                f"/references?error={quote('ลบไฟล์นี้ไม่ได้ เพราะมีรายการที่ถูกใช้จับคู่แล้ว กรุณาตรวจ match ก่อน')}"
            )
        db.execute("DELETE FROM reference_items WHERE source_filename = ?", (filename,))
        audit(db, "delete_reference_source", "reference_items", filename, f"deleted {stats['total']} rows")
        db.commit()
    return redirect("/references?deleted=1")


@app.post("/references/manual")
async def create_reference_item(
    reference: str = Form(...),
    amount: float = Form(...),
    transaction_date: str = Form(""),
    transaction_time: str = Form(""),
    party_name: str = Form(""),
    has_attachment: int = Form(0),
    notes: str = Form(""),
) -> RedirectResponse:
    clean_reference = reference.strip()[:200]
    if not clean_reference:
        return redirect(f"/references?error={quote('กรุณาระบุ reference')}")
    parsed_date = parse_date(transaction_date) if transaction_date else None
    parsed_time = parse_time(transaction_time) if transaction_time else None
    row_hash = hashlib.sha256(
        "|".join(
            [
                clean_reference,
                parsed_date or "",
                parsed_time or "",
                f"{round(float(amount), 2):.2f}",
                party_name.strip()[:200],
            ]
        ).encode("utf-8")
    ).hexdigest()
    with closing(connect()) as db:
        db.execute(
            """
            INSERT INTO reference_items (
                source_filename, reference, transaction_date, transaction_time,
                amount, party_name, has_attachment, notes, row_hash
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "manual",
                clean_reference,
                parsed_date,
                parsed_time,
                round(float(amount), 2),
                party_name.strip()[:200] or None,
                1 if has_attachment else 0,
                notes.strip()[:1000] or None,
                row_hash,
            ),
        )
        audit(db, "create_reference", "reference_item", clean_reference, "manual create")
        db.commit()
    return redirect("/references")


@app.post("/matches/reference")
async def match_with_reference(
    transaction_ids: str = Form(...),
    reference_item_id: int = Form(...),
    notes: str = Form(""),
    next_url: str = Form(""),
) -> RedirectResponse:
    ids = sorted({int(value) for value in re.findall(r"\d+", transaction_ids)})
    if not ids:
        return redirect_manual_error("กรุณาเลือกรายการ Statement")
    placeholders = ",".join("?" for _ in ids)
    with closing(connect()) as db:
        tx_rows = db.execute(
            f"""
            SELECT id, statement_id, amount, match_status
            FROM transactions
            WHERE id IN ({placeholders})
            """,
            ids,
        ).fetchall()
        reference = db.execute(
            "SELECT * FROM reference_items WHERE id = ?",
            (reference_item_id,),
        ).fetchone()
        if len(tx_rows) != len(ids) or not reference:
            return redirect_manual_error("ข้อมูลสำหรับจับคู่ไม่ครบ")
        if any(row["match_status"] == "matched" for row in tx_rows):
            return redirect_manual_error("มีรายการ Statement ที่ถูกจับคู่แล้ว")
        if reference["match_status"] == "matched":
            return redirect_manual_error("รายการฝั่งเปรียบเทียบนี้ถูกจับคู่แล้ว")
        try:
            create_match_group(
                db,
                tx_rows,
                reference=reference["reference"],
                expected_cents=money_cents(reference["amount"]),
                has_attachment=int(reference["has_attachment"] or 0),
                method="reference_group" if len(ids) > 1 else "reference_single",
                notes=notes,
                reference_item_id=reference_item_id,
            )
        except ValueError as exc:
            return redirect_manual_error(str(exc))
        db.commit()
    # redirect กลับหน้าเดิม (statement filter เดิม) ถ้า next_url ปลอดภัย
    safe_next = next_url if next_url.startswith("/") and not next_url.startswith("//") else "/review"
    return redirect(safe_next)


@app.post("/matches/auto")
async def auto_match(statement_id: int | None = Form(None)) -> RedirectResponse:
    with closing(connect()) as db:
        clauses = ["t.match_status = 'unmatched'", "t.is_duplicate = 0"]
        values: list[Any] = []
        if statement_id:
            clauses.append("t.statement_id = ?")
            values.append(statement_id)
        rows = db.execute(
            f"""
            SELECT t.*
            FROM transactions t
            WHERE {" AND ".join(clauses)}
            ORDER BY t.transaction_date, t.transaction_time, t.id
            """,
            values,
        ).fetchall()
        matched = 0
        for row in rows:
            candidates = candidate_reference_items(db, row, limit=2)
            if not candidates:
                continue
            candidate = candidates[0]
            next_score = candidates[1]["match_score"] if len(candidates) > 1 else 0
            if candidate["match_score"] < 82 or candidate["match_score"] - next_score < 12:
                continue
            # auto-match ต้องการวันที่ตรงเสมอ (text-only match ข้ามไป)
            if parse_iso_date(candidate.get("transaction_date")) != parse_iso_date(row["transaction_date"]):
                continue
            create_match_group(
                db,
                [row],
                reference=candidate["reference"],
                expected_cents=money_cents(candidate["amount"]),
                has_attachment=int(candidate["has_attachment"] or 0),
                method="auto_exact_amount_date_name",
                notes=f"auto matched by exact amount/date/name; score={candidate['match_score']}; {candidate['match_reason']}",
                reference_item_id=int(candidate["id"]),
            )
            matched += 1
        audit(db, "auto_match", "transactions", statement_id or "all", f"matched {matched} rows")
        db.commit()
    return redirect("/review")


@app.post("/matches/{group_id}/remove")
async def remove_match_group(group_id: int, request: Request) -> RedirectResponse:
    """Unmatch — reset all transactions in the group back to unmatched."""
    with closing(connect()) as db:
        group = db.execute(
            "SELECT * FROM match_groups WHERE id = ?", (group_id,)
        ).fetchone()
        if not group:
            return redirect("/review")

        tx_rows = db.execute(
            """
            SELECT t.id, t.statement_id
            FROM match_group_items mgi
            JOIN transactions t ON t.id = mgi.transaction_id
            WHERE mgi.match_group_id = ?
            """,
            (group_id,),
        ).fetchall()
        tx_ids = [int(r["id"]) for r in tx_rows]
        stmt_ids = {int(r["statement_id"]) for r in tx_rows}

        if tx_ids:
            placeholders = ",".join("?" for _ in tx_ids)
            db.execute(
                f"""
                UPDATE transactions
                SET match_status = 'unmatched',
                    match_group_id = NULL,
                    match_method = NULL,
                    reference = NULL,
                    has_attachment = 0,
                    notes = NULL,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id IN ({placeholders})
                """,
                tx_ids,
            )

        if group["reference_item_id"]:
            db.execute(
                """
                UPDATE reference_items
                SET match_status = 'unmatched', updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (group["reference_item_id"],),
            )

        db.execute("DELETE FROM match_group_items WHERE match_group_id = ?", (group_id,))
        db.execute("DELETE FROM match_groups WHERE id = ?", (group_id,))

        for sid in stmt_ids:
            sync_statement_stats(db, sid)

        audit(
            db, "remove_match", "match_group", group_id,
            f"tx={','.join(map(str, tx_ids))}; ref={group['target_reference']}",
        )
        db.commit()
    return redirect("/review")


@app.get("/audit", response_class=HTMLResponse)
async def audit_page(request: Request) -> HTMLResponse:
    with closing(connect()) as db:
        logs = db.execute(
            """
            SELECT *
            FROM statement_audit_logs
            ORDER BY created_at DESC, id DESC
            LIMIT 300
            """
        ).fetchall()
    return page(request, "audit.html", active="audit", logs=logs)


@app.get("/export/{kind}.csv")
async def export_csv(
    kind: str,
    date_from: str | None = None,
    date_to: str | None = None,
    card_last4: str | None = None,
    platform: str | None = None,
    status: str | None = None,
    statement_id: int | None = None,
) -> Response:
    if kind not in {"matched", "unmatched", "missing-attachments"}:
        raise HTTPException(status_code=404, detail="Export not found")
    export_clause = {
        "matched": "t.match_status = 'matched'",
        "unmatched": "t.match_status = 'unmatched'",
        "missing-attachments": "t.match_status = 'matched' AND t.has_attachment = 0",
    }[kind]
    clauses, values = summary_transaction_filter_parts(
        date_from=date_from,
        date_to=date_to,
        card_last4=card_last4,
        platform=platform,
        status=status,
        statement_id=statement_id,
    )
    clauses.append(export_clause)
    where = f"WHERE {' AND '.join(clauses)}"
    with closing(connect()) as db:
        rows = db.execute(
            f"""
            SELECT
                t.transaction_date,
                t.transaction_time,
                t.description,
                t.amount,
                t.match_status,
                t.reference,
                t.has_attachment,
                t.is_duplicate,
                s.original_filename,
                mg.target_reference,
                mg.match_type,
                t.match_method,
                ri.reference AS reference_item,
                ri.party_name
            FROM transactions t
            JOIN statements s ON s.id = t.statement_id
            LEFT JOIN match_groups mg ON mg.id = t.match_group_id
            LEFT JOIN reference_items ri ON ri.id = mg.reference_item_id
            {where}
            ORDER BY t.transaction_date, t.transaction_time, t.id
            """,
            values,
        ).fetchall()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "transaction_date",
            "transaction_time",
            "description",
            "amount",
            "match_status",
            "reference",
            "target_reference",
            "reference_item",
            "party_name",
            "has_attachment",
            "is_duplicate",
            "statement_file",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row["transaction_date"],
                row["transaction_time"] or "",
                row["description"],
                f"{float(row['amount'] or 0):.2f}",
                row["match_status"],
                row["reference"] or "",
                row["target_reference"] or "",
                row["reference_item"] or "",
                row["party_name"] or "",
                int(row["has_attachment"] or 0),
                int(row["is_duplicate"] or 0),
                row["original_filename"],
            ]
        )
    return Response(
        output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{kind}.csv"'},
    )


_EXCEL_COLS: list[tuple[str, str]] = [
    ("transaction_date", "วันที่"),
    ("transaction_time", "เวลา"),
    ("description", "รายละเอียด"),
    ("amount", "จำนวนเงิน"),
    ("card_last4", "บัตร (4 หลัก)"),
    ("match_status", "สถานะ"),
    ("reference", "Reference ที่ตั้งไว้"),
    ("target_reference", "Reference ที่จับคู่"),
    ("party_name", "ชื่อฝั่งเปรียบเทียบ"),
    ("has_attachment", "มีเอกสาร"),
    ("is_duplicate", "ซ้ำ"),
    ("original_filename", "ไฟล์ Statement"),
]


def _xlsx_header(ws: Any, fill_hex: str) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 52)


def _xlsx_data_sheet(wb: Any, title: str, rows: list[dict], fill_hex: str) -> None:
    ws = wb.create_sheet(title)
    ws.append([label for _, label in _EXCEL_COLS])
    for row in rows:
        values: list[Any] = []
        for key, _ in _EXCEL_COLS:
            val = row.get(key)
            if val is None:
                values.append("")
            elif key in ("has_attachment", "is_duplicate"):
                values.append("ใช่" if int(val or 0) else "ไม่")
            elif key == "amount":
                values.append(float(val or 0))
            else:
                values.append(str(val))
        ws.append(values)
    _xlsx_header(ws, fill_hex)


@app.get("/export/report.xlsx")
async def export_excel_report(
    date_from: str | None = None,
    date_to: str | None = None,
    card_last4: str | None = None,
    platform: str | None = None,
    status: str | None = None,
    statement_id: int | None = None,
) -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    base_clauses, base_values = summary_transaction_filter_parts(
        date_from=date_from,
        date_to=date_to,
        card_last4=card_last4,
        platform=platform,
        status=status,
        statement_id=statement_id,
    )

    def fetch(where_clause: str) -> list[dict]:
        clauses = [*base_clauses, where_clause]
        return db.execute(
            f"""
            SELECT
                t.transaction_date, t.transaction_time, t.description,
                t.amount, t.card_last4, t.match_status, t.reference,
                t.has_attachment, t.is_duplicate,
                s.original_filename,
                mg.target_reference,
                ri.party_name
            FROM transactions t
            JOIN statements s ON s.id = t.statement_id
            LEFT JOIN match_groups mg ON mg.id = t.match_group_id
            LEFT JOIN reference_items ri ON ri.id = mg.reference_item_id
            WHERE {' AND '.join(clauses)}
            ORDER BY t.transaction_date DESC, t.transaction_time DESC, t.id DESC
            """,
            base_values,
        ).fetchall()

    with closing(connect()) as db:
        matched = fetch("t.match_status = 'matched' AND t.is_duplicate = 0")
        unmatched = fetch("t.match_status = 'unmatched' AND t.is_duplicate = 0")
        duplicates = fetch("t.is_duplicate = 1")
        totals_where = f"WHERE {' AND '.join(base_clauses)}" if base_clauses else ""
        totals = db.execute(
            f"""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN match_status='matched' AND is_duplicate=0 THEN 1 ELSE 0 END) AS matched,
                SUM(CASE WHEN match_status='unmatched' AND is_duplicate=0 THEN 1 ELSE 0 END) AS unmatched,
                SUM(CASE WHEN is_duplicate=1 THEN 1 ELSE 0 END) AS duplicates
            FROM transactions t
            {totals_where}
            """,
            base_values,
        ).fetchone()
        statement_name = ""
        if statement_id:
            statement_row = db.execute(
                "SELECT original_filename FROM statements WHERE id = ?", (statement_id,)
            ).fetchone()
            statement_name = str(statement_row["original_filename"] or "") if statement_row else ""
        card_name = ""
        if card_last4:
            clean_last4 = re.sub(r"\D", "", card_last4)[-4:]
            card_row = db.execute("SELECT name FROM cards WHERE last4 = ?", (clean_last4,)).fetchone()
            card_name = str(card_row["name"] or "") if card_row else ""

    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws_s = wb.create_sheet("สรุป")
    ws_s.append(["รายงาน Reconciliation", ""])
    ws_s.append(["", ""])
    ws_s.append(["รายการทั้งหมด", int(totals["total"] or 0)])
    ws_s.append(["✅ Matched", int(totals["matched"] or 0)])
    ws_s.append(["❌ Unmatched", int(totals["unmatched"] or 0)])
    ws_s.append(["⚠ Duplicate", int(totals["duplicates"] or 0)])
    selected_filters: list[tuple[str, str]] = []
    if date_from or date_to:
        selected_filters.append(("ช่วงวันที่", f"{date_from or 'เริ่มต้น'} ถึง {date_to or 'ปัจจุบัน'}"))
    if card_last4:
        clean_last4 = re.sub(r"\D", "", card_last4)[-4:]
        selected_filters.append(("บัตร", f"{card_name + ' ' if card_name else ''}••••{clean_last4}"))
    if platform:
        selected_filters.append(("แพลตฟอร์ม", SUMMARY_PLATFORM_LABELS.get(platform, platform)))
    if status:
        selected_filters.append(("สถานะ", SUMMARY_STATUS_LABELS.get(status, status)))
    if statement_id:
        selected_filters.append(("ไฟล์ Statement", statement_name or f"ID {statement_id}"))
    ws_s.append(["ตัวกรอง", "ข้อมูลทั้งหมด" if not selected_filters else "ตามรายการด้านล่าง"])
    for label, selected_value in selected_filters:
        ws_s.append([label, selected_value])
    ws_s["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws_s["A1"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws_s.column_dimensions["A"].width = 28
    ws_s.column_dimensions["B"].width = 16

    _xlsx_data_sheet(wb, "✅ Matched", matched, "217346")
    _xlsx_data_sheet(wb, "❌ Unmatched", unmatched, "C0392B")
    _xlsx_data_sheet(wb, "⚠ Duplicate", duplicates, "D35400")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="reconciliation.xlsx"'},
    )


@app.get("/summary", response_class=HTMLResponse)
async def summary_page(request: Request) -> HTMLResponse:
    with closing(connect()) as db:
        totals = db.execute(
            """
            SELECT
                COUNT(*) AS transaction_count,
                COALESCE(SUM(CASE WHEN amount > 0 THEN amount ELSE 0 END), 0) AS charges,
                COALESCE(SUM(CASE WHEN amount < 0 THEN ABS(amount) ELSE 0 END), 0) AS refunds,
                COALESCE(SUM(amount), 0) AS net,
                SUM(CASE WHEN match_status = 'matched' THEN 1 ELSE 0 END) AS matched,
                SUM(CASE WHEN match_status = 'unmatched' THEN 1 ELSE 0 END) AS unmatched,
                SUM(CASE WHEN match_status = 'ignored' THEN 1 ELSE 0 END) AS ignored,
                SUM(CASE WHEN is_duplicate = 1 THEN 1 ELSE 0 END) AS duplicates,
                SUM(CASE WHEN match_status = 'matched' AND has_attachment = 0 THEN 1 ELSE 0 END) AS missing_attachments
            FROM transactions
            """
        ).fetchone()
        match_groups = db.execute(
            """
            SELECT
                COUNT(*) AS count,
                SUM(CASE WHEN match_type = 'group' THEN 1 ELSE 0 END) AS group_count,
                SUM(CASE WHEN has_attachment = 0 THEN 1 ELSE 0 END) AS no_attachment_count
            FROM match_groups
            WHERE status = 'confirmed'
            """
        ).fetchone()
        categories = db.execute(
            """
            SELECT category, COUNT(*) AS count, SUM(amount) AS total
            FROM transactions
            GROUP BY category
            ORDER BY ABS(SUM(amount)) DESC
            """
        ).fetchall()
        months = db.execute(
            """
            SELECT substr(transaction_date, 1, 7) AS month,
                   COUNT(*) AS count,
                   SUM(amount) AS total
            FROM transactions
            GROUP BY substr(transaction_date, 1, 7)
            ORDER BY month DESC
            LIMIT 12
            """
        ).fetchall()
    return page(
        request,
        "summary.html",
        active="summary",
        totals=totals,
        match_groups=match_groups,
        categories=categories,
        months=months,
    )


@app.get("/cards", response_class=HTMLResponse)
async def cards_page(request: Request) -> HTMLResponse:
    with closing(connect()) as db:
        cards = db.execute(
            """
            SELECT c.*,
                   COUNT(t.id) AS transaction_count,
                   COALESCE(SUM(t.amount), 0) AS total
            FROM cards c
            LEFT JOIN transactions t ON t.card_last4 = c.last4
            GROUP BY c.id
            ORDER BY c.name
            """
        ).fetchall()
        unknown_cards = db.execute(
            """
            SELECT card_last4, COUNT(*) AS transaction_count
            FROM transactions
            WHERE card_last4 IS NOT NULL
              AND card_last4 NOT IN (SELECT last4 FROM cards)
            GROUP BY card_last4
            ORDER BY card_last4
            """
        ).fetchall()
    return page(
        request,
        "cards.html",
        active="cards",
        cards=cards,
        unknown_cards=unknown_cards,
    )


@app.post("/cards")
async def create_card(
    name: str = Form(...),
    last4: str = Form(...),
    holder_name: str = Form(""),
    bank_name: str = Form(""),
) -> RedirectResponse:
    clean_last4 = re.sub(r"\D", "", last4)[-4:]
    if len(clean_last4) != 4:
        raise HTTPException(status_code=422, detail="เลขท้ายบัตรต้องมี 4 หลัก")
    with closing(connect()) as db:
        db.execute(
            """
            INSERT INTO cards (name, last4, holder_name, bank_name)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(last4) DO UPDATE SET
                name = excluded.name,
                holder_name = excluded.holder_name,
                bank_name = excluded.bank_name
            """,
            (
                name.strip()[:100],
                clean_last4,
                holder_name.strip()[:100] or None,
                bank_name.strip()[:100] or None,
            ),
        )
        db.commit()
    return redirect("/cards")


@app.post("/cards/{card_id}/delete")
async def delete_card(card_id: int) -> RedirectResponse:
    with closing(connect()) as db:
        db.execute("DELETE FROM cards WHERE id = ?", (card_id,))
        db.commit()
    return redirect("/cards")


# Imported here (not at module top, and not interleaved earlier) so
# `app.api` can safely do `from app.main import ...` for every helper
# defined throughout this file without a circular-import error — by the
# end of module execution all of them exist.
from app.api import api_router  # noqa: E402

app.include_router(api_router)
