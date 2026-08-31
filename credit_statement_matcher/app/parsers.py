from __future__ import annotations

import csv
import io
import os
import re
import shutil
import subprocess
import tempfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable


@dataclass
class ParsedTransaction:
    transaction_date: str | None
    description: str
    amount: float | None
    card_last4: str | None
    category: str
    transaction_time: str | None = None
    deposit_amount: float | None = None
    withdraw_amount: float | None = None
    channel: str | None = None
    tr_code: str | None = None
    row_hash: str | None = None
    source_page: int | None = None
    confidence: float = 100.0
    warnings: list[str] = field(default_factory=list)


@dataclass
class ParsedStatement:
    issuer: str
    statement_type: str
    extraction_method: str
    masked_reference: str | None
    statement_date_from: str | None
    statement_date_to: str | None
    summary_totals: dict[str, float]
    transactions: list[ParsedTransaction]
    warnings: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class PositionedWord:
    text: str
    left: float
    top: float
    width: float
    height: float
    confidence: float


@dataclass
class OcrLine:
    text: str
    confidence: float
    page: int
    words: list[PositionedWord] = field(default_factory=list)


@dataclass(frozen=True)
class TrustedAmount:
    page: int
    top: float
    amount: float
    is_credit: bool
    merchant: str | None = None
    reference: str | None = None
    location: str | None = None


HEADER_ALIASES = {
    "date": {
        "date",
        "transactiondate",
        "postingdate",
        "วันที่",
        "วันที่ทำรายการ",
        "วันที่รายการ",
        "วันเดือนปี",
    },
    "description": {
        "description",
        "details",
        "detail",
        "merchant",
        "รายการ",
        "รายละเอียด",
        "ชื่อร้านค้า",
    },
    "amount": {
        "amount",
        "transactionamount",
        "ยอด",
        "ยอดเงิน",
        "จำนวนเงิน",
        "ยอดใช้จ่าย",
    },
    "debit": {"debit", "charge", "ถอน", "เดบิต", "ยอดเรียกเก็บ"},
    "credit": {"credit", "payment", "refund", "ฝาก", "เครดิต", "ยอดชำระ"},
    "deposit": {"deposit", "เงินเข้า", "ยอดเงินเข้า", "ยอดรับ", "รับเงิน", "ยอดฝาก"},
    "withdraw": {"withdraw", "withdrawal", "เงินออก", "ยอดเงินออก", "ยอดถอน", "ถอนเงิน"},
    "time": {"time", "transactiontime", "เวลา", "เวลาทำรายการ", "เวลารายการ"},
    "channel": {"channel", "ช่องทาง", "ช่องทางรายการ"},
    "tr_code": {"trcode", "transactioncode", "code", "รหัสรายการ"},
    "card": {
        "card",
        "cardnumber",
        "cardno",
        "เลขบัตร",
        "หมายเลขบัตร",
        "บัตร",
    },
}


CATEGORY_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("อาหารและเครื่องดื่ม", ("restaurant", "cafe", "coffee", "food", "grabfood", "lineman", "ร้านอาหาร", "กาแฟ")),
    ("เดินทาง", ("grab", "bolt", "taxi", "airasia", "thai air", "flight", "bts", "mrt", "ทางด่วน", "น้ำมัน")),
    ("ซอฟต์แวร์และออนไลน์", ("google", "apple.com", "openai", "microsoft", "adobe", "aws", "cloud", "hosting", "subscription")),
    ("สำนักงาน", ("office", "stationery", "ออฟฟิศ", "เครื่องเขียน", "อุปกรณ์สำนักงาน")),
    ("สาธารณูปโภค", ("electric", "water", "internet", "telephone", "ไฟฟ้า", "ประปา", "อินเทอร์เน็ต", "โทรศัพท์")),
    ("ค่าธรรมเนียม", ("fee", "interest", "charge", "ค่าธรรมเนียม", "ดอกเบี้ย")),
    ("ชำระบัตร/คืนเงิน", ("payment received", "refund", "cashback", "ชำระ", "คืนเงิน")),
]


def _normalise_header(value: Any) -> str:
    return re.sub(r"[\s_\-./()]+", "", str(value or "").strip().lower())


def _column_map(row: Iterable[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    normalised = [_normalise_header(value) for value in row]
    for field, aliases in HEADER_ALIASES.items():
        for index, value in enumerate(normalised):
            if value in aliases:
                result[field] = index
                break
    return result


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[str, int]] | None:
    for index, row in enumerate(rows[:25]):
        mapping = _column_map(row)
        if "date" in mapping and "description" in mapping and (
            "amount" in mapping
            or "debit" in mapping
            or "credit" in mapping
            or "deposit" in mapping
            or "withdraw" in mapping
        ):
            return index, mapping
    return None


def parse_amount(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    negative = text.startswith("(") and text.endswith(")")
    if re.search(r"\b(?:CR|CREDIT)\b", text, re.IGNORECASE):
        negative = True

    cleaned = re.sub(r"[^\d.,+\-]", "", text)
    if not cleaned:
        return None

    if cleaned.count(",") and not cleaned.count("."):
        tail = cleaned.rsplit(",", 1)[-1]
        cleaned = cleaned.replace(",", ".") if len(tail) == 2 else cleaned.replace(",", "")
    else:
        cleaned = cleaned.replace(",", "")

    try:
        amount = float(cleaned)
    except ValueError:
        return None
    return -abs(amount) if negative else amount


def parse_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 20_000 < float(value) < 80_000:
        # Excel serial date, including Excel's 1900 leap-year compatibility.
        origin = datetime(1899, 12, 30)
        return (origin + timedelta(days=float(value))).date().isoformat()

    text = str(value or "").strip()
    text = re.sub(r"\s+.*$", "", text)
    if not text:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%m/%d/%Y",
    ):
        try:
            parsed = datetime.strptime(text, fmt).date()
            if parsed.year > 2400:
                parsed = parsed.replace(year=parsed.year - 543)
            return parsed.isoformat()
        except ValueError:
            continue
    return None


def parse_time(value: Any) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")
    if isinstance(value, (int, float)):
        number = float(value)
        if 0 <= number < 1:
            total_seconds = int(round(number * 24 * 60 * 60))
            hours = (total_seconds // 3600) % 24
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).strftime("%H:%M:%S")
        except ValueError:
            continue
    match = re.match(r"^(\d{1,2})[:.](\d{2})(?:[:.](\d{2}))?$", text)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        seconds = int(match.group(3) or 0)
        if 0 <= hours <= 23 and 0 <= minutes <= 59 and 0 <= seconds <= 59:
            return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    return None


def _last4(value: Any) -> str | None:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits[-4:] if len(digits) >= 4 else None


def guess_category(description: str, amount: float) -> str:
    text = description.lower()
    for category, keywords in CATEGORY_RULES:
        if any(keyword in text for keyword in keywords):
            return category
    if amount < 0:
        return "ชำระบัตร/คืนเงิน"
    return "ยังไม่จัดหมวดหมู่"


def _clean_optional(value: Any, limit: int = 100) -> str | None:
    text = str(value or "").strip()
    return text[:limit] if text else None


def make_row_hash(
    transaction_date: str,
    transaction_time: str | None,
    amount: float,
    description: str,
    channel: str | None,
    tr_code: str | None,
) -> str:
    import hashlib

    basis = "|".join(
        [
            transaction_date,
            transaction_time or "",
            f"{amount:.2f}",
            re.sub(r"\s+", " ", description.strip()).lower(),
            channel or "",
            tr_code or "",
        ]
    )
    return hashlib.sha256(basis.encode("utf-8")).hexdigest()


def parse_rows(rows: list[list[Any]]) -> list[ParsedTransaction]:
    header = _find_header(rows)
    if not header:
        raise ValueError(
            "ไม่พบหัวตาราง กรุณาใช้คอลัมน์ Date/วันที่, Description/รายละเอียด และ Amount/จำนวนเงิน"
        )

    header_index, columns = header
    transactions: list[ParsedTransaction] = []

    for row in rows[header_index + 1 :]:
        if not row or all(value in (None, "") for value in row):
            continue

        def cell(field: str) -> Any:
            index = columns.get(field)
            return row[index] if index is not None and index < len(row) else None

        transaction_date = parse_date(cell("date"))
        description = str(cell("description") or "").strip()

        deposit = parse_amount(cell("deposit"))
        withdraw = parse_amount(cell("withdraw"))

        if "amount" in columns:
            amount = parse_amount(cell("amount"))
        else:
            debit = parse_amount(cell("debit"))
            credit = parse_amount(cell("credit"))
            if deposit not in (None, 0):
                amount = abs(deposit)
            elif withdraw not in (None, 0):
                amount = -abs(withdraw)
            else:
                amount = abs(debit) if debit not in (None, 0) else (-abs(credit) if credit not in (None, 0) else None)

        if not transaction_date or not description or amount is None:
            continue

        transaction_time = parse_time(cell("time"))
        channel = _clean_optional(cell("channel"))
        tr_code = _clean_optional(cell("tr_code"), 50)
        signed_amount = round(amount, 2)

        transactions.append(
            ParsedTransaction(
                transaction_date=transaction_date,
                description=description[:500],
                amount=signed_amount,
                card_last4=_last4(cell("card")),
                category=guess_category(description, amount),
                transaction_time=transaction_time,
                deposit_amount=abs(deposit) if deposit not in (None, 0) else (signed_amount if signed_amount > 0 else None),
                withdraw_amount=abs(withdraw) if withdraw not in (None, 0) else (abs(signed_amount) if signed_amount < 0 else None),
                channel=channel,
                tr_code=tr_code,
                row_hash=make_row_hash(transaction_date, transaction_time, signed_amount, description, channel, tr_code),
            )
        )

    if not transactions:
        raise ValueError("อ่านไฟล์ได้แต่ไม่พบรายการที่มีวันที่ รายละเอียด และจำนวนเงินครบ")
    return transactions


def _decode_csv(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp874", "tis-620"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("ไม่สามารถอ่าน encoding ของไฟล์ CSV ได้")


def parse_csv(data: bytes) -> list[ParsedTransaction]:
    text = _decode_csv(data)
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    rows = [list(row) for row in csv.reader(io.StringIO(text), dialect)]
    return parse_rows(rows)


def parse_xlsx(data: bytes) -> list[ParsedTransaction]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    errors: list[str] = []
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        try:
            return parse_rows(rows)
        except ValueError as exc:
            errors.append(f"{sheet.title}: {exc}")
    raise ValueError("ไม่พบตารางรายการในไฟล์ Excel — " + "; ".join(errors))


PDF_LINE_PATTERN = re.compile(
    r"^\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}-\d{1,2}-\d{1,2})\s+"
    r"(.+?)\s+"
    r"(\(?[-+]?(?:\d{1,3}(?:,\d{3})*|\d+)(?:\.\d{2})?\)?(?:\s*CR)?)\s*$",
    re.IGNORECASE,
)

AMOUNT_AT_END_PATTERN = re.compile(
    r"(\(?[-+]?(?:\d{1,3}(?:,\d{3})*|\d+)(?:\.\d{2})\)?(?:\s*(?:CR|DR))?)\s*$",
    re.IGNORECASE,
)
MAX_PDF_PAGES = int(os.getenv("STATEMENT_MAX_PDF_PAGES", "50"))
OCR_DPI = int(os.getenv("STATEMENT_OCR_DPI", "200"))
OCR_THRESHOLD = int(os.getenv("STATEMENT_OCR_THRESHOLD", "175"))
OCR_ADAPTIVE_RADIUS = int(os.getenv("STATEMENT_OCR_ADAPTIVE_RADIUS", "18"))
OCR_ADAPTIVE_OFFSET = int(os.getenv("STATEMENT_OCR_ADAPTIVE_OFFSET", "8"))
THAI_MONTHS = {
    "มกราคม": 1,
    "กุมภาพันธ์": 2,
    "มีนาคม": 3,
    "เมษายน": 4,
    "พฤษภาคม": 5,
    "มิถุนายน": 6,
    "กรกฎาคม": 7,
    "สิงหาคม": 8,
    "กันยายน": 9,
    "ตุลาคม": 10,
    "พฤศจิกายน": 11,
    "ธันวาคม": 12,
}


def _mask_reference(value: Any) -> str | None:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) < 4:
        return None
    return f"•••• {digits[-4:]}"


def _statement_dates(
    transactions: list[ParsedTransaction],
) -> tuple[str | None, str | None]:
    values = sorted(
        item.transaction_date for item in transactions if item.transaction_date
    )
    return (values[0], values[-1]) if values else (None, None)


def _confidence_warnings(
    confidence: float,
    transaction_date: str | None,
    amount: float | None,
    description: str,
) -> list[str]:
    warnings: list[str] = []
    if confidence < 60:
        warnings.append("OCR confidence ต่ำ ต้องตรวจทานก่อนนำเข้า")
    elif confidence < 80:
        warnings.append("OCR confidence ปานกลาง กรุณาตรวจทาน")
    if not transaction_date:
        warnings.append("อ่านวันที่ไม่สำเร็จ")
    if amount is None:
        warnings.append("อ่านจำนวนเงินไม่สำเร็จ")
    if not description.strip():
        warnings.append("อ่านรายละเอียดไม่สำเร็จ")
    return warnings


def _build_transaction(
    *,
    transaction_date: str | None,
    description: str,
    amount: float | None,
    card_last4: str | None = None,
    transaction_time: str | None = None,
    channel: str | None = None,
    tr_code: str | None = None,
    source_page: int | None = None,
    confidence: float = 100.0,
    warnings: list[str] | None = None,
) -> ParsedTransaction:
    clean_description = re.sub(r"\s+", " ", description).strip()[:500]
    signed_amount = round(float(amount), 2) if amount is not None else None
    row_warnings = list(warnings or [])
    for warning in _confidence_warnings(
        confidence, transaction_date, signed_amount, clean_description
    ):
        if warning not in row_warnings:
            row_warnings.append(warning)
    row_hash = None
    if transaction_date and signed_amount is not None and clean_description:
        row_hash = make_row_hash(
            transaction_date,
            transaction_time,
            signed_amount,
            clean_description,
            channel,
            tr_code,
        )
    return ParsedTransaction(
        transaction_date=transaction_date,
        description=clean_description,
        amount=signed_amount,
        card_last4=card_last4,
        category=guess_category(clean_description, signed_amount or 0),
        transaction_time=transaction_time,
        deposit_amount=(
            signed_amount if signed_amount is not None and signed_amount > 0 else None
        ),
        withdraw_amount=(
            abs(signed_amount)
            if signed_amount is not None and signed_amount < 0
            else None
        ),
        channel=channel,
        tr_code=tr_code,
        row_hash=row_hash,
        source_page=source_page,
        confidence=round(max(0.0, min(100.0, confidence)), 1),
        warnings=row_warnings,
    )

def _tabular_statement(
    transactions: list[ParsedTransaction], extraction_method: str
) -> ParsedStatement:
    start, end = _statement_dates(transactions)
    return ParsedStatement(
        issuer="ไม่ระบุ",
        statement_type="tabular",
        extraction_method=extraction_method,
        masked_reference=None,
        statement_date_from=start,
        statement_date_to=end,
        summary_totals={},
        transactions=transactions,
    )


def _crop_text(page: Any, bbox: tuple[float, float, float, float]) -> str:
    try:
        return re.sub(
            r"\s+",
            " ",
            page.crop(bbox).extract_text(x_tolerance=1, y_tolerance=1) or "",
        ).strip()
    except (ValueError, TypeError):
        return ""


def _parse_scb_pdf(document: Any, full_text: str) -> ParsedStatement:
    transactions: list[ParsedTransaction] = []
    for page_number, page in enumerate(document.pages, start=1):
        words = page.extract_words(x_tolerance=1, y_tolerance=2) or []
        for word in words:
            raw_date = str(word.get("text") or "")
            if float(word.get("x0", 999)) > 75 or not re.fullmatch(
                r"\d{2}/\d{2}/\d{2}", raw_date
            ):
                continue
            top = float(word["top"])
            bottom = float(word["bottom"])
            row_top = max(0.0, top - 1.5)
            row_bottom = min(float(page.height), bottom + 1.5)
            transaction_date = parse_date(raw_date)
            transaction_time = parse_time(
                _crop_text(page, (62, row_top, 93, row_bottom))
            )
            tr_code = _clean_optional(
                _crop_text(page, (93, row_top, 116, row_bottom)), 50
            )
            channel = _clean_optional(
                _crop_text(page, (116, row_top, 160, row_bottom)), 100
            )
            debit = parse_amount(
                _crop_text(page, (160, row_top, 250, row_bottom))
            )
            credit = parse_amount(
                _crop_text(page, (250, row_top, 320, row_bottom))
            )
            description = _crop_text(
                page, (394, row_top, float(page.width), row_bottom)
            )
            warnings: list[str] = []
            amount: float | None
            if debit not in (None, 0) and credit not in (None, 0):
                amount = None
                warnings.append("พบทั้งยอด Debit และ Credit ในแถวเดียว")
            elif credit not in (None, 0):
                amount = abs(float(credit))
            elif debit not in (None, 0):
                amount = -abs(float(debit))
            else:
                amount = None
                warnings.append("ไม่พบยอด Debit หรือ Credit")
            if not description:
                description = " ".join(
                    value for value in (channel, tr_code) if value
                ) or "รายการจาก SCB Statement"
            transactions.append(
                _build_transaction(
                    transaction_date=transaction_date,
                    transaction_time=transaction_time,
                    description=description,
                    amount=amount,
                    channel=channel,
                    tr_code=tr_code,
                    source_page=page_number,
                    confidence=100,
                    warnings=warnings,
                )
            )

    account_match = re.search(
        r"(?:Account\s*No\.?|เลขที่บัญชี)\s*[:.]?\s*([0-9][0-9-]{7,})",
        full_text,
        re.IGNORECASE,
    )
    period_match = re.search(
        r"(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})", full_text
    )
    start, end = _statement_dates(transactions)
    if period_match:
        start = parse_date(period_match.group(1)) or start
        end = parse_date(period_match.group(2)) or end
    valid_amounts = [
        item.amount for item in transactions if item.amount is not None
    ]
    return ParsedStatement(
        issuer="ธนาคารไทยพาณิชย์ (SCB)",
        statement_type="bank_savings",
        extraction_method="pdf_text",
        masked_reference=_mask_reference(
            account_match.group(1) if account_match else None
        ),
        statement_date_from=start,
        statement_date_to=end,
        summary_totals={
            "total_credit": round(sum(value for value in valid_amounts if value > 0), 2),
            "total_debit": round(abs(sum(value for value in valid_amounts if value < 0)), 2),
        },
        transactions=transactions,
        warnings=[] if transactions else ["ไม่พบรายการเคลื่อนไหวใน SCB Statement"],
    )


def _parse_tesseract_tsv(tsv: str, page_number: int) -> list[OcrLine]:
    grouped: dict[tuple[int, int, int], list[PositionedWord]] = {}
    # Tesseract TSV does not quote fields. Treating a literal OCR `"` as CSV
    # quoting can otherwise swallow many following rows into one word.
    reader = csv.DictReader(
        io.StringIO(tsv), delimiter="\t", quoting=csv.QUOTE_NONE
    )
    for row in reader:
        text = str(row.get("text") or "").strip()
        try:
            confidence = float(row.get("conf") or -1)
            if not text or confidence < 0:
                continue
            key = (
                int(row.get("block_num") or 0),
                int(row.get("par_num") or 0),
                int(row.get("line_num") or 0),
            )
            grouped.setdefault(key, []).append(
                PositionedWord(
                    text=text,
                    left=float(row.get("left") or 0),
                    top=float(row.get("top") or 0),
                    width=float(row.get("width") or 0),
                    height=float(row.get("height") or 0),
                    confidence=confidence,
                )
            )
        except (TypeError, ValueError):
            continue

    lines: list[OcrLine] = []
    for words in grouped.values():
        words.sort(key=lambda item: item.left)
        weight = sum(max(1, len(item.text)) for item in words)
        confidence = sum(
            item.confidence * max(1, len(item.text)) for item in words
        ) / max(1, weight)
        lines.append(
            OcrLine(
                text=" ".join(item.text for item in words),
                confidence=confidence,
                page=page_number,
                words=words,
            )
        )
    return sorted(
        lines,
        key=lambda line: (
            min((word.top for word in line.words), default=0),
            min((word.left for word in line.words), default=0),
        ),
    )


def _ocr_image_variants(source: Any) -> dict[str, Any]:
    """Build OCR-ready images without assuming the whole page has one background.

    Krungsri statements mix white cells, grey bands and small text. A single
    global threshold can erase numbers printed on grey cells, so the summary
    page is also read from contrast-preserving grayscale and locally adaptive
    black/white variants.
    """
    from PIL import ImageChops, ImageEnhance, ImageFilter, ImageOps

    grayscale = ImageOps.autocontrast(ImageOps.grayscale(source), cutoff=1)
    grayscale = ImageEnhance.Contrast(grayscale).enhance(1.25)
    denoised = grayscale.filter(ImageFilter.MedianFilter(3))
    global_bw = denoised.point(
        lambda value: 255 if value > OCR_THRESHOLD else 0
    )

    local_mean = denoised.filter(ImageFilter.BoxBlur(OCR_ADAPTIVE_RADIUS))
    local_difference = ImageChops.subtract(
        denoised,
        local_mean,
        scale=1.0,
        offset=128,
    )
    adaptive_cutoff = max(0, min(255, 128 - OCR_ADAPTIVE_OFFSET))
    adaptive_bw = local_difference.point(
        lambda value: 255 if value > adaptive_cutoff else 0
    )
    return {
        "binary": global_bw,
        "adaptive": adaptive_bw,
        "grayscale": denoised,
    }


def _run_tesseract_tsv(
    image_path: Path,
    page_number: int,
    *,
    psm: int,
    environment: dict[str, str],
    error_label: str,
    timeout: int = 60,
    languages: str = "tha+eng",
) -> list[OcrLine]:
    try:
        completed = subprocess.run(
            [
                "tesseract",
                str(image_path),
                "stdout",
                "-l",
                languages,
                "--psm",
                str(psm),
                "tsv",
            ],
            check=True,
            capture_output=True,
            text=True,
            timeout=timeout,
            env=environment,
        )
    except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
        raise ValueError(
            f"{error_label}หน้า {page_number} ไม่สำเร็จ กรุณาตรวจคุณภาพไฟล์"
        ) from exc
    return _parse_tesseract_tsv(completed.stdout, page_number)


def _map_scaled_ocr_lines(
    lines: list[OcrLine],
    *,
    crop_left: float,
    crop_top: float,
    scale: float,
) -> list[OcrLine]:
    """Map OCR coordinates from an enlarged crop back onto the source image."""
    result: list[OcrLine] = []
    for line in lines:
        words = [
            PositionedWord(
                text=word.text,
                left=crop_left + word.left / scale,
                top=crop_top + word.top / scale,
                width=word.width / scale,
                height=word.height / scale,
                confidence=word.confidence,
            )
            for word in line.words
        ]
        result.append(
            OcrLine(
                text=line.text,
                confidence=line.confidence,
                page=line.page,
                words=words,
            )
        )
    return result


def _deduplicate_ocr_lines(lines: list[OcrLine]) -> list[OcrLine]:
    """Merge identical lines from OCR variants, keeping the clearest reading."""
    selected: dict[tuple[int, str, int, int], OcrLine] = {}
    for line in lines:
        top, left = _line_position(line)
        key = (
            line.page,
            re.sub(r"\s+", " ", line.text).strip().casefold(),
            round(top / 12),
            round(left / 12),
        )
        current = selected.get(key)
        if current is None or line.confidence > current.confidence:
            selected[key] = line
    return sorted(
        selected.values(),
        key=lambda line: (line.page, *_line_position(line)),
    )


def _ocr_pdf(
    data: bytes,
    page_count: int,
    sparse_pages: set[int] | None = None,
    stop_before_amex_supplement: bool = False,
) -> list[OcrLine]:
    from PIL import Image

    if not shutil.which("pdftoppm"):
        raise ValueError("เซิร์ฟเวอร์ยังไม่ได้ติดตั้ง Poppler สำหรับแปลง PDF")
    if not shutil.which("tesseract"):
        raise ValueError("เซิร์ฟเวอร์ยังไม่ได้ติดตั้ง Tesseract OCR")
    with tempfile.TemporaryDirectory(prefix="statement-ocr-") as temp_dir:
        pdf_path = Path(temp_dir) / "statement.pdf"
        output_prefix = Path(temp_dir) / "page"
        pdf_path.write_bytes(data)
        try:
            subprocess.run(
                [
                    "pdftoppm",
                    "-png",
                    "-r",
                    str(OCR_DPI),
                    "-f",
                    "1",
                    "-l",
                    str(page_count),
                    str(pdf_path),
                    str(output_prefix),
                ],
                check=True,
                capture_output=True,
                timeout=300,
            )
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
            raise ValueError("ไม่สามารถแปลง PDF เป็นภาพสำหรับ OCR ได้") from exc

        images = sorted(Path(temp_dir).glob("page-*.png"))
        if len(images) != page_count:
            raise ValueError("จำนวนหน้าที่แปลงเพื่อ OCR ไม่ครบ")
        result: list[OcrLine] = []
        ocr_environment = {**os.environ, "OMP_THREAD_LIMIT": "1"}
        for page_number, image_path in enumerate(images, start=1):
            page_lines: list[OcrLine] = []
            processed_paths: dict[str, Path] = {}
            with Image.open(image_path) as source:
                variants = _ocr_image_variants(source)
                for variant_name, processed in variants.items():
                    processed_path = (
                        Path(temp_dir)
                        / f"processed-{page_number:03d}-{variant_name}.png"
                    )
                    processed.save(processed_path)
                    processed_paths[variant_name] = processed_path
            page_lines.extend(
                _run_tesseract_tsv(
                    processed_paths["binary"],
                    page_number,
                    psm=6,
                    environment=ocr_environment,
                    error_label="OCR ",
                )
            )
            if page_number in (sparse_pages or set()):
                for variant_name in ("binary", "adaptive", "grayscale"):
                    page_lines.extend(
                        _run_tesseract_tsv(
                            processed_paths[variant_name],
                            page_number,
                            psm=11,
                            environment=ocr_environment,
                            error_label="OCR แบบแยกบล็อก ",
                        )
                    )
            if stop_before_amex_supplement and _is_amex_supplement_page(
                "\n".join(line.text for line in page_lines)
            ):
                break
            result.extend(page_lines)
        return _deduplicate_ocr_lines(result)


def _infer_statement_year(text: str) -> tuple[int, int]:
    parsed_dates = [
        parse_date(value)
        for value in re.findall(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", text)
    ]
    valid = [datetime.fromisoformat(value) for value in parsed_dates if value]
    if valid:
        latest = max(valid)
        return latest.year, latest.month
    today = date.today()
    return today.year, today.month


def _parse_thai_day_month(text: str) -> tuple[int, int, int] | None:
    clean = _normalise_ocr_line(text)
    clean = re.sub(r"^([([]?)\s*[Oo0][fIl1]\b", r"\g<1>01", clean)
    for month_name, month in THAI_MONTHS.items():
        match = re.search(rf"(\d{{1,2}})\s*{month_name}", clean)
        if match:
            return int(match.group(1)), month, match.end()
    return None


def _normalise_ocr_line(text: str) -> str:
    clean = re.sub(r"\s+", " ", text).strip()
    previous = None
    while clean != previous:
        previous = clean
        clean = re.sub(r"(?<=[ก-๙])\s+(?=[ก-๙])", "", clean)
    return clean


def _is_amex_supplement_page(text: str) -> bool:
    """Identify the Amex rewards/news page that follows transaction pages."""
    compact = re.sub(r"\s+", "", _normalise_ocr_line(text)).casefold()
    has_membership_rewards = bool(
        re.search(r"membershipreward(?:s)?", compact)
    )
    return "รายการmembershiprewards" in compact or (
        has_membership_rewards and "สิทธิพิเศษ" in compact
    )


def _ocr_card_last4(text: str, filename: str) -> str | None:
    filename_match = re.search(
        r"(?:x{2,}|amex).*?(\d{4})$", Path(filename).stem, re.IGNORECASE
    )
    if filename_match:
        return filename_match.group(1)
    labelled = re.search(
        r"(?:Card\s*(?:Number|No\.?)|หมายเลขบัตร).*?([0-9Xx* -]{10,})",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    if labelled:
        value = _last4(labelled.group(1))
        if value:
            return value
    return _last4(Path(filename).stem)


def _deduplicate_transactions(
    transactions: list[ParsedTransaction],
) -> list[ParsedTransaction]:
    seen: set[str] = set()
    result: list[ParsedTransaction] = []
    for item in transactions:
        key = item.row_hash or "|".join(
            [
                item.transaction_date or "",
                item.description,
                "" if item.amount is None else f"{item.amount:.2f}",
                str(item.source_page or ""),
            ]
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _extract_amex_trusted_amounts(
    document: Any, max_page: int | None = None
) -> list[TrustedAmount]:
    result: list[TrustedAmount] = []
    scale = OCR_DPI / 72
    known_locations = {
        "DUBLIN",
        "SINGAPORE",
        "FACEBOOK.COM",
    }
    for page_number, page in enumerate(document.pages, start=1):
        if max_page is not None and page_number > max_page:
            break
        words = page.extract_words(x_tolerance=1, y_tolerance=2) or []
        credit_tops = [
            float(word["top"])
            for word in words
            if str(word.get("text") or "").strip().upper() == "CR"
        ]
        for word in words:
            text = str(word.get("text") or "").strip()
            if float(word.get("x0") or 0) < float(page.width) * 0.72:
                continue
            if not re.fullmatch(r"\d{1,3}(?:,\d{3})*\.\d{2}", text):
                continue
            amount = parse_amount(text)
            if amount is None:
                continue
            top = float(word["top"])
            same_row = sorted(
                (
                    candidate
                    for candidate in words
                    if 70 <= float(candidate.get("x0") or 0) < float(page.width) * 0.72
                    and abs(float(candidate.get("top") or 0) - top) <= 4
                    and "(cid:" not in str(candidate.get("text") or "")
                ),
                key=lambda candidate: float(candidate.get("x0") or 0),
            )
            row_tokens = [
                str(candidate.get("text") or "").strip()
                for candidate in same_row
                if str(candidate.get("text") or "").strip()
            ]
            reference_index = next(
                (
                    index
                    for index, token in enumerate(row_tokens)
                    if token.startswith("*") and len(token) >= 5
                ),
                None,
            )
            reference = None
            location = None
            merchant_tokens = list(row_tokens)
            if reference_index is not None:
                reference = row_tokens[reference_index].lstrip("*").strip() or None
                merchant_tokens = row_tokens[:reference_index]
                trailing = row_tokens[reference_index + 1 :]
                location = " ".join(trailing).strip() or None
            elif row_tokens and row_tokens[-1].upper() in known_locations:
                location = row_tokens[-1]
                merchant_tokens = row_tokens[:-1]
            merchant = " ".join(merchant_tokens).strip() or None
            result.append(
                TrustedAmount(
                    page=page_number,
                    top=top * scale,
                    amount=abs(amount),
                    is_credit=any(0 <= credit_top - top <= 15 for credit_top in credit_tops),
                    merchant=merchant,
                    reference=reference,
                    location=location,
                )
            )
    return result


def _parse_amex_ocr(
    lines: list[OcrLine],
    filename: str,
    full_text: str,
    trusted_amounts: list[TrustedAmount] | None = None,
) -> ParsedStatement:
    year, statement_month = _infer_statement_year(full_text)
    card_last4 = _ocr_card_last4(full_text, filename)
    page_texts: dict[int, list[str]] = {}
    for line in lines:
        page_texts.setdefault(line.page, []).append(line.text)
    supplement_page = next(
        (
            page
            for page, page_lines in sorted(page_texts.items())
            if _is_amex_supplement_page("\n".join(page_lines))
        ),
        None,
    )
    transactions: list[ParsedTransaction] = []
    for index, line in enumerate(lines):
        if supplement_page is not None and line.page >= supplement_page:
            break
        stripped = _normalise_ocr_line(line.text)
        # The issuer's total label is an additional row-level boundary before
        # the Membership Rewards/news pages begin.
        if "รายการทั้งหมดสำหรับ" in stripped:
            break
        if re.match(r"^[([{]", stripped):
            continue
        parsed_day = _parse_thai_day_month(stripped)
        if not parsed_day:
            continue
        day, month, date_end = parsed_day
        transaction_year = year - 1 if month > statement_month + 1 else year
        try:
            transaction_date = date(transaction_year, month, day).isoformat()
        except ValueError:
            transaction_date = None

        combined = stripped
        combined_confidence = line.confidence
        amount_match = AMOUNT_AT_END_PATTERN.search(combined)
        if not amount_match and index + 1 < len(lines):
            following = lines[index + 1]
            if following.page == line.page and not _parse_thai_day_month(following.text):
                combined = f"{stripped} {_normalise_ocr_line(following.text)}"
                combined_confidence = min(line.confidence, following.confidence)
                amount_match = AMOUNT_AT_END_PATTERN.search(combined)
        line_top = min((word.top for word in line.words), default=-1)
        trusted = min(
            (
                item
                for item in (trusted_amounts or [])
                if item.page == line.page and abs(item.top - line_top) <= 20
            ),
            key=lambda item: abs(item.top - line_top),
            default=None,
        )
        if not amount_match:
            description = combined[date_end:].strip(" -|")
            if trusted and trusted.merchant:
                description = trusted.merchant
            if trusted and not description:
                nearby = min(
                    (
                        candidate
                        for candidate in lines
                        if candidate.page == line.page
                        and candidate is not line
                        and abs(
                            min(
                                (word.top for word in candidate.words),
                                default=-1000,
                            )
                            - line_top
                        )
                        <= 18
                        and re.search(
                            r"PAYMENT|TIKTOK|FACEBK|FACEBOOK",
                            candidate.text,
                            re.IGNORECASE,
                        )
                    ),
                    key=lambda candidate: abs(
                        min(
                            (word.top for word in candidate.words),
                            default=-1000,
                        )
                        - line_top
                    ),
                    default=None,
                )
                if nearby:
                    description = _normalise_ocr_line(nearby.text)
            if trusted and description:
                transactions.append(
                    _build_transaction(
                        transaction_date=transaction_date,
                        description=description,
                        amount=-trusted.amount if trusted.is_credit else trusted.amount,
                        card_last4=card_last4,
                        channel=trusted.location,
                        tr_code=trusted.reference,
                        source_page=line.page,
                        confidence=combined_confidence,
                    )
                )
                continue
            if description and len(description) > 3:
                transactions.append(
                    _build_transaction(
                        transaction_date=transaction_date,
                        description=description,
                        amount=None,
                        card_last4=card_last4,
                        source_page=line.page,
                        confidence=combined_confidence,
                    )
                )
            continue
        raw_amount = amount_match.group(1)
        amount = parse_amount(raw_amount)
        description = combined[date_end:amount_match.start()].strip(" -|")
        if trusted and trusted.merchant:
            description = trusted.merchant
        next_text = (
            _normalise_ocr_line(lines[index + 1].text)
            if index + 1 < len(lines) and lines[index + 1].page == line.page
            else ""
        )
        is_credit = bool(
            re.search(r"\bCR\b", f"{combined} {next_text}", re.IGNORECASE)
            or re.search(
                r"PAYMENT\s+AT\s+BANK|PAYMENT\s+RECEIVED|REFUND|CASHBACK",
                description,
                re.IGNORECASE,
            )
        )
        if trusted:
            amount = trusted.amount
            is_credit = trusted.is_credit or is_credit
        if amount is not None:
            amount = -abs(amount) if is_credit else abs(amount)
        if not description or re.search(
            r"วันที่|ยอดเดิม|ยอดรวม|วงเงิน|payment due", description, re.IGNORECASE
        ):
            continue
        transactions.append(
            _build_transaction(
                transaction_date=transaction_date,
                description=description,
                amount=amount,
                card_last4=card_last4,
                channel=trusted.location if trusted else None,
                tr_code=trusted.reference if trusted else None,
                source_page=line.page,
                confidence=combined_confidence,
            )
        )

    transactions = _deduplicate_transactions(transactions)
    start, end = _statement_dates(transactions)
    warnings = [
        "Amex ใช้ OCR เพราะ text layer ภาษาไทยไม่สมบูรณ์",
        "กรุณาเทียบยอดรวมรายการกับหน้าสรุป Amex ก่อนยืนยัน เพราะ OCR อาจอ่านแถวพื้นเทาหรือตัวเลขบางรายการไม่ครบ",
    ]
    if not transactions:
        warnings.append("ไม่พบรายการธุรกรรมจาก Amex Statement")
    valid_amounts = [
        item.amount for item in transactions if item.amount is not None
    ]
    return ParsedStatement(
        issuer="American Express",
        statement_type="credit_card",
        extraction_method="ocr",
        masked_reference=_mask_reference(card_last4),
        statement_date_from=start,
        statement_date_to=end,
        summary_totals={
            "total_charges": round(sum(value for value in valid_amounts if value > 0), 2),
            "total_credits": round(abs(sum(value for value in valid_amounts if value < 0)), 2),
        },
        transactions=transactions,
        warnings=warnings,
    )


def _summary_amount(lines: list[OcrLine], pattern: str) -> float | None:
    for line in lines:
        if re.search(pattern, line.text, re.IGNORECASE):
            amounts = list(
                re.finditer(
                    r"[-+]?(?:\d{1,3}(?:,\d{3})*|\d+)(?:\.\d{2})",
                    line.text,
                )
            )
            if amounts:
                return parse_amount(amounts[-1].group(0))
    return None


def _line_position(line: OcrLine) -> tuple[float, float]:
    return (
        min((word.top for word in line.words), default=0),
        min((word.left for word in line.words), default=0),
    )


def _line_right(line: OcrLine) -> float:
    return max(
        (word.left + word.width for word in line.words),
        default=_line_position(line)[1],
    )


def _parse_ocr_currency_values(text: str) -> list[float]:
    """Parse currency values after common OCR separator/decimal damage."""
    pattern = re.compile(
        r"[-+]?(?:"
        r"\d{1,3}(?:,\d{3})+\.\d{2}"
        r"|\d{1,3}[.,]\d{3}(?:\s*(?:[.,]\s*)?(?:\d{2}|[oO]{2}))?"
        r"|\d{2,}\s+\d{2}"
        r"|\d+\.\d{2}"
        r"|\d{5,}"
        r"|0+"
        r")"
    )
    values: list[float] = []
    for match in pattern.finditer(text.replace("฿", "")):
        token = match.group(0).strip()
        sign = -1 if token.startswith("-") else 1
        unsigned = token.lstrip("+-")
        value: float | None = None
        if re.fullmatch(r"\d{1,3}(?:,\d{3})+\.\d{2}", unsigned):
            value = parse_amount(token)
        elif re.fullmatch(
            r"\d{1,3}[.,]\d{3}(?:\s*(?:[.,]\s*)?(?:\d{2}|[oO]{2}))?",
            unsigned,
        ):
            digits = re.sub(r"\D", "", unsigned.replace("o", "0").replace("O", "0"))
            has_cents = bool(
                re.search(
                    r"[.,]\d{3}(?:\s*(?:[.,]\s*)?(?:\d{2}|[oO]{2}))$",
                    unsigned,
                )
            )
            value = float(digits) / 100 if has_cents else float(digits)
            value *= sign
        elif re.fullmatch(r"\d{2,}\s+\d{2}", unsigned):
            value = sign * float(re.sub(r"\D", "", unsigned)) / 100
        elif re.fullmatch(r"\d{5,}", unsigned):
            value = sign * float(unsigned) / 100
        else:
            value = parse_amount(token)
        if value is not None and value not in values:
            values.append(value)
    return values


def _line_numeric_values(line: OcrLine) -> list[float]:
    return _parse_ocr_currency_values(line.text)


def _summary_amount_nearby(
    lines: list[OcrLine], pattern: str, *, max_vertical_distance: float = 40
) -> float | None:
    matches: list[tuple[bool, float, float, float, float]] = []
    for label in lines:
        label_match = re.search(pattern, label.text, re.IGNORECASE)
        if not label_match:
            continue
        # Only consider numbers printed after the matched label. OCR variants
        # sometimes merge barcode/account digits appearing before a label into
        # the same line.
        direct_values = _parse_ocr_currency_values(
            label.text[label_match.end() :]
        )
        for direct in direct_values:
            matches.append((True, 0, 0, label.confidence, direct))
        label_top, label_left = _line_position(label)
        label_right = _line_right(label)
        for candidate in lines:
            if candidate.page != label.page or candidate is label:
                continue
            candidate_top, candidate_left = _line_position(candidate)
            vertical_distance = abs(candidate_top - label_top)
            if (
                vertical_distance > max_vertical_distance
                or candidate_left <= max(label_left + 80, label_right)
            ):
                continue
            for value in _line_numeric_values(candidate):
                horizontal_gap = candidate_left - label_right
                matches.append(
                    (
                        False,
                        horizontal_gap,
                        vertical_distance,
                        candidate.confidence,
                        value,
                    )
                )
    if not matches:
        return None
    # Direct label/value rows are strongest. For split columns, the closest
    # value to the label's right edge is safer than the largest amount on the
    # same horizontal band (which may belong to a neighbouring summary card).
    matches.sort(
        key=lambda item: (
            0 if item[0] else 1,
            item[1],
            0 if item[0] else -abs(item[4]),
            item[2],
            -item[3],
        )
    )
    return matches[0][4]


def _amount_above_label(
    lines: list[OcrLine], pattern: str, *, max_vertical_distance: float = 80
) -> float | None:
    candidates: list[tuple[float, float, float, float]] = []
    for label in lines:
        if not re.search(pattern, label.text, re.IGNORECASE):
            continue
        label_top, _ = _line_position(label)
        label_right = _line_right(label)
        for candidate in lines:
            candidate_top, candidate_left = _line_position(candidate)
            vertical_distance = label_top - candidate_top
            if (
                candidate.page != label.page
                or not 12 <= vertical_distance <= max_vertical_distance
                or candidate_left <= label_right + 20
            ):
                continue
            for value in _line_numeric_values(candidate):
                candidates.append(
                    (
                        vertical_distance,
                        -abs(value),
                        -candidate.confidence,
                        value,
                    )
                )
    if not candidates:
        return None
    candidates.sort(key=lambda item: (item[0], item[1], item[2]))
    return candidates[0][3]


def _amount_below_column_label(
    lines: list[OcrLine], pattern: str, *, max_vertical_distance: float = 85
) -> float | None:
    candidates: list[tuple[float, float, float, float, float]] = []
    for label in lines:
        if not re.search(pattern, label.text, re.IGNORECASE):
            continue
        label_top, label_left = _line_position(label)
        label_right = _line_right(label)
        for candidate in lines:
            candidate_top, candidate_left = _line_position(candidate)
            vertical_distance = candidate_top - label_top
            if (
                candidate.page != label.page
                or not 8 <= vertical_distance <= max_vertical_distance
                or not label_left - 40 <= candidate_left <= label_right + 40
            ):
                continue
            for value in _line_numeric_values(candidate):
                candidates.append(
                    (
                        abs(candidate_left - label_left),
                        vertical_distance,
                        -abs(value),
                        -candidate.confidence,
                        value,
                    )
                )
    if not candidates:
        return None
    # Prefer the value aligned immediately below the most specific column
    # heading; a PSM 6 line can contain every heading and start far to the left.
    candidates.sort(key=lambda item: (item[0], item[1], item[2], item[3]))
    return candidates[0][4]


def _parse_krungsri_ocr(
    lines: list[OcrLine], filename: str, full_text: str
) -> ParsedStatement:
    card_last4 = _ocr_card_last4(full_text, filename)
    transactions: list[ParsedTransaction] = []
    in_transactions = False
    for line in lines:
        text = line.text.strip()
        if re.search(r"Transaction\s*Date", text, re.IGNORECASE):
            in_transactions = True
            continue
        if re.search(
            r"Total\s*Payments|Summary\s*Of\s*All\s*Accounts", text, re.IGNORECASE
        ):
            in_transactions = False
        if not in_transactions:
            continue
        match = PDF_LINE_PATTERN.match(text)
        if not match:
            continue
        transaction_date = parse_date(match.group(1))
        description = match.group(2).strip()
        amount = parse_amount(match.group(3))
        if re.search(r"payment|ชำระ|refund|คืนเงิน", description, re.IGNORECASE):
            amount = -abs(amount) if amount is not None else None
        transactions.append(
            _build_transaction(
                transaction_date=transaction_date,
                description=description,
                amount=amount,
                card_last4=card_last4,
                source_page=line.page,
                confidence=line.confidence,
            )
        )
    transactions = _deduplicate_transactions(transactions)
    start, end = _statement_dates(transactions)
    closing_matches = [
        parse_date(value)
        for value in re.findall(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", full_text)
    ]
    closing_dates = sorted(value for value in closing_matches if value)
    if closing_dates:
        end = closing_dates[-1]
    warnings: list[str] = []
    if not transactions:
        warnings.append("Statement นี้ไม่พบรายการธุรกรรมในรอบบัญชี")
    summary_totals: dict[str, float] = {}
    total_due = _summary_amount_nearby(
        lines,
        r"Total\s*Payment\s*(?:Due|Oue)\s*For\s*Credit\s*Card",
    )
    if total_due is None:
        total_due = _summary_amount_nearby(
            lines, r"Total\s*Payment\s*(?:Due|Oue)"
        )
    minimum_due = _summary_amount_nearby(
        lines, r"Minimum\s*Payment\s*(?:Due|Oue)"
    )
    shared_credit_line = _summary_amount_nearby(
        lines,
        r"Shared\s*Credit\s*Line(?!\s*Used)",
    )
    shared_credit_line_used = _summary_amount_nearby(
        lines,
        r"Shared\s*Credit\s*Line\s*Used",
        max_vertical_distance=60,
    )
    if shared_credit_line_used in (None, 0):
        shared_credit_line_used = _amount_below_column_label(
            lines,
            r"Outstanding\s*Balance",
        )
    shared_available_credit = _summary_amount_nearby(
        lines,
        r"Shared\s*(?:Available|Avalable)\s*(?:Credit|Cre.?it)\s*(?:Limit|Lime)",
    )
    previous_balance = _summary_amount_nearby(
        lines,
        r"Previous\s*Statement\s*Balance",
    )
    total_payments = _summary_amount_nearby(
        lines,
        r"Total\s*Payments",
    )
    if previous_balance is None:
        previous_balance = _amount_above_label(
            lines,
            r"Total\s*Payments",
        )
    compact_text = re.sub(r"\s+", "", _normalise_ocr_line(full_text))
    if (
        total_due is not None
        and total_due > 0
        and (
            re.search(r"ไม่ม.?ยอด(?:ที่)?ต้องช", compact_text)
            or (
                previous_balance is not None
                and previous_balance < 0
                and total_payments == 0
            )
        )
    ):
        total_due = -abs(total_due)
    if total_due is not None:
        summary_totals["total_payment_due"] = round(total_due, 2)
    if minimum_due is not None:
        summary_totals["minimum_payment_due"] = round(minimum_due, 2)
    if shared_credit_line is not None:
        summary_totals["shared_credit_line"] = round(shared_credit_line, 2)
    if shared_credit_line_used is not None:
        summary_totals["shared_credit_line_used"] = round(
            shared_credit_line_used, 2
        )
    if shared_available_credit is not None:
        summary_totals["shared_available_credit_limit"] = round(
            shared_available_credit, 2
        )
    if previous_balance is not None:
        summary_totals["previous_statement_balance"] = round(
            previous_balance, 2
        )
    if total_payments is not None:
        summary_totals["total_payments"] = round(total_payments, 2)
    if not summary_totals:
        warnings.append("อ่านยอดสรุปไม่สำเร็จ กรุณาตรวจหน้าสรุปของ Krungsri")
    return ParsedStatement(
        issuer="Krungsriayudhya Card",
        statement_type="credit_card",
        extraction_method="ocr",
        masked_reference=_mask_reference(card_last4),
        statement_date_from=start,
        statement_date_to=end,
        summary_totals=summary_totals,
        transactions=transactions,
        warnings=warnings,
    )


def _parse_generic_ocr(
    lines: list[OcrLine], filename: str
) -> ParsedStatement:
    transactions: list[ParsedTransaction] = []
    for line in lines:
        match = PDF_LINE_PATTERN.match(line.text)
        if not match:
            continue
        amount = parse_amount(match.group(3))
        transactions.append(
            _build_transaction(
                transaction_date=parse_date(match.group(1)),
                description=match.group(2),
                amount=amount,
                card_last4=_last4(Path(filename).stem),
                source_page=line.page,
                confidence=line.confidence,
            )
        )
    if not transactions:
        raise ValueError(
            "OCR อ่านตัวอักษรได้ แต่ยังแยกวันที่ รายละเอียด และยอดเงินไม่ได้"
        )
    start, end = _statement_dates(transactions)
    return ParsedStatement(
        issuer="ไม่ระบุ",
        statement_type="unknown",
        extraction_method="ocr",
        masked_reference=None,
        statement_date_from=start,
        statement_date_to=end,
        summary_totals={},
        transactions=_deduplicate_transactions(transactions),
        warnings=["ไม่พบ adapter เฉพาะ จึงใช้ตัวแยก OCR รูปแบบทั่วไป"],
    )


def _parse_pdf_text_generic(document: Any) -> list[ParsedTransaction]:
    fallback: list[ParsedTransaction] = []
    for page_number, page in enumerate(document.pages, start=1):
        for table in page.extract_tables() or []:
            try:
                rows = parse_rows([list(row) for row in table if row])
                for item in rows:
                    item.source_page = page_number
                return rows
            except ValueError:
                pass
        for line in (page.extract_text() or "").splitlines():
            match = PDF_LINE_PATTERN.match(line)
            if not match:
                continue
            amount = parse_amount(match.group(3))
            description = match.group(2).strip()
            if parse_date(match.group(1)) and amount is not None and description:
                fallback.append(
                    _build_transaction(
                        transaction_date=parse_date(match.group(1)),
                        description=description,
                        amount=amount,
                        source_page=page_number,
                    )
                )
    return fallback


def parse_pdf_with_metadata(filename: str, data: bytes) -> ParsedStatement:
    import pdfplumber

    try:
        document = pdfplumber.open(io.BytesIO(data))
    except Exception as exc:
        raise ValueError("ไฟล์ PDF เสียหรือไม่สามารถเปิดอ่านได้") from exc
    with document:
        page_count = len(document.pages)
        if page_count < 1:
            raise ValueError("PDF ไม่มีหน้าเอกสาร")
        if page_count > MAX_PDF_PAGES:
            raise ValueError(f"PDF มีเกิน {MAX_PDF_PAGES} หน้า")
        page_texts = [page.extract_text() or "" for page in document.pages]
        full_text = "\n".join(page_texts)
        normalised = full_text.lower()
        if (
            "the siam commercial bank public company limited" in normalised
            and "statement of saving account" in normalised
        ):
            return _parse_scb_pdf(document, full_text)

        is_amex = (
            "americanexpress.com" in normalised
            or "payment at bank" in normalised
            or "american express" in normalised
            or "amex" in filename.lower()
        )
        amex_supplement_page = next(
            (
                page_number
                for page_number, page_text in enumerate(page_texts, start=1)
                if _is_amex_supplement_page(page_text)
            ),
            None,
        )
        if amex_supplement_page is not None:
            is_amex = True
        last_amex_transaction_page = (
            amex_supplement_page - 1
            if amex_supplement_page is not None
            else page_count
        )
        trusted_amex_amounts = (
            _extract_amex_trusted_amounts(
                document, max_page=last_amex_transaction_page
            )
            if is_amex
            else []
        )
        corrupted_text = full_text.count("(cid:") >= 20
        if not is_amex and not corrupted_text:
            transactions = _parse_pdf_text_generic(document)
            if transactions:
                start, end = _statement_dates(transactions)
                return ParsedStatement(
                    issuer="ไม่ระบุ",
                    statement_type="unknown",
                    extraction_method="pdf_text",
                    masked_reference=None,
                    statement_date_from=start,
                    statement_date_to=end,
                    summary_totals={},
                    transactions=transactions,
                )

    ocr_page_count = last_amex_transaction_page if is_amex else page_count
    ocr_lines = (
        _ocr_pdf(
            data,
            ocr_page_count,
            sparse_pages={1},
            stop_before_amex_supplement=True,
        )
        if ocr_page_count > 0
        else []
    )
    ocr_text = "\n".join(line.text for line in ocr_lines)
    ocr_normalised = ocr_text.lower()
    if (
        is_amex
        or "american express" in ocr_normalised
        or "payment at bank" in ocr_normalised
        or "amex" in filename.lower()
    ):
        return _parse_amex_ocr(
            ocr_lines, filename, ocr_text, trusted_amex_amounts
        )
    if (
        "krungsriayudhya card" in ocr_normalised
        or "krungsri" in ocr_normalised
        or "กรุงศรี" in ocr_text
    ):
        return _parse_krungsri_ocr(ocr_lines, filename, ocr_text)
    return _parse_generic_ocr(ocr_lines, filename)


def parse_pdf(data: bytes) -> list[ParsedTransaction]:
    return parse_pdf_with_metadata("statement.pdf", data).transactions


def parse_statement_with_metadata(filename: str, data: bytes) -> ParsedStatement:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        return _tabular_statement(parse_csv(data), "csv")
    if extension == ".xlsx":
        return _tabular_statement(parse_xlsx(data), "spreadsheet")
    if extension == ".pdf":
        return parse_pdf_with_metadata(filename, data)
    raise ValueError("รองรับเฉพาะไฟล์ .csv, .xlsx และ .pdf")


def parse_statement(filename: str, data: bytes) -> list[ParsedTransaction]:
    return parse_statement_with_metadata(filename, data).transactions


# ── Lightweight local OCR — image evidence uploads ───────────────────────────

EVIDENCE_IMAGE_MAX_DIMENSION = int(os.getenv("STATEMENT_IMAGE_MAX_DIMENSION", "1800"))
EVIDENCE_OCR_TIMEOUT_SECONDS = int(os.getenv("STATEMENT_IMAGE_OCR_TIMEOUT_SECONDS", "60"))


def prepare_evidence_image(data: bytes) -> bytes:
    """Validate, orient and shrink evidence without adding JPEG artifacts."""
    from PIL import Image, ImageOps, UnidentifiedImageError

    try:
        with Image.open(io.BytesIO(data)) as source:
            image = ImageOps.exif_transpose(source).convert("RGB")
            image.thumbnail(
                (EVIDENCE_IMAGE_MAX_DIMENSION, EVIDENCE_IMAGE_MAX_DIMENSION),
                Image.Resampling.LANCZOS,
            )
            output = io.BytesIO()
            # Ads evidence is usually a screenshot with 8-12px text. Re-encoding
            # it as JPEG makes character edges fuzzy and hurts Thai OCR, while a
            # lossless PNG remains small for these mostly-white screens.
            image.save(output, format="PNG", optimize=True)
    except (UnidentifiedImageError, OSError, ValueError) as exc:
        raise ValueError("ไฟล์รูปภาพเสียหรือไม่ใช่รูปภาพที่ระบบรองรับ") from exc
    return output.getvalue()


def _ads_platform(text: str) -> str:
    clean = text.casefold()
    if "facebook" in clean or "meta" in clean or "fbads" in clean:
        return "Facebook/Meta Ads"
    if "google" in clean:
        return "Google Ads"
    if "tiktok" in clean:
        return "TikTok Ads"
    return "หลักฐานค่าโฆษณา"


META_MONTH_MARKERS = {
    1: ("มค", "มกราคม"),
    2: ("กพ", "กุมภาพันธ์"),
    3: ("มีค", "มีนาคม"),
    4: ("เมย", "เมษายน"),
    5: ("พค", "พฤษภาคม"),
    6: ("มิย", "มิถุนายน"),
    7: ("กค", "กรกฎาคม"),
    8: ("สค", "สิงหาคม"),
    9: ("กย", "กันยายน"),
    10: ("ตค", "ตุลาคม"),
    11: ("พย", "พฤศจิกายน"),
    12: ("ธค", "ธันวาคม"),
}


def _meta_compact_text(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9ก-๙]+", "", text).casefold()


def _meta_month(text: str) -> int | None:
    compact = _meta_compact_text(text)
    for month, markers in META_MONTH_MARKERS.items():
        if any(marker in compact for marker in markers):
            return month
    return None


def _meta_year(text: str) -> int | None:
    years = [
        int(value)
        for value in re.findall(r"(?<!\d)(20\d{2}|25\d{2})(?!\d)", text)
    ]
    if not years:
        return None
    year = years[-1]
    return year - 543 if year > 2400 else year


def _meta_dominant_period(lines: list[OcrLine]) -> tuple[int | None, int | None]:
    months: list[int] = []
    years: list[int] = []
    for line in lines:
        month = _meta_month(line.text)
        year = _meta_year(line.text)
        if month:
            months.append(month)
        if year and 2020 <= year <= date.today().year + 2:
            years.append(year)
    month = Counter(months).most_common(1)[0][0] if months else None
    year = Counter(years).most_common(1)[0][0] if years else None
    return month, year


def _meta_row_date(
    text: str,
    *,
    fallback_month: int | None,
    fallback_year: int | None,
) -> str | None:
    day_match = re.search(r"(?<!\d)([0-3]?\d)(?!\d)", text)
    if not day_match:
        return None
    day = int(day_match.group(1))
    month = _meta_month(text) or fallback_month
    year = _meta_year(text) or fallback_year
    if not month or not year:
        return None
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def _meta_amount(text: str) -> tuple[float | None, bool]:
    """Return an amount and whether a leading OCR baht artifact was removed."""
    compact = re.sub(r"\s+", "", text)
    numeric = re.sub(r"[^\d,.]", "", compact)
    corrected_baht = False
    # On Meta's font, Tesseract sees the leading ฿ glyph as an extra 8 or 6:
    # ฿64.50 -> 864.50 and ฿5,000.00 -> 65,000.00.
    if compact[:1].isdigit() and numeric.startswith(("8", "6")) and re.fullmatch(
        r"\d{1,3}(?:,\d{3})*\.\d{2}|\d+\.\d{2}", numeric[1:]
    ):
        numeric = numeric[1:]
        corrected_baht = True
    return parse_amount(numeric), corrected_baht


def _meta_filename_card(filename: str) -> str | None:
    values = re.findall(r"(?<!\d)(\d{4})(?!\d)", Path(filename).stem)
    return next(
        (value for value in reversed(values) if not value.startswith("20")),
        None,
    )


def _meta_line_words(lines: list[OcrLine]) -> list[PositionedWord]:
    selected: dict[tuple[int, int, str], PositionedWord] = {}
    for line in lines:
        for word in line.words:
            key = (
                round(word.left / 2),
                round(word.top / 2),
                _meta_compact_text(word.text),
            )
            current = selected.get(key)
            if current is None or word.confidence > current.confidence:
                selected[key] = word
    return sorted(selected.values(), key=lambda word: (word.top, word.left))


def _meta_reference(
    row_lines: list[OcrLine], left: float, right: float
) -> str | None:
    candidates: list[tuple[int, float, str]] = []
    for line in row_lines:
        relevant = [
            word for word in line.words
            if left <= word.left + word.width / 2 <= right
        ]
        if not relevant:
            continue
        groups: list[list[PositionedWord]] = []
        centers: list[float] = []
        for word in sorted(
            relevant,
            key=lambda item: (item.top + item.height / 2, item.left),
        ):
            center = word.top + word.height / 2
            if not groups or abs(center - centers[-1]) > 6:
                groups.append([word])
                centers.append(center)
            else:
                groups[-1].append(word)
        for group in groups:
            joined = "".join(
                word.text for word in sorted(group, key=lambda word: word.left)
            )
            for raw in re.findall(r"[A-Za-z0-9?]{8,12}", joined):
                clean = raw.upper()
                if not re.search(r"[A-Z]", clean) or not re.search(r"\d", clean):
                    continue
                if (
                    clean.startswith(("MASTERCARD", "AMERICAN", "EXPRESS"))
                    or "CARD" in clean
                    or "CEARD" in clean
                ):
                    continue
                candidates.append(
                    (
                        clean.count("?"),
                        -sum(word.confidence for word in group),
                        clean,
                    )
                )
    return min(candidates)[2] if candidates else None


def _parse_meta_payment_ocr(
    lines: list[OcrLine],
    filename: str,
    image_width: int,
    image_height: int,
) -> list[ParsedTransaction]:
    """Parse Meta payment-history screenshots by their stable table columns."""
    words = _meta_line_words(lines)
    amount_words: list[PositionedWord] = []
    for word in words:
        center_x = word.left + word.width / 2
        if not (image_width * 0.25 <= center_x <= image_width * 0.43):
            continue
        if word.top < image_height * 0.30:
            continue
        amount, _ = _meta_amount(word.text)
        if amount is not None and re.search(r"[.,]\d{2}\D*$", word.text):
            amount_words.append(word)

    anchors: list[PositionedWord] = []
    for word in sorted(amount_words, key=lambda item: item.top):
        if anchors and abs(word.top - anchors[-1].top) <= 4:
            if word.confidence > anchors[-1].confidence:
                anchors[-1] = word
        else:
            anchors.append(word)
    if not anchors:
        return []

    gaps = [anchors[index + 1].top - anchors[index].top for index in range(len(anchors) - 1)]
    row_height = sorted(gaps)[len(gaps) // 2] if gaps else max(50.0, image_height * 0.10)
    fallback_month, fallback_year = _meta_dominant_period(lines)
    filename_card = _meta_filename_card(filename)
    transactions: list[ParsedTransaction] = []

    for index, anchor in enumerate(anchors):
        top = (anchors[index - 1].top + anchor.top) / 2 if index else anchor.top - row_height / 2
        bottom = (anchor.top + anchors[index + 1].top) / 2 if index + 1 < len(anchors) else anchor.top + row_height / 2
        row_words = [
            word for word in words
            if top <= word.top + word.height / 2 < bottom
        ]
        row_lines = [
            line for line in lines
            if any(top <= word.top + word.height / 2 < bottom for word in line.words)
        ]

        date_text = " ".join(
            word.text for word in sorted(row_words, key=lambda item: (item.top, item.left))
            if image_width * 0.13 <= word.left + word.width / 2 <= image_width * 0.31
        )
        transaction_date = _meta_row_date(
            date_text,
            fallback_month=fallback_month,
            fallback_year=fallback_year,
        )
        amount, _ = _meta_amount(anchor.text)
        if amount is None:
            continue

        payment_text = " ".join(
            word.text for word in sorted(row_words, key=lambda item: (item.top, item.left))
            if image_width * 0.42 <= word.left + word.width / 2 <= image_width * 0.66
        )
        card_match = re.search(
            r"(?:master\s*card|american\s*express|visa|บัตร).*?(\d{4})(?!\d)",
            payment_text,
            re.IGNORECASE,
        )
        card_last4 = filename_card or (card_match.group(1) if card_match else None)
        reference = _meta_reference(row_lines, image_width * 0.42, image_width * 0.66)

        left_text = " ".join(
            word.text for word in sorted(row_words, key=lambda item: (item.top, item.left))
            if word.left + word.width / 2 < image_width * 0.16
        )
        id_parts = re.findall(r"\d{12,}", left_text)
        transaction_id = "-".join(id_parts[:2]) if id_parts else None

        invoice_text = " ".join(
            word.text for word in sorted(row_words, key=lambda item: (item.top, item.left))
            if image_width * 0.70 <= word.left + word.width / 2 <= image_width * 0.94
        )
        invoice_match = re.search(r"FBADS[-\s]\d{3}[-\s]\d{6,}", invoice_text, re.IGNORECASE)
        invoice = re.sub(r"\s", "-", invoice_match.group(0)).upper() if invoice_match else None

        warnings: list[str] = []
        if not transaction_date:
            warnings.append("อ่านวันที่ไม่ชัด กรุณาตรวจจากรูป")
        if not reference:
            warnings.append("อ่านเลขอ้างอิงไม่ชัด กรุณาตรวจจากรูป")
        elif "?" in reference:
            warnings.append("เลขอ้างอิงมีตัวอักษรที่อ่านไม่ชัด กรุณาตรวจจากรูป")
        if not card_last4:
            warnings.append("อ่านเลขท้ายบัตรไม่ชัด กรุณาตรวจจากรูป")

        description_parts = ["Facebook/Meta Ads"]
        if transaction_id:
            description_parts.append(f"รายการ {transaction_id}")
        if invoice:
            description_parts.append(f"ใบกำกับ {invoice}")
        confidence_values = [anchor.confidence]
        confidence_values.extend(word.confidence for word in row_words)
        confidence = sum(confidence_values) / max(1, len(confidence_values))
        if confidence < 80:
            warnings.insert(0, "ข้อความตัวเล็กบางส่วนอาจอ่านคลาดเคลื่อน กรุณาเทียบกับรูป")
        row_hash = make_row_hash(
            transaction_date or "",
            None,
            amount,
            " | ".join(description_parts),
            filename,
            reference,
        )
        transactions.append(
            ParsedTransaction(
                transaction_date=transaction_date,
                description=" | ".join(description_parts),
                amount=round(abs(amount), 2),
                card_last4=card_last4,
                category="ค่าโฆษณาออนไลน์",
                deposit_amount=round(abs(amount), 2),
                channel=Path(filename).name[:100],
                tr_code=reference,
                row_hash=row_hash,
                confidence=round(max(0.0, min(100.0, confidence)), 2),
                warnings=warnings,
            )
        )
    return _deduplicate_transactions(transactions)


def _ads_row_texts(lines: list[OcrLine]) -> list[tuple[str, float]]:
    words = [word for line in lines for word in line.words if word.text.strip()]
    if not words:
        return [(line.text, line.confidence) for line in lines if line.text.strip()]

    heights = sorted(max(1.0, word.height) for word in words)
    median_height = heights[len(heights) // 2]
    tolerance = max(10.0, median_height * 0.8)
    grouped: list[list[PositionedWord]] = []
    centers: list[float] = []
    for word in sorted(words, key=lambda item: (item.top + item.height / 2, item.left)):
        center = word.top + word.height / 2
        if not grouped or abs(center - centers[-1]) > tolerance:
            grouped.append([word])
            centers.append(center)
        else:
            grouped[-1].append(word)
            centers[-1] = sum(item.top + item.height / 2 for item in grouped[-1]) / len(grouped[-1])

    result: list[tuple[str, float]] = []
    for group in grouped:
        group.sort(key=lambda item: item.left)
        text = " ".join(item.text for item in group).strip()
        weight = sum(max(1, len(item.text)) for item in group)
        confidence = sum(item.confidence * max(1, len(item.text)) for item in group) / max(1, weight)
        if text:
            result.append((text, confidence))
    return result


def _labelled_code(text: str, labels: str) -> str | None:
    match = re.search(
        rf"(?:{labels})\s*[:#-]?\s*([A-Z0-9][A-Z0-9|/_-]{{5,}})",
        text,
        re.IGNORECASE,
    )
    return match.group(1).strip("|/_-") if match else None


def _parse_ads_ocr_lines(lines: list[OcrLine], filename: str) -> list[ParsedTransaction]:
    transactions: list[ParsedTransaction] = []
    all_text = "\n".join(line.text for line in lines)
    platform = _ads_platform(all_text)
    amount_pattern = re.compile(
        r"(?<![\d:/-])(?:฿\s*)?(\d{1,3}(?:,\d{3})*\.\d{2}|\d+\.\d{2})(?:\s*(?:THB|บาท))?(?!\d)",
        re.IGNORECASE,
    )

    row_texts = _ads_row_texts(lines)
    for row_index, (text, confidence) in enumerate(row_texts):
        clean = re.sub(r"\s+", " ", text).strip()
        lowered = clean.casefold()
        if any(label in lowered for label in ("ยอดรวม", "total amount", "available balance", "current balance")):
            continue
        amount_matches = list(amount_pattern.finditer(clean))
        if not amount_matches:
            continue
        amount = parse_amount(amount_matches[-1].group(1))
        if amount is None or amount <= 0:
            continue

        context_parts = [clean]
        for prior_index in range(row_index - 1, max(-1, row_index - 4), -1):
            prior = re.sub(r"\s+", " ", row_texts[prior_index][0]).strip()
            if amount_pattern.search(prior):
                break
            context_parts.insert(0, prior)
        context = " ".join(context_parts)

        date_match = re.search(
            r"\b(?:\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\b",
            context,
        )
        time_match = re.search(r"\b\d{1,2}:\d{2}(?::\d{2})?\b", context)
        transaction_date = parse_date(date_match.group(0)) if date_match else None
        transaction_time = parse_time(time_match.group(0)) if time_match else None

        card_match = re.search(
            r"(?:•{2,}|\*{2,}|x{2,}|ending\s*(?:in)?|ลงท้าย)\s*[- ]?(\d{4})\b",
            context,
            re.IGNORECASE,
        )
        if not card_match:
            card_match = re.search(
                r"(?:visa|mastercard|master card|amex|card|บัตร)[^\d]{0,24}(\d{4})\b",
                context,
                re.IGNORECASE,
            )
        card_last4 = card_match.group(1) if card_match else None

        invoice = _labelled_code(context, r"invoice(?:\s*(?:no|number))?|ใบแจ้งหนี้|เลขที่เอกสาร")
        if not invoice:
            invoice_match = re.search(r"\b(?:THTT|TH|INV)[A-Z0-9_-]{6,}\b", context, re.IGNORECASE)
            invoice = invoice_match.group(0) if invoice_match else None
        transaction_id = _labelled_code(context, r"transaction(?:\s*id)?|รหัสธุรกรรม|รหัสรายการ")
        payment_reference = _labelled_code(context, r"reference|ref|รหัสอ้างอิง")

        code_candidates = re.findall(r"\b(?=[A-Z0-9_-]{7,}\b)(?=[A-Z0-9_-]*[A-Z])(?=[A-Z0-9_-]*\d)[A-Z0-9_-]+\b", context.upper())
        excluded = {value.upper() for value in (invoice, transaction_id) if value}
        if not payment_reference:
            payment_reference = next((value for value in code_candidates if value not in excluded), None)

        description_parts = [value for value in (platform, invoice, transaction_id) if value]
        description = " | ".join(description_parts)
        if len(description_parts) == 1:
            without_amount = re.sub(amount_pattern, "", context).strip(" |-")
            if without_amount:
                description = f"{platform} | {without_amount[:220]}"

        warnings: list[str] = []
        if confidence < 70:
            warnings.append("ข้อความบางส่วนไม่ชัด กรุณาเทียบกับรูป")
        if not transaction_date:
            warnings.append("ไม่พบวันที่ กรุณาตรวจหรือกรอกเพิ่ม")
        if not payment_reference:
            warnings.append("ไม่พบเลขอ้างอิง กรุณาตรวจหรือกรอกเพิ่ม")
        if not card_last4:
            warnings.append("ไม่พบเลขท้ายบัตร กรุณาตรวจหรือกรอกเพิ่ม")

        rounded_amount = round(abs(float(amount)), 2)
        row_hash = make_row_hash(
            transaction_date or "",
            transaction_time,
            rounded_amount,
            description,
            filename,
            payment_reference,
        )
        transactions.append(
            ParsedTransaction(
                transaction_date=transaction_date,
                transaction_time=transaction_time,
                description=description[:500],
                amount=rounded_amount,
                card_last4=card_last4,
                category="ค่าโฆษณาออนไลน์",
                deposit_amount=rounded_amount,
                withdraw_amount=None,
                channel=Path(filename).name[:100],
                tr_code=payment_reference or transaction_id or invoice,
                row_hash=row_hash,
                confidence=round(max(0.0, min(100.0, confidence)), 2),
                warnings=warnings,
            )
        )
    return _deduplicate_transactions(transactions)


def _ocr_ads_image(filename: str, data: bytes) -> tuple[list[ParsedTransaction], str]:
    from PIL import Image, ImageEnhance, ImageFilter

    if not shutil.which("tesseract"):
        raise ValueError("เซิร์ฟเวอร์ยังไม่ได้ติดตั้ง Tesseract OCR")
    with tempfile.TemporaryDirectory(prefix="ads-evidence-ocr-") as temp_dir:
        source_path = Path(temp_dir) / "source.png"
        source_path.write_bytes(data)
        environment = {**os.environ, "OMP_THREAD_LIMIT": "2"}
        with Image.open(source_path) as source:
            image_width, image_height = source.size
            variants = _ocr_image_variants(source)
            grayscale_path = Path(temp_dir) / "grayscale.png"
            variants["grayscale"].save(grayscale_path)
            binary_path = Path(temp_dir) / "binary.png"
            variants["binary"].save(binary_path)

        # Sparse layout mode preserves Meta's separated table columns much
        # better than the generic denoised page pass.
        sparse_source_lines = _run_tesseract_tsv(
            source_path,
            1,
            psm=11,
            environment=environment,
            error_label="อ่านข้อความจากรูป ",
            timeout=max(10, min(25, EVIDENCE_OCR_TIMEOUT_SECONDS // 2)),
        )
        sparse_text = "\n".join(line.text for line in sparse_source_lines)
        looks_like_meta = (
            "fbads" in sparse_text.casefold()
            or (
                ("mastercard" in sparse_text.casefold() or "american express" in sparse_text.casefold())
                and ("บัญชีโฆษณา" in _normalise_ocr_line(sparse_text) or "ธุรกรรม" in sparse_text)
            )
        )
        if looks_like_meta:
            with Image.open(source_path) as source:
                crop_left = int(source.width * 0.42)
                crop_top = int(source.height * 0.30)
                crop_right = int(source.width * 0.66)
                payment_crop = source.crop((crop_left, crop_top, crop_right, source.height))
                scale = 3.0
                payment_crop = payment_crop.resize(
                    (round(payment_crop.width * scale), round(payment_crop.height * scale)),
                    Image.Resampling.LANCZOS,
                )
                payment_crop = ImageEnhance.Contrast(payment_crop).enhance(1.2)
                payment_crop = payment_crop.filter(ImageFilter.SHARPEN)
                payment_path = Path(temp_dir) / "meta-payment-column.png"
                payment_crop.save(payment_path, format="PNG", optimize=True)
            payment_lines = _run_tesseract_tsv(
                payment_path,
                1,
                psm=6,
                languages="eng",
                environment=environment,
                error_label="อ่านคอลัมน์บัตรและเลขอ้างอิง ",
                timeout=max(10, min(25, EVIDENCE_OCR_TIMEOUT_SECONDS // 2)),
            )
            mapped_payment_lines = _map_scaled_ocr_lines(
                payment_lines,
                crop_left=crop_left,
                crop_top=crop_top,
                scale=scale,
            )
            meta_lines = [*sparse_source_lines, *mapped_payment_lines]
            transactions = _parse_meta_payment_ocr(
                meta_lines,
                filename,
                image_width,
                image_height,
            )
            return transactions, "\n".join(line.text for line in meta_lines)

        lines = _run_tesseract_tsv(
            grayscale_path,
            1,
            psm=6,
            environment=environment,
            error_label="อ่านข้อความจากรูป ",
            timeout=max(10, min(25, EVIDENCE_OCR_TIMEOUT_SECONDS // 2)),
        )
        transactions = _parse_ads_ocr_lines(lines, filename)
        if not transactions:
            fallback_lines = _run_tesseract_tsv(
                binary_path,
                1,
                psm=11,
                environment=environment,
                error_label="อ่านข้อความแบบแยกส่วน ",
                timeout=max(10, min(25, EVIDENCE_OCR_TIMEOUT_SECONDS // 2)),
            )
            lines = _deduplicate_ocr_lines([*lines, *sparse_source_lines, *fallback_lines])
            transactions = _parse_ads_ocr_lines(lines, filename)
        return transactions, "\n".join(line.text for line in lines)


async def parse_images_with_local_ocr(
    image_items: list[tuple[str, bytes]],
) -> ParsedStatement:
    """Read evidence sequentially with local Thai/English Tesseract OCR."""
    import asyncio as _asyncio
    import logging as _logging

    transactions: list[ParsedTransaction] = []
    warnings: list[str] = []
    detected_text: list[str] = []
    for filename, data in image_items:
        try:
            rows, text = await _asyncio.wait_for(
                _asyncio.to_thread(_ocr_ads_image, filename, data),
                timeout=EVIDENCE_OCR_TIMEOUT_SECONDS,
            )
            transactions.extend(rows)
            detected_text.append(text)
            if not rows:
                warnings.append(f"{filename}: ยังแยกรายการไม่ได้ กรุณากรอกในหน้าตรวจสอบ")
        except TimeoutError:
            warnings.append(f"{filename}: ใช้เวลาอ่านเกินกำหนด กรุณากรอกในหน้าตรวจสอบ")
        except ValueError as exc:
            _logging.getLogger(__name__).warning("[OCR] %s failed: %s", filename, exc)
            warnings.append(f"{filename}: {exc}")

    transactions = _deduplicate_transactions(transactions)
    start, end = _statement_dates(transactions)
    issuer = _ads_platform("\n".join(detected_text))
    return ParsedStatement(
        issuer=issuer,
        statement_type="ads_screenshot",
        extraction_method="tesseract_local",
        masked_reference=None,
        statement_date_from=start,
        statement_date_to=end,
        summary_totals={
            "total_charges": round(sum(item.amount or 0 for item in transactions), 2)
        },
        transactions=transactions,
        warnings=warnings,
    )
